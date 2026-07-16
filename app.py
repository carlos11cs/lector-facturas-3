import calendar
import csv
import hashlib
import io
import json
import logging
import multiprocessing as mp
import os
import re
import secrets
import tempfile
import zipfile
from datetime import date, datetime, timedelta
from functools import wraps
from xml.etree import ElementTree as ET

import httpx
import fitz
import openpyxl
try:
    import stripe
except ImportError:  # pragma: no cover - dependency is required in deployment
    stripe = None
from flask import Flask, jsonify, redirect, render_template, request, session, url_for, g, send_file
from sqlalchemy import (
    Boolean,
    Column,
    Float,
    Integer,
    MetaData,
    String,
    Text,
    Table,
    create_engine,
    func,
    inspect,
    select,
    text,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from services.ai_invoice_service import (
    _extract_first_date,
    _find_payment_date_by_keywords,
    _find_payment_dates_by_keywords,
    analyze_invoice,
    _extract_pdf_text_from_bytes,
    _extract_pdf_text_ocr_from_bytes,
    _extract_image_text_ocr_from_bytes,
    extract_loan_schedule,
)
from services.storage_service import get_public_url, upload_bytes

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "data.db")
ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png"}
DOCUMENT_CENTER_ALLOWED_EXTENSIONS = {
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".zip",
    ".csv",
    ".xlsx",
    ".xls",
}
ANALYSIS_TIMEOUT_SECONDS = int(os.getenv("ANALYSIS_TIMEOUT_SECONDS", "120"))
DEFAULT_USER_ID = int(os.getenv("DEFAULT_USER_ID", "1"))
OWNER_EMAIL = (os.getenv("OWNER_EMAIL") or "").strip().lower()
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "")
APP_FROM_EMAIL = os.getenv("APP_FROM_EMAIL", "no-reply@tuapp.com")
ACCESS_FROM_EMAIL = os.getenv("ACCESS_FROM_EMAIL", "").strip()
REPORTS_FROM_EMAIL = os.getenv("REPORTS_FROM_EMAIL", "").strip()
APP_BASE_URL = (os.getenv("APP_BASE_URL") or "").strip().rstrip("/")
STRIPE_SECRET_KEY = (os.getenv("STRIPE_SECRET_KEY") or "").strip()
STRIPE_WEBHOOK_SECRET = (os.getenv("STRIPE_WEBHOOK_SECRET") or "").strip()
STRIPE_PRICE_STARTER = (os.getenv("STRIPE_PRICE_STARTER") or "").strip()
STRIPE_PRICE_PRO = (os.getenv("STRIPE_PRICE_PRO") or "").strip()
STRIPE_PRICE_ADVANCED = (os.getenv("STRIPE_PRICE_ADVANCED") or "").strip()
STRIPE_PRICE_STARTER_ANNUAL = (os.getenv("STRIPE_PRICE_STARTER_ANNUAL") or "").strip()
STRIPE_PRICE_PRO_ANNUAL = (os.getenv("STRIPE_PRICE_PRO_ANNUAL") or "").strip()
STRIPE_PRICE_ADVANCED_ANNUAL = (os.getenv("STRIPE_PRICE_ADVANCED_ANNUAL") or "").strip()
STRIPE_PROMO_COUPON_3M_50 = (os.getenv("STRIPE_PROMO_COUPON_3M_50") or "").strip()
RENT_WITHHOLDING_TYPES = {"alquiler_local", "alquiler_cabina"}
EXCLUDED_111_TYPES = {"prestamo", "seguridad_social", "amortizacion", "kilometraje"}
ALLOWED_NO_INVOICE_EXPENSE_TYPES = {
    "nomina",
    "seguridad_social",
    "alquiler_local",
    "alquiler_cabina",
    "amortizacion",
    "kilometraje",
    "prestamo",
    "otro",
}
MONTH_LABELS_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
STRIPE_PRICE_IDS = {
    "monthly": {
        "starter": STRIPE_PRICE_STARTER,
        "pro": STRIPE_PRICE_PRO,
        "advanced": STRIPE_PRICE_ADVANCED,
    },
    "annual": {
        "starter": STRIPE_PRICE_STARTER_ANNUAL,
        "pro": STRIPE_PRICE_PRO_ANNUAL,
        "advanced": STRIPE_PRICE_ADVANCED_ANNUAL,
    },
}
STRIPE_PRICE_TO_PLAN = {
    price_id: {"plan": plan, "billing_period": billing_period}
    for billing_period, period_prices in STRIPE_PRICE_IDS.items()
    for plan, price_id in period_prices.items()
    if price_id
}

PLAN_LIMITS = {
    "starter": {"companies": 5, "staff": 2},
    "pro": {"companies": 25, "staff": 8},
    "advanced": {"companies": 75, "staff": 20},
}

_raw_db_url = os.getenv("DATABASE_URL")
DATABASE_URL = _raw_db_url.strip() if _raw_db_url else ""
if not DATABASE_URL:
    DATABASE_URL = f"sqlite:///{DB_PATH}"
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if DATABASE_URL.startswith("sqlite"):
    logging.warning("DATABASE_URL no configurada. Usando SQLite local.")

engine = create_engine(DATABASE_URL, pool_pre_ping=True, future=True)
metadata = MetaData()

companies_table = Table(
    "companies",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False),
    Column("agency_id", Integer, nullable=True),
    Column("display_name", String, nullable=False),
    Column("legal_name", String, nullable=False),
    Column("tax_id", String, nullable=False),
    Column("company_type", String, nullable=False),  # individual | company
    Column("email", String),
    Column("phone", String),
    Column("assigned_user_id", Integer),
    Column("vat_regime", String, nullable=False, server_default="general"),
    Column("tax_periodicity", String, nullable=False, server_default="quarterly"),
    Column("files_model_303", Boolean, nullable=False, server_default=text("true")),
    Column("files_model_111", Boolean, nullable=False, server_default=text("false")),
    Column("files_model_115", Boolean, nullable=False, server_default=text("false")),
    Column("files_model_130", Boolean, nullable=False, server_default=text("false")),
    Column("files_model_202", Boolean, nullable=False, server_default=text("false")),
    Column("balance_manual_data", Text),
    Column("created_at", String, nullable=False),
)

users_table = Table(
    "users",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("email", String, nullable=False, unique=True),
    Column("password_hash", String, nullable=False),
    Column("role", String, nullable=False),  # owner | agency | staff
    Column("plan", String, nullable=False),  # internal | trial | standard | premium
    Column("agency_id", Integer),
    Column("created_at", String, nullable=False),
    Column("is_active", Boolean, nullable=False, server_default=text("true")),
)

agencies_table = Table(
    "agencies",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("name", String, nullable=False),
    Column("email", String, nullable=False),
    Column("phone", String),
    Column("plan", String, nullable=False),  # starter | pro | advanced
    Column("status", String, nullable=False),  # trial | active | suspended
    Column("stripe_customer_id", String),
    Column("stripe_subscription_id", String),
    Column("stripe_price_id", String),
    Column("stripe_subscription_status", String),
    Column("stripe_current_period_end", String),
    Column("trial_ends_at", String),
    Column("created_at", String, nullable=False),
    Column("last_login_at", String),
)

password_resets_table = Table(
    "password_resets",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False),
    Column("token", String, nullable=False),
    Column("expires_at", String, nullable=False),
    Column("used_at", String),
)

user_invitations_table = Table(
    "user_invitations",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("email", String, nullable=False),
    Column("role", String, nullable=False),  # agency | staff
    Column("agency_id", Integer, nullable=True),
    Column("name", String),
    Column("plan", String),
    Column("token", String, nullable=False),
    Column("expires_at", String, nullable=False),
    Column("accepted_at", String),
    Column("created_by_user_id", Integer, nullable=False),
    Column("created_at", String, nullable=False),
)

invoices_table = Table(
    "invoices",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False, server_default=str(DEFAULT_USER_ID)),
    Column("company_id", Integer, nullable=False),
    Column("original_filename", String, nullable=False),
    Column("stored_filename", String),
    Column("invoice_date", String, nullable=False),
    Column("supplier", String, nullable=False),
    Column("base_amount", Float, nullable=False),
    Column("vat_deductible", Boolean),
    Column("vat_rate", Integer),
    Column("vat_amount", Float),
    Column("total_amount", Float, nullable=False),
    Column("vat_breakdown", Text),
    Column("withholding_amount", Float),
    Column("payment_date", String),
    Column("payment_dates", Text),
    Column("payment_completed_dates", Text),
    Column("ocr_text", Text),
    Column("extraction_source", String),
    Column("confidence_score", Float),
    Column("expense_category", String, nullable=False, server_default="with_invoice"),
    Column("expense_family", String),
    Column("expense_subtype", String),
    Column("pnl_bucket", String),
    Column("tax_model_targets", Text),
    Column("created_at", String, nullable=False),
)

income_invoices_table = Table(
    "income_invoices",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False, server_default=str(DEFAULT_USER_ID)),
    Column("company_id", Integer, nullable=False),
    Column("original_filename", String, nullable=False),
    Column("stored_filename", String),
    Column("invoice_date", String, nullable=False),
    Column("client", String, nullable=False),
    Column("base_amount", Float, nullable=False),
    Column("vat_rate", Integer),
    Column("vat_amount", Float),
    Column("total_amount", Float, nullable=False),
    Column("vat_breakdown", Text),
    Column("payment_date", String),
    Column("payment_dates", Text),
    Column("payment_completed_dates", Text),
    Column("ocr_text", Text),
    Column("extraction_source", String),
    Column("confidence_score", Float),
    Column("created_at", String, nullable=False),
)

known_suppliers_table = Table(
    "known_suppliers",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False, server_default=str(DEFAULT_USER_ID)),
    Column("company_id", Integer, nullable=True),
    Column("name", String, nullable=False),
    Column("tax_id", String),
    Column("confirmed_at", String, nullable=False),
)

facturacion_table = Table(
    "facturacion",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False, server_default=str(DEFAULT_USER_ID)),
    Column("company_id", Integer, nullable=False),
    Column("mes", Integer, nullable=False),
    Column("anio", Integer, nullable=False),
    Column("invoice_date", String),
    Column("concept", String),
    Column("base_facturada", Float, nullable=False),
    Column("tipo_iva", Integer, nullable=False),
    Column("iva_repercutido", Float, nullable=False),
    Column("total_amount", Float),
)

no_invoice_table = Table(
    "no_invoice_expenses",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False, server_default=str(DEFAULT_USER_ID)),
    Column("company_id", Integer, nullable=False),
    Column("expense_date", String, nullable=False),
    Column("payment_date", String),
    Column("payment_dates", Text),
    Column("payment_completed_dates", Text),
    Column("concept", String, nullable=False),
    Column("amount", Float, nullable=False),
    Column("interest_amount", Float),
    Column("vat_deductible", Boolean),
    Column("vat_rate", Integer),
    Column("vat_amount", Float),
    Column("base_amount", Float),
    Column("withholding_amount", Float),
    Column("payroll_employee_name", String),
    Column("payroll_period", String),
    Column("payroll_net_amount", Float),
    Column("payroll_total_deductions_amount", Float),
    Column("payroll_employer_cost_amount", Float),
    Column("expense_type", String, nullable=False),
    Column("deductible", Boolean, nullable=False),
    Column("expense_family", String),
    Column("expense_subtype", String),
    Column("pnl_bucket", String),
    Column("tax_model_targets", Text),
    Column("created_at", String, nullable=False),
)

loan_installments_table = Table(
    "loan_installments",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("user_id", Integer, nullable=False, server_default=str(DEFAULT_USER_ID)),
    Column("company_id", Integer, nullable=False),
    Column("bank_name", String),
    Column("concept", String, nullable=False),
    Column("payment_date", String, nullable=False),
    Column("payment_completed_dates", Text),
    Column("total_amount", Float, nullable=False),
    Column("interest_amount", Float, nullable=False),
    Column("principal_amount", Float, nullable=False),
    Column("created_at", String, nullable=False),
)

document_batches_table = Table(
    "document_batches",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("company_id", Integer, nullable=False),
    Column("uploaded_by_user_id", Integer, nullable=False),
    Column("period", String, nullable=False),
    Column("total_documents", Integer, nullable=False, server_default=text("0")),
    Column("processed_documents", Integer, nullable=False, server_default=text("0")),
    Column("ready_documents", Integer, nullable=False, server_default=text("0")),
    Column("review_documents", Integer, nullable=False, server_default=text("0")),
    Column("duplicate_documents", Integer, nullable=False, server_default=text("0")),
    Column("failed_documents", Integer, nullable=False, server_default=text("0")),
    Column("status", String, nullable=False, server_default="uploaded"),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
)

processed_documents_table = Table(
    "processed_documents",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("company_id", Integer, nullable=False),
    Column("uploaded_by_user_id", Integer, nullable=False),
    Column("source_batch_id", Integer, nullable=False),
    Column("original_filename", String, nullable=False),
    Column("storage_path", String, nullable=False),
    Column("file_url", String),
    Column("file_type", String),
    Column("uploaded_at", String, nullable=False),
    Column("processing_status", String, nullable=False, server_default="uploaded"),
    Column("detected_document_type", String, nullable=False, server_default="unknown"),
    Column("confidence_score", Float),
    Column("extracted_data_json", Text),
    Column("original_extracted_data_json", Text),
    Column("corrected_data_json", Text),
    Column("validation_status", String, nullable=False, server_default="needs_review"),
    Column("issue_type", String),
    Column("issue_description", Text),
    Column("linked_accounting_record_id", String),
    Column("linked_accounting_record_type", String),
    Column("approved_at", String),
    Column("approved_by_user_id", Integer),
    Column("rejected_at", String),
    Column("rejected_by_user_id", Integer),
    Column("registered_at", String),
    Column("registered_by_user_id", Integer),
    Column("period", String, nullable=False),
    Column("extracted_text", Text),
    Column("duplicate_of_document_id", Integer),
    Column("content_hash", String),
    Column("audit_log_json", Text),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
)

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY") or secrets.token_urlsafe(32)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.getenv("ENV", "").lower() == "production"
if stripe and STRIPE_SECRET_KEY:
    stripe.api_key = STRIPE_SECRET_KEY


def static_asset_url(filename: str) -> str:
    static_path = os.path.join(app.static_folder or "", filename)
    version = None
    try:
        version = int(os.path.getmtime(static_path))
    except OSError:
        version = None
    if version is None:
        return url_for("static", filename=filename)
    return url_for("static", filename=filename, v=version)


@app.context_processor
def inject_static_asset_url():
    return {"static_asset_url": static_asset_url}


def get_app_base_url():
    if APP_BASE_URL:
        return APP_BASE_URL
    return request.url_root.rstrip("/")


def stripe_is_configured():
    return bool(
        stripe
        and
        STRIPE_SECRET_KEY
        and STRIPE_WEBHOOK_SECRET
        and STRIPE_PRICE_STARTER
        and STRIPE_PRICE_PRO
        and STRIPE_PRICE_ADVANCED
    )


def get_plan_label(plan):
    return {
        "starter": "Starter",
        "pro": "Pro",
        "advanced": "Business",
    }.get((plan or "").strip().lower(), (plan or "-").title())


def get_role_label(role):
    return {
        "agency": "gestoría",
        "staff": "trabajador",
    }.get((role or "").strip().lower(), role or "usuario")


def get_access_from_email():
    return ACCESS_FROM_EMAIL or APP_FROM_EMAIL


def get_reports_from_email():
    return REPORTS_FROM_EMAIL or APP_FROM_EMAIL


def annual_billing_is_configured():
    return bool(
        STRIPE_PRICE_STARTER_ANNUAL
        and STRIPE_PRICE_PRO_ANNUAL
        and STRIPE_PRICE_ADVANCED_ANNUAL
    )


def get_billing_period_label(billing_period):
    return {
        "monthly": "Mensual",
        "annual": "Anual",
    }.get((billing_period or "").strip().lower(), "Mensual")


def get_stripe_price_id(plan, billing_period="monthly"):
    normalized_period = (billing_period or "monthly").strip().lower()
    normalized_plan = (plan or "").strip().lower()
    return STRIPE_PRICE_IDS.get(normalized_period, {}).get(normalized_plan, "")


def get_plan_from_price_id(price_id):
    return STRIPE_PRICE_TO_PLAN.get((price_id or "").strip(), {})


def get_plan_limits(plan):
    return PLAN_LIMITS.get((plan or "").strip().lower(), PLAN_LIMITS["starter"])


def get_default_stripe_discounts():
    if STRIPE_PROMO_COUPON_3M_50:
        return [{"coupon": STRIPE_PROMO_COUPON_3M_50}]
    return []


def count_pending_staff_invitations(conn, agency_id):
    if not agency_id:
        return 0
    now = datetime.utcnow().isoformat()
    return conn.execute(
        select(func.count())
        .select_from(user_invitations_table)
        .where(user_invitations_table.c.agency_id == agency_id)
        .where(user_invitations_table.c.role == "staff")
        .where(user_invitations_table.c.accepted_at.is_(None))
        .where(user_invitations_table.c.expires_at >= now)
    ).scalar_one()


def get_agency_usage(conn, agency_id):
    company_count = conn.execute(
        select(func.count())
        .select_from(companies_table)
        .where(companies_table.c.agency_id == agency_id)
    ).scalar_one()
    staff_count = conn.execute(
        select(func.count())
        .select_from(users_table)
        .where(users_table.c.agency_id == agency_id)
        .where(users_table.c.role == "staff")
        .where(users_table.c.is_active.is_(True))
    ).scalar_one()
    pending_staff_invitations = count_pending_staff_invitations(conn, agency_id)
    return {
        "companies": int(company_count or 0),
        "staff": int(staff_count or 0),
        "pending_staff_invitations": int(pending_staff_invitations or 0),
    }


def get_agency_limits_and_usage(conn, agency_id, plan=None):
    resolved_plan = (plan or "").strip().lower()
    if not resolved_plan:
        resolved_plan = (
            conn.execute(select(agencies_table.c.plan).where(agencies_table.c.id == agency_id))
            .scalar_one_or_none()
            or "starter"
        )
    limits = get_plan_limits(resolved_plan)
    usage = get_agency_usage(conn, agency_id)
    return {
        "plan": resolved_plan,
        "limits": limits,
        "usage": usage,
    }


def get_limit_error(resource, limit):
    resource_label = "empresas" if resource == "companies" else "usuarios staff"
    return (
        f"Has alcanzado el límite de {resource_label} de tu plan actual "
        f"({limit}). Sube de plan para ampliar capacidad."
    )


def get_agency_id_for_user(user):
    if not user:
        return None
    role = user.get("role")
    if role == "agency":
        return int(user.get("agency_id") or user["id"])
    if role == "staff":
        agency_id = user.get("agency_id")
        return int(agency_id) if agency_id else None
    return None


def get_agency_row_for_user(user):
    agency_id = get_agency_id_for_user(user)
    if not agency_id:
        return None
    with engine.connect() as conn:
        return conn.execute(
            select(agencies_table).where(agencies_table.c.id == agency_id)
        ).mappings().first()


def get_user_email(user_id):
    if not user_id:
        return None
    with engine.connect() as conn:
        return conn.execute(
            select(users_table.c.email).where(users_table.c.id == user_id)
        ).scalar_one_or_none()


def format_iso_date_label(value):
    if not value:
        return None
    try:
        if isinstance(value, (int, float)):
            return datetime.utcfromtimestamp(value).strftime("%d/%m/%Y")
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def map_stripe_subscription_status(subscription_status):
    status = (subscription_status or "").strip().lower()
    if status in {"trialing", "active", "past_due"}:
        return "active"
    if status in {"canceled", "unpaid", "paused", "incomplete_expired"}:
        return "suspended"
    if status == "incomplete":
        return "trial"
    return None


def sync_agency_billing_state(
    conn,
    agency_id,
    *,
    plan=None,
    app_status=None,
    stripe_customer_id=None,
    stripe_subscription_id=None,
    stripe_price_id=None,
    stripe_subscription_status=None,
    stripe_current_period_end=None,
):
    values = {}
    if plan:
        values["plan"] = plan
    if app_status:
        values["status"] = app_status
        if app_status == "active":
            values["trial_ends_at"] = None
    if stripe_customer_id is not None:
        values["stripe_customer_id"] = stripe_customer_id
    if stripe_subscription_id is not None:
        values["stripe_subscription_id"] = stripe_subscription_id
    if stripe_price_id is not None:
        values["stripe_price_id"] = stripe_price_id
    if stripe_subscription_status is not None:
        values["stripe_subscription_status"] = stripe_subscription_status
    if stripe_current_period_end is not None:
        values["stripe_current_period_end"] = stripe_current_period_end
    if values:
        conn.execute(
            agencies_table.update().where(agencies_table.c.id == agency_id).values(**values)
        )
    if app_status:
        conn.execute(
            users_table.update()
            .where(users_table.c.agency_id == agency_id)
            .values(is_active=app_status != "suspended")
        )


def clear_agency_stripe_state(conn, agency_id, *, reset_status=False):
    values = {
        "stripe_customer_id": None,
        "stripe_subscription_id": None,
        "stripe_price_id": None,
        "stripe_subscription_status": None,
        "stripe_current_period_end": None,
    }
    if reset_status:
        values["status"] = "trial"
        values["trial_ends_at"] = (datetime.utcnow() + timedelta(days=14)).isoformat()
    conn.execute(
        agencies_table.update().where(agencies_table.c.id == agency_id).values(**values)
    )


def stripe_customer_exists(customer_id):
    if not stripe or not customer_id:
        return False
    try:
        customer = stripe.Customer.retrieve(customer_id)
        return not getattr(customer, "deleted", False)
    except Exception:
        app.logger.warning("Stripe customer %s no disponible en el entorno actual", customer_id)
        return False


def get_billing_context_for_user(user):
    agency = get_agency_row_for_user(user)
    if not agency or user.get("role") != "agency":
        return None
    subscription_status = agency.get("stripe_subscription_status")
    status = agency.get("status") or "trial"
    with engine.connect() as conn:
        limits_and_usage = get_agency_limits_and_usage(
            conn, agency["id"], plan=agency.get("plan") or "starter"
        )
    return {
        "plan": agency.get("plan") or "starter",
        "plan_label": get_plan_label(agency.get("plan")),
        "status": status,
        "status_label": {
            "active": "Activa",
            "trial": "Trial",
            "suspended": "Suspendida",
        }.get(status, status.title()),
        "stripe_status_label": subscription_status.replace("_", " ").title()
        if subscription_status
        else None,
        "trial_ends_label": format_iso_date_label(agency.get("trial_ends_at")),
        "current_period_end_label": format_iso_date_label(
            agency.get("stripe_current_period_end")
        ),
        "has_customer": bool(agency.get("stripe_customer_id")),
        "has_subscription": bool(agency.get("stripe_subscription_id")),
        "can_manage": user.get("role") == "agency",
        "stripe_ready": stripe_is_configured(),
        "companies_used": limits_and_usage["usage"]["companies"],
        "companies_limit": limits_and_usage["limits"]["companies"],
        "staff_used": limits_and_usage["usage"]["staff"],
        "staff_limit": limits_and_usage["limits"]["staff"],
        "pending_staff_invitations": limits_and_usage["usage"][
            "pending_staff_invitations"
        ],
        "launch_promo_enabled": bool(STRIPE_PROMO_COUPON_3M_50),
        "annual_billing_available": annual_billing_is_configured(),
    }


def get_account_context_for_user(user):
    if not user:
        return None
    agency = get_agency_row_for_user(user)
    context = {
        "role": user.get("role"),
        "email": user.get("email") or "",
        "agency_name": "",
        "phone": "",
        "can_manage_subscription": user.get("role") == "agency",
    }
    if agency:
        context["agency_name"] = agency.get("name") or ""
        context["phone"] = agency.get("phone") or ""
    return context


def get_billing_message():
    billing_state = (request.args.get("billing") or "").strip().lower()
    billing_error = (request.args.get("billing_error") or "").strip().lower()
    if billing_state == "success":
        return {"type": "success", "text": "Suscripción activada. Stripe ha confirmado el alta."}
    if billing_state == "cancelled":
        return {"type": "warning", "text": "El proceso de pago se canceló antes de completarse."}
    if billing_error == "not_configured":
        return {
            "type": "warning",
            "text": "Stripe no está configurado todavía en producción. Faltan variables de entorno.",
        }
    if billing_error == "not_available":
        return {
            "type": "warning",
            "text": "No se pudo abrir la facturación porque esta gestoría no tiene cliente de Stripe asociado.",
        }
    if billing_error == "checkout_failed":
        return {
            "type": "warning",
            "text": "No se pudo iniciar el checkout de Stripe. Revisa la configuración de precios y claves.",
        }
    if billing_error == "annual_not_configured":
        return {
            "type": "warning",
            "text": "El cobro anual todavía no está configurado en Stripe. Faltan los price_id anuales.",
        }
    if billing_error == "portal_failed":
        return {
            "type": "warning",
            "text": "No se pudo abrir el portal de facturación de Stripe.",
        }
    return None


def build_invitation_email_html(recipient_email, role, invite_link, sender_name=None, agency_name=None):
    intro = (
        "Has recibido una invitacion para acceder a Ledged como trabajador."
        if role == "staff"
        else "Has recibido una invitacion para dar de alta tu gestoría en Ledged."
    )
    sender_block = f"<p><strong>Invitado por:</strong> {sender_name}</p>" if sender_name else ""
    agency_block = f"<p><strong>Gestoría:</strong> {agency_name}</p>" if agency_name else ""
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:640px;margin:0 auto;color:#1d2420">
      <h2 style="margin-bottom:12px">Invitacion a Ledged</h2>
      <p>Hola {recipient_email},</p>
      <p>{intro}</p>
      {sender_block}
      {agency_block}
      <p>Para activar tu acceso y definir tu contraseña, usa este enlace:</p>
      <p><a href="{invite_link}" style="display:inline-block;padding:12px 18px;background:#227c65;color:#ffffff;text-decoration:none;border-radius:999px">Aceptar invitacion</a></p>
      <p>Si el boton no funciona, copia y pega esta URL en tu navegador:</p>
      <p><a href="{invite_link}">{invite_link}</a></p>
      <p>Este enlace caduca en 7 dias.</p>
    </div>
    """


def create_user_invitation(
    *,
    email,
    role,
    created_by_user_id,
    agency_id=None,
    name=None,
    plan=None,
):
    email = _normalize_email(email)
    token = secrets.token_urlsafe(32)
    created_at = datetime.utcnow().isoformat()
    expires_at = (datetime.utcnow() + timedelta(days=7)).isoformat()
    with engine.begin() as conn:
        existing_user = conn.execute(
            select(users_table.c.id).where(users_table.c.email == email)
        ).first()
        if existing_user:
            return {"ok": False, "errors": ["El email ya está registrado."]}
        existing_invite = conn.execute(
            select(user_invitations_table.c.id)
            .where(user_invitations_table.c.email == email)
            .where(user_invitations_table.c.role == role)
            .where(user_invitations_table.c.accepted_at.is_(None))
            .where(user_invitations_table.c.expires_at >= created_at)
        ).first()
        if existing_invite:
            return {"ok": False, "errors": ["Ya existe una invitación pendiente para ese email."]}
        conn.execute(
            user_invitations_table.insert().values(
                email=email,
                role=role,
                agency_id=agency_id,
                name=name,
                plan=plan,
                token=token,
                expires_at=expires_at,
                accepted_at=None,
                created_by_user_id=created_by_user_id,
                created_at=created_at,
            )
        )
    return {"ok": True, "token": token}


def send_user_invitation_email(*, email, role, token, sender_name=None, agency_name=None, reply_to=None):
    invite_link = f"{get_app_base_url()}{url_for('accept_invitation', token=token)}"
    subject = (
        "Invitación a Ledged"
        if role == "agency"
        else "Invitación a tu cuenta de Ledged"
    )
    html = build_invitation_email_html(
        email,
        role,
        invite_link,
        sender_name=sender_name,
        agency_name=agency_name,
    )
    return send_email(
        email,
        subject,
        html,
        reply_to=reply_to,
        from_email=get_access_from_email(),
    )


def init_db():
    metadata.create_all(engine)
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())

    def add_column_if_missing(table_name, column_name, column_type):
        if table_name not in table_names:
            return
        columns = {col["name"] for col in inspector.get_columns(table_name)}
        if column_name in columns:
            return
        with engine.begin() as conn:
            conn.execute(
                text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")
            )

    def drop_not_null_if_needed(table_name, column_name):
        if table_name not in table_names:
            return
        columns = {col["name"]: col for col in inspector.get_columns(table_name)}
        column = columns.get(column_name)
        if not column or column.get("nullable", True):
            return
        if engine.dialect.name != "postgresql":
            return
        with engine.begin() as conn:
            conn.execute(
                text(f"ALTER TABLE {table_name} ALTER COLUMN {column_name} DROP NOT NULL")
            )

    add_column_if_missing("invoices", "user_id", "INTEGER")
    add_column_if_missing("invoices", "company_id", "INTEGER")
    add_column_if_missing("invoices", "payment_date", "VARCHAR")
    add_column_if_missing("invoices", "payment_dates", "TEXT")
    add_column_if_missing("invoices", "payment_completed_dates", "TEXT")
    add_column_if_missing("invoices", "vat_breakdown", "TEXT")
    add_column_if_missing("invoices", "withholding_amount", "FLOAT")
    add_column_if_missing("invoices", "vat_deductible", "BOOLEAN")
    add_column_if_missing("invoices", "extraction_source", "VARCHAR")
    add_column_if_missing("invoices", "confidence_score", "FLOAT")
    add_column_if_missing("invoices", "expense_family", "VARCHAR")
    add_column_if_missing("invoices", "expense_subtype", "VARCHAR")
    add_column_if_missing("invoices", "pnl_bucket", "VARCHAR")
    add_column_if_missing("invoices", "tax_model_targets", "TEXT")
    drop_not_null_if_needed("invoices", "vat_rate")
    drop_not_null_if_needed("invoices", "stored_filename")
    if "invoices" in table_names:
        with engine.begin() as conn:
            conn.execute(
                invoices_table.update()
                .where(invoices_table.c.stored_filename.is_not(None))
                .values(stored_filename="")
            )
            conn.execute(
                invoices_table.update()
                .where(invoices_table.c.ocr_text.is_not(None))
                .values(ocr_text=None)
            )
            conn.execute(
                invoices_table.update()
                .where(invoices_table.c.user_id.is_(None))
                .values(user_id=DEFAULT_USER_ID)
            )
            invoice_rows = conn.execute(
                select(
                    invoices_table.c.id,
                    invoices_table.c.expense_category,
                    invoices_table.c.vat_deductible,
                    invoices_table.c.vat_amount,
                    invoices_table.c.withholding_amount,
                    invoices_table.c.expense_family,
                    invoices_table.c.expense_subtype,
                    invoices_table.c.pnl_bucket,
                    invoices_table.c.tax_model_targets,
                )
            ).mappings().all()
            for row in invoice_rows:
                existing_targets = parse_tax_model_targets(row.get("tax_model_targets"))
                inferred_targets = parse_tax_model_targets(
                    derive_invoice_profile(
                        row.get("expense_category"),
                        row.get("vat_deductible"),
                        row.get("vat_amount"),
                        row.get("withholding_amount"),
                    ).get("tax_model_targets")
                )
                if (
                    row.get("expense_family")
                    and row.get("expense_subtype")
                    and row.get("pnl_bucket")
                    and row.get("tax_model_targets")
                    and (
                        "111" not in inferred_targets
                        or "111" in existing_targets
                    )
                ):
                    continue
                profile = derive_invoice_profile(
                    row.get("expense_category"),
                    row.get("vat_deductible"),
                    row.get("vat_amount"),
                    row.get("withholding_amount"),
                )
                conn.execute(
                    invoices_table.update()
                    .where(invoices_table.c.id == row["id"])
                    .values(**profile)
                )
            conn.execute(
                invoices_table.update()
                .where(invoices_table.c.vat_deductible.is_(None))
                .values(vat_deductible=True)
            )

    add_column_if_missing("facturacion", "user_id", "INTEGER")
    add_column_if_missing("facturacion", "company_id", "INTEGER")
    add_column_if_missing("facturacion", "invoice_date", "VARCHAR")
    add_column_if_missing("facturacion", "concept", "VARCHAR")
    add_column_if_missing("facturacion", "total_amount", "FLOAT")
    if "facturacion" in table_names:
        with engine.begin() as conn:
            conn.execute(
                facturacion_table.update()
                .where(facturacion_table.c.user_id.is_(None))
                .values(user_id=DEFAULT_USER_ID)
            )

    add_column_if_missing("no_invoice_expenses", "user_id", "INTEGER")
    add_column_if_missing("no_invoice_expenses", "company_id", "INTEGER")
    add_column_if_missing("no_invoice_expenses", "interest_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "vat_deductible", "BOOLEAN")
    add_column_if_missing("no_invoice_expenses", "vat_rate", "INTEGER")
    add_column_if_missing("no_invoice_expenses", "vat_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "base_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "withholding_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "payroll_employee_name", "VARCHAR")
    add_column_if_missing("no_invoice_expenses", "payroll_period", "VARCHAR")
    add_column_if_missing("no_invoice_expenses", "payroll_net_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "payroll_total_deductions_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "payroll_employer_cost_amount", "FLOAT")
    add_column_if_missing("no_invoice_expenses", "payment_date", "VARCHAR")
    add_column_if_missing("no_invoice_expenses", "payment_dates", "TEXT")
    add_column_if_missing("no_invoice_expenses", "payment_completed_dates", "TEXT")
    add_column_if_missing("no_invoice_expenses", "expense_family", "VARCHAR")
    add_column_if_missing("no_invoice_expenses", "expense_subtype", "VARCHAR")
    add_column_if_missing("no_invoice_expenses", "pnl_bucket", "VARCHAR")
    add_column_if_missing("no_invoice_expenses", "tax_model_targets", "TEXT")
    add_column_if_missing("loan_installments", "bank_name", "VARCHAR")
    add_column_if_missing("loan_installments", "payment_completed_dates", "TEXT")
    add_column_if_missing("companies", "vat_regime", "VARCHAR DEFAULT 'general'")
    add_column_if_missing("companies", "tax_periodicity", "VARCHAR DEFAULT 'quarterly'")
    add_column_if_missing("companies", "files_model_303", "BOOLEAN DEFAULT TRUE")
    add_column_if_missing("companies", "files_model_111", "BOOLEAN DEFAULT FALSE")
    add_column_if_missing("companies", "files_model_115", "BOOLEAN DEFAULT FALSE")
    add_column_if_missing("companies", "files_model_130", "BOOLEAN DEFAULT FALSE")
    add_column_if_missing("companies", "files_model_202", "BOOLEAN DEFAULT FALSE")
    add_column_if_missing("companies", "balance_manual_data", "TEXT")
    if "no_invoice_expenses" in table_names:
        with engine.begin() as conn:
            conn.execute(
                no_invoice_table.update()
                .where(no_invoice_table.c.user_id.is_(None))
                .values(user_id=DEFAULT_USER_ID)
            )
            conn.execute(
                no_invoice_table.update()
                .where(no_invoice_table.c.vat_deductible.is_(None))
                .values(vat_deductible=False)
            )
            conn.execute(
                no_invoice_table.update()
                .where(no_invoice_table.c.base_amount.is_(None))
                .values(base_amount=no_invoice_table.c.amount)
            )
            conn.execute(
                no_invoice_table.update()
                .where(no_invoice_table.c.withholding_amount.is_(None))
                .values(withholding_amount=0.0)
            )
            expense_rows = conn.execute(
                select(
                    no_invoice_table.c.id,
                    no_invoice_table.c.expense_type,
                    no_invoice_table.c.vat_deductible,
                    no_invoice_table.c.withholding_amount,
                    no_invoice_table.c.expense_family,
                    no_invoice_table.c.expense_subtype,
                    no_invoice_table.c.pnl_bucket,
                    no_invoice_table.c.tax_model_targets,
                )
            ).mappings().all()
            for row in expense_rows:
                existing_targets = parse_tax_model_targets(row.get("tax_model_targets"))
                inferred_targets = parse_tax_model_targets(
                    derive_no_invoice_profile(
                        row.get("expense_type"),
                        row.get("vat_deductible"),
                        row.get("withholding_amount"),
                    ).get("tax_model_targets")
                )
                if (
                    row.get("expense_family")
                    and row.get("expense_subtype")
                    and row.get("pnl_bucket")
                    and row.get("tax_model_targets")
                    and (
                        ("115" not in inferred_targets or "115" in existing_targets)
                        and ("111" not in inferred_targets or "111" in existing_targets)
                    )
                ):
                    continue
                profile = derive_no_invoice_profile(
                    row.get("expense_type"),
                    row.get("vat_deductible"),
                    row.get("withholding_amount"),
                )
                conn.execute(
                    no_invoice_table.update()
                    .where(no_invoice_table.c.id == row["id"])
                    .values(**profile)
                )

    add_column_if_missing("income_invoices", "user_id", "INTEGER")
    add_column_if_missing("income_invoices", "company_id", "INTEGER")
    add_column_if_missing("income_invoices", "payment_date", "VARCHAR")
    add_column_if_missing("income_invoices", "payment_dates", "TEXT")
    add_column_if_missing("income_invoices", "payment_completed_dates", "TEXT")
    add_column_if_missing("income_invoices", "vat_breakdown", "TEXT")
    add_column_if_missing("income_invoices", "extraction_source", "VARCHAR")
    add_column_if_missing("income_invoices", "confidence_score", "FLOAT")
    drop_not_null_if_needed("income_invoices", "vat_rate")
    drop_not_null_if_needed("income_invoices", "stored_filename")
    if "income_invoices" in table_names:
        with engine.begin() as conn:
            conn.execute(
                income_invoices_table.update()
                .where(income_invoices_table.c.stored_filename.is_not(None))
                .values(stored_filename="")
            )
            conn.execute(
                income_invoices_table.update()
                .where(income_invoices_table.c.ocr_text.is_not(None))
                .values(ocr_text=None)
            )
    if "income_invoices" in table_names:
        with engine.begin() as conn:
            conn.execute(
                income_invoices_table.update()
                .where(income_invoices_table.c.user_id.is_(None))
                .values(user_id=DEFAULT_USER_ID)
            )

    add_column_if_missing("companies", "agency_id", "INTEGER")
    add_column_if_missing("companies", "email", "VARCHAR")
    add_column_if_missing("companies", "phone", "VARCHAR")
    add_column_if_missing("companies", "assigned_user_id", "INTEGER")
    if "companies" in table_names:
        with engine.begin() as conn:
            conn.execute(
                companies_table.update()
                .where(companies_table.c.user_id.is_(None))
                .values(user_id=DEFAULT_USER_ID)
            )
            conn.execute(
                companies_table.update()
                .where(companies_table.c.agency_id.is_(None))
                .values(agency_id=companies_table.c.user_id)
            )

    add_column_if_missing("users", "agency_id", "INTEGER")
    if "users" in table_names:
        with engine.begin() as conn:
            conn.execute(
                users_table.update()
                .where(users_table.c.agency_id.is_(None))
                .values(agency_id=users_table.c.id)
            )

    add_column_if_missing("agencies", "phone", "VARCHAR")
    add_column_if_missing("agencies", "name", "VARCHAR")
    add_column_if_missing("agencies", "email", "VARCHAR")
    add_column_if_missing("agencies", "plan", "VARCHAR")
    add_column_if_missing("agencies", "status", "VARCHAR")
    add_column_if_missing("agencies", "stripe_customer_id", "VARCHAR")
    add_column_if_missing("agencies", "stripe_subscription_id", "VARCHAR")
    add_column_if_missing("agencies", "stripe_price_id", "VARCHAR")
    add_column_if_missing("agencies", "stripe_subscription_status", "VARCHAR")
    add_column_if_missing("agencies", "stripe_current_period_end", "VARCHAR")
    add_column_if_missing("agencies", "trial_ends_at", "VARCHAR")
    add_column_if_missing("agencies", "last_login_at", "VARCHAR")
    if "agencies" in table_names:
        with engine.begin() as conn:
            existing_ids = {
                row[0] for row in conn.execute(select(agencies_table.c.id)).all()
            }
            agency_users = conn.execute(
                select(
                    users_table.c.id,
                    users_table.c.email,
                    users_table.c.created_at,
                ).where(users_table.c.role == "agency")
            ).mappings().all()
            for user in agency_users:
                if user["id"] in existing_ids:
                    continue
                created_at = user.get("created_at") or datetime.utcnow().isoformat()
                trial_ends = (datetime.utcnow() + timedelta(days=14)).isoformat()
                conn.execute(
                    agencies_table.insert().values(
                        id=user["id"],
                        name=user["email"],
                        email=user["email"],
                        phone=None,
                        plan="starter",
                        status="trial",
                        stripe_customer_id=None,
                        stripe_subscription_id=None,
                        stripe_price_id=None,
                        stripe_subscription_status=None,
                        stripe_current_period_end=None,
                        trial_ends_at=trial_ends,
                        created_at=created_at,
                        last_login_at=None,
                    )
                )


def allowed_file(filename):
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_EXTENSIONS


def _row_to_user(row):
    if not row:
        return None
    return {
        "id": row["id"],
        "email": row["email"],
        "role": row["role"],
        "plan": row["plan"],
        "agency_id": row.get("agency_id") if isinstance(row, dict) else None,
        "is_active": bool(row["is_active"]),
    }


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    with engine.connect() as conn:
        row = conn.execute(
            select(
                users_table.c.id,
                users_table.c.email,
                users_table.c.role,
                users_table.c.plan,
                users_table.c.agency_id,
                users_table.c.is_active,
            ).where(users_table.c.id == user_id)
        ).mappings().first()
    user = _row_to_user(row)
    if user and not user["is_active"]:
        return None
    return user


def plan_allows(user, allowed_plans):
    if not user:
        return False
    return user.get("plan") in allowed_plans


def require_plan(allowed_plans):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            user = getattr(g, "current_user", None)
            if not plan_allows(user, allowed_plans):
                return jsonify({"ok": False, "errors": ["Plan insuficiente."]}), 403
            return view(*args, **kwargs)

        return wrapped

    return decorator


def require_owner(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = getattr(g, "current_user", None)
        if not user or user.get("role") != "owner":
            return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
        return view(*args, **kwargs)

    return wrapped


@app.before_request
def load_user_and_enforce_auth():
    g.current_user = get_current_user()
    path = request.path or ""
    if path.startswith("/static/"):
        return None
    if path == "/api/stripe/webhook":
        return None
    if path.startswith("/invite/"):
        return None
    if path.startswith("/login") or path.startswith("/register"):
        return None
    if path.startswith("/reset") or path.startswith("/reset-password"):
        return None
    if path.startswith("/health"):
        return None
    if path in {
        "/",
        "/landing",
        "/aviso-legal",
        "/privacidad",
        "/cookies",
        "/terminos",
    }:
        return None
    if not g.current_user:
        if path.startswith("/api/"):
            return jsonify({"ok": False, "error": "auth_required"}), 401
        return redirect(url_for("login"))
    return None


def parse_amount(value):
    if value is None:
        return None
    cleaned = value.replace("EUR", "").replace("euro", "").strip()
    cleaned = cleaned.replace(".", "").replace(",", ".") if "," in cleaned else cleaned
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_entity_name(value: str) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9]", "", value.lower())


def get_company_names_for_supplier_check(company_id: int, conn) -> list:
    if not company_id:
        return []
    row = conn.execute(
        select(companies_table.c.display_name, companies_table.c.legal_name).where(
            companies_table.c.id == company_id
        )
    ).first()
    if not row:
        return []
    return [row[0], row[1]]


def is_supplier_same_as_company(supplier: str, company_id: int, conn) -> bool:
    if not supplier or not company_id:
        return False
    normalized_supplier = normalize_entity_name(supplier)
    if not normalized_supplier:
        return False
    for name in get_company_names_for_supplier_check(company_id, conn):
        if normalize_entity_name(name) == normalized_supplier:
            return True
    return False


def normalize_vat_amounts(base_amount, vat_rate, vat_amount, total_amount):
    if vat_rate is None:
        return base_amount, vat_amount, total_amount
    if base_amount is None and total_amount is None:
        return base_amount, vat_amount, total_amount
    rate = vat_rate / 100
    if base_amount is None and total_amount is not None:
        base_amount = round(total_amount / (1 + rate), 2)
        vat_amount = round(total_amount - base_amount, 2)
        total_amount = round(total_amount, 2)
        return base_amount, vat_amount, total_amount
    if base_amount is not None:
        vat_amount = round(base_amount * rate, 2)
        total_amount = round(base_amount + vat_amount, 2)
    return base_amount, vat_amount, total_amount


def get_current_user_id():
    if getattr(g, "current_user", None):
        return int(g.current_user["id"])
    try:
        header_value = request.headers.get("X-User-Id")
        if header_value and str(header_value).isdigit():
            return int(header_value)
    except Exception:
        pass
    return DEFAULT_USER_ID


def get_data_owner_id():
    user = g.current_user
    if not user:
        return DEFAULT_USER_ID
    if user.get("role") == "staff":
        return int(user.get("agency_id") or user["id"])
    return int(user["id"])


def _resolve_company_id():
    company_id = None
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        company_id = payload.get("company_id") or payload.get("companyId")
    if company_id is None:
        company_id = request.args.get("company_id") or request.form.get("company_id")
    if company_id is None:
        return None
    try:
        return int(company_id)
    except (TypeError, ValueError):
        return None


def get_company_id(required=True):
    user_id = get_current_user_id()
    user_role = (g.current_user or {}).get("role")
    company_id = _resolve_company_id()
    with engine.connect() as conn:
        if company_id is None:
            if user_role == "staff":
                row = conn.execute(
                    select(companies_table.c.id).where(
                        companies_table.c.assigned_user_id == user_id
                    )
                ).first()
            elif user_role == "owner":
                row = conn.execute(select(companies_table.c.id)).first()
            else:
                row = conn.execute(
                    select(companies_table.c.id).where(
                        companies_table.c.agency_id == user_id
                    )
                ).first()
            if row:
                company_id = int(row[0])
        else:
            if user_role == "staff":
                exists = conn.execute(
                    select(companies_table.c.id)
                    .where(companies_table.c.assigned_user_id == user_id)
                    .where(companies_table.c.id == company_id)
                ).first()
            elif user_role == "owner":
                exists = conn.execute(
                    select(companies_table.c.id).where(
                        companies_table.c.id == company_id
                    )
                ).first()
            else:
                exists = conn.execute(
                    select(companies_table.c.id)
                    .where(companies_table.c.agency_id == user_id)
                    .where(companies_table.c.id == company_id)
                ).first()
            if not exists:
                company_id = None

    if required and company_id is None:
        return None
    return company_id


def is_company_accessible(company_id):
    if company_id is None:
        return False
    user_id = get_current_user_id()
    role = (g.current_user or {}).get("role")
    with engine.connect() as conn:
        if role == "staff":
            exists = conn.execute(
                select(companies_table.c.id)
                .where(companies_table.c.assigned_user_id == user_id)
                .where(companies_table.c.id == company_id)
            ).first()
        elif role == "owner":
            exists = conn.execute(
                select(companies_table.c.id).where(companies_table.c.id == company_id)
            ).first()
        else:
            exists = conn.execute(
                select(companies_table.c.id)
                .where(companies_table.c.agency_id == user_id)
                .where(companies_table.c.id == company_id)
            ).first()
    return bool(exists)


def _validate_nif(nif):
    if not nif:
        return False
    nif = nif.strip().upper()
    match = re.match(r"^(\d{8})([A-Z])$", nif)
    if not match:
        return False
    number, letter = match.groups()
    letters = "TRWAGMYFPDXBNJZSQVHLCKE"
    return letters[int(number) % 23] == letter


def _validate_cif(cif):
    if not cif:
        return False
    cif = cif.strip().upper()
    match = re.match(r"^([ABCDEFGHJKLMNPQRSUVW])(\d{7})([0-9A-J])$", cif)
    if not match:
        return False
    letter, digits, control = match.groups()
    total = 0
    for idx, char in enumerate(digits, start=1):
        n = int(char)
        if idx % 2 == 1:
            n *= 2
            total += n // 10 + n % 10
        else:
            total += n
    control_num = (10 - (total % 10)) % 10
    control_digit = str(control_num)
    control_letter = "JABCDEFGHI"[control_num]
    if letter in "PQRSW":
        return control == control_letter
    if letter in "ABEH":
        return control == control_digit
    return control in {control_digit, control_letter}


def validate_tax_id(tax_id, company_type):
    if company_type == "individual":
        return _validate_nif(tax_id)
    if company_type == "company":
        return _validate_cif(tax_id)
    return False


def resolve_assigned_staff(agency_id, staff_id):
    if not staff_id:
        return None
    try:
        staff_id = int(staff_id)
    except (TypeError, ValueError):
        return None
    with engine.connect() as conn:
        staff = conn.execute(
            select(users_table.c.id)
            .where(users_table.c.id == staff_id)
            .where(users_table.c.role == "staff")
            .where(users_table.c.agency_id == agency_id)
        ).first()
    if not staff:
        return None
    return staff_id


def normalize_date(value):
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw).isoformat()
    except ValueError:
        return None


def compute_payment_date(invoice_date_value, payment_date_value=None):
    payment_date = normalize_date(payment_date_value)
    if payment_date:
        return payment_date
    invoice_date = normalize_date(invoice_date_value)
    if not invoice_date:
        return None
    try:
        base_date = date.fromisoformat(invoice_date)
    except ValueError:
        return None
    return (base_date + timedelta(days=30)).isoformat()


def parse_payment_dates(raw_value):
    if not raw_value:
        return []
    if isinstance(raw_value, list):
        values = raw_value
    else:
        value = raw_value
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return []
            try:
                parsed = json.loads(value)
                if isinstance(parsed, list):
                    values = parsed
                else:
                    values = [parsed]
            except json.JSONDecodeError:
                values = [item.strip() for item in re.split(r"[;,]\s*", value) if item.strip()]
        else:
            values = [value]

    normalized = []
    for item in values:
        norm = normalize_date(str(item)) if item else None
        if norm:
            normalized.append(norm)
    return sorted(set(normalized))


def serialize_payment_dates(values):
    normalized = parse_payment_dates(values)
    return json.dumps(normalized) if normalized else None


DOCUMENT_PROCESSING_TYPES = {
    "purchase_invoice",
    "sales_invoice",
    "receipt",
    "bank_statement",
    "loan_document",
    "payroll",
    "social_security",
    "tax_model",
    "non_accounting",
    "unknown",
}


def allowed_processing_file(filename):
    return os.path.splitext(filename.lower())[1] in DOCUMENT_CENTER_ALLOWED_EXTENSIONS


def local_uploaded_file_path(key: str) -> str:
    base_dir = os.getenv("UPLOAD_FOLDER") or os.path.join(tempfile.gettempdir(), "uploads")
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, key)


def normalize_period_label(raw_period: str) -> str:
    value = (raw_period or "").strip()
    return value or f"{date.today().year}"


def json_dumps(value):
    return json.dumps(value, ensure_ascii=False)


def json_loads_dict(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    try:
        loaded = json.loads(value)
    except Exception:
        return {}
    return loaded if isinstance(loaded, dict) else {}


def json_loads_list(value):
    if not value:
        return []
    if isinstance(value, list):
        return value
    try:
        loaded = json.loads(value)
    except Exception:
        return []
    return loaded if isinstance(loaded, list) else []


def append_document_audit(existing_value, action, user_id=None, payload=None):
    history = json_loads_list(existing_value)
    history.append(
        {
            "action": action,
            "user_id": user_id,
            "timestamp": datetime.utcnow().isoformat(),
            "payload": payload or {},
        }
    )
    return json_dumps(history)


def document_file_type_from_name(filename):
    ext = os.path.splitext((filename or "").lower())[1]
    return ext.lstrip(".") or "unknown"


def document_content_type_from_name(filename):
    extension = os.path.splitext((filename or "").lower())[1]
    return {
        ".pdf": "application/pdf",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".csv": "text/csv",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
    }.get(extension, "application/octet-stream")


def document_storage_key(company_id, batch_id, filename):
    safe_name = secure_filename(filename or "documento")
    return f"document-center/company-{company_id}/batch-{batch_id}/{secrets.token_hex(8)}-{safe_name}"


def persist_document_bytes(company_id, batch_id, filename, data, content_type=None):
    key = document_storage_key(company_id, batch_id, filename)
    upload_bytes(data, key, content_type=content_type)
    return key, get_public_url(key)


def extract_text_from_spreadsheet_archive(file_bytes):
    lines = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            shared_strings = []
            if "xl/sharedStrings.xml" in archive.namelist():
                try:
                    shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                    for node in shared_root.iter():
                        if node.tag.endswith("}t") and node.text:
                            shared_strings.append(node.text.strip())
                except Exception:
                    shared_strings = []

            sheet_names = sorted(
                name
                for name in archive.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            )[:3]
            for sheet_name in sheet_names:
                try:
                    root = ET.fromstring(archive.read(sheet_name))
                except Exception:
                    continue
                for row in root.iter():
                    if not row.tag.endswith("}row"):
                        continue
                    values = []
                    for cell in row:
                        if not cell.tag.endswith("}c"):
                            continue
                        raw_value = None
                        cell_type = cell.attrib.get("t")
                        for child in cell:
                            if child.tag.endswith("}v") and child.text not in (None, ""):
                                raw_value = child.text.strip()
                                break
                            if child.tag.endswith("}is"):
                                texts = [node.text.strip() for node in child.iter() if node.tag.endswith("}t") and node.text]
                                if texts:
                                    raw_value = " ".join(texts)
                                    break
                        if raw_value in (None, ""):
                            continue
                        if cell_type == "s":
                            try:
                                shared_index = int(raw_value)
                                raw_value = shared_strings[shared_index] if 0 <= shared_index < len(shared_strings) else raw_value
                            except (TypeError, ValueError):
                                pass
                        values.append(str(raw_value).strip())
                    if values:
                        lines.append(" | ".join(values))
                    if len(lines) >= 200:
                        return "\n".join(lines)
    except zipfile.BadZipFile:
        return ""
    except Exception:
        app.logger.exception("No se pudo extraer texto del spreadsheet como archivo ZIP estructurado")
        return ""
    return "\n".join(lines)


def extract_text_from_document_bytes(file_bytes, filename):
    extension = os.path.splitext((filename or "").lower())[1]
    if extension == ".pdf":
        text_value = _extract_pdf_text_from_bytes(file_bytes) or ""
        if not text_value or len(text_value.strip()) < 60:
            text_value = _extract_pdf_text_ocr_from_bytes(file_bytes) or ""
        return text_value
    if extension in {".jpg", ".jpeg", ".png"}:
        return _extract_image_text_ocr_from_bytes(file_bytes) or ""
    if extension == ".csv":
        try:
            return file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return file_bytes.decode("latin-1", errors="ignore")
    if extension in {".xlsx", ".xls"}:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            lines = []
            for sheet in workbook.worksheets[:3]:
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if values:
                        lines.append(" | ".join(values))
                    if len(lines) >= 200:
                        return "\n".join(lines)
            return "\n".join(lines)
        except Exception:
            app.logger.warning(
                "Fallo leyendo spreadsheet %s con openpyxl. Se intenta extracción degradada.",
                filename,
                exc_info=True,
            )
            fallback_text = extract_text_from_spreadsheet_archive(file_bytes)
            if fallback_text.strip():
                return fallback_text
            try:
                return file_bytes.decode("latin-1", errors="ignore")[:12000]
            except Exception:
                return ""
    return ""


def iter_uploaded_documents(files):
    for uploaded in files:
        if not uploaded or not uploaded.filename:
            continue
        original_name = os.path.basename(uploaded.filename)
        if not allowed_processing_file(original_name):
            continue
        file_bytes = uploaded.read()
        extension = os.path.splitext(original_name.lower())[1]
        if extension == ".zip":
            try:
                with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                    for member in archive.infolist():
                        if member.is_dir():
                            continue
                        member_name = os.path.basename(member.filename)
                        if not member_name or not allowed_processing_file(member_name):
                            continue
                        inner_bytes = archive.read(member)
                        inner_ext = os.path.splitext(member_name.lower())[1]
                        if inner_ext == ".zip":
                            continue
                        yield {
                            "filename": member_name,
                            "bytes": inner_bytes,
                            "content_type": None,
                        }
            except zipfile.BadZipFile:
                continue
        else:
            yield {
                "filename": original_name,
                "bytes": file_bytes,
                "content_type": uploaded.mimetype,
            }


def detect_document_type(text_value, filename="", company_names=None):
    text_lower = (text_value or "").lower()
    filename_lower = (filename or "").lower()
    company_names = [item.lower() for item in (company_names or []) if item]

    if any(token in text_lower for token in ("modelo 303", "modelo 111", "modelo 115", "modelo 200", "modelo 202")):
        return "tax_model"
    if any(token in text_lower for token in ("seguridad social", "rlc", "rnt")):
        return "social_security"
    if any(token in text_lower for token in ("nómina", "nomina", "devengos", "deducciones")):
        return "payroll"
    if any(token in text_lower for token in ("préstamo", "prestamo", "cuadro de amortización", "cuadro de amortizacion", "tin", "tae")):
        return "loan_document"
    if any(token in text_lower for token in ("extracto", "saldo", "movimientos", "iban")) and len(re.findall(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text_lower)) >= 2:
        return "bank_statement"
    if "factura simplificada" in text_lower or "ticket" in text_lower:
        return "receipt"
    if "factura" in text_lower or "nº factura" in text_lower or "base imponible" in text_lower:
        if any(name and name in text_lower for name in company_names) and any(
            token in text_lower for token in ("cliente", "destinatario", "factura emitida")
        ):
            return "sales_invoice"
        if any(token in filename_lower for token in ("venta", "ingreso", "sales", "emitida")):
            return "sales_invoice"
        return "purchase_invoice"
    if any(token in filename_lower for token in ("extracto", "banco", "bank")):
        return "bank_statement"
    if any(token in filename_lower for token in ("nomina", "nómina")):
        return "payroll"
    return "unknown"


def extract_tax_id_from_text(text_value):
    if not text_value:
        return None
    match = re.search(r"\b([A-Z]\d{7}[A-Z0-9]|\d{8}[A-Z])\b", text_value.upper())
    return match.group(1) if match else None


def extract_invoice_number_from_text(text_value):
    if not text_value:
        return None
    patterns = [
        r"(?:factura|fra\.?|n[ºo]\s*factura)[:\s#-]*([A-Z0-9\/\.-]{3,})",
        r"(?:invoice\s*(?:no|number)?)[:\s#-]*([A-Z0-9\/\.-]{3,})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text_value, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def extract_labeled_amount(text_value, labels):
    if not text_value:
        return None
    patterns = [
        rf"(?:{'|'.join(labels)})\s*[:\-]?\s*([\d\.\s,]+)",
        rf"([\d\.\s,]+)\s*(?:€|eur)?\s*(?:{'|'.join(labels)})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text_value, flags=re.IGNORECASE)
        if match:
            amount = parse_amount(match.group(1))
            if amount is not None:
                return amount
    return None


def extract_payroll_period(text_value):
    if not text_value:
        return None
    month_regex = r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)"
    patterns = [
        rf"(?:periodo|período|nomina|nómina)\s*(?:de)?\s*{month_regex}\s*(?:de)?\s*(20\d{{2}})",
        rf"{month_regex}\s*(20\d{{2}})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text_value, flags=re.IGNORECASE)
        if match:
            month_value = match.group(1)
            year_value = match.group(2)
            return f"{month_value.title()} {year_value}"
    return None


def extract_model_period(text_value):
    if not text_value:
        return None
    compact_quarter_match = re.search(
        r"([1-4])\s*T\s*(20\d{2})",
        text_value,
        flags=re.IGNORECASE,
    )
    if compact_quarter_match:
        return f"T{compact_quarter_match.group(1)} {compact_quarter_match.group(2)}"
    quarter_match = re.search(
        r"([1-4])\s*[ºo]?\s*(?:trimestre|trim)\s*(20\d{2})",
        text_value,
        flags=re.IGNORECASE,
    )
    if quarter_match:
        return f"T{quarter_match.group(1)} {quarter_match.group(2)}"
    year_match = re.search(r"(?:ejercicio|año)\s*(20\d{2})", text_value, flags=re.IGNORECASE)
    if year_match:
        return year_match.group(1)
    return None


def extract_tax_model_data(text_value):
    lower_text = (text_value or "").lower()
    model_name = next(
        (
            model
            for model in ("303", "111", "115", "200", "202", "390", "190", "180", "130")
            if f"modelo {model}" in lower_text
        ),
        None,
    )
    payable_amount = extract_labeled_amount(
        text_value,
        [
            "a ingresar",
            "resultado liquidaci[oó]n",
            "resultado de la liquidaci[oó]n",
            "importe a ingresar",
            "total a ingresar",
            "cuota",
        ],
    )
    refund_amount = extract_labeled_amount(
        text_value,
        [
            "a devolver",
            "importe a devolver",
            "solicitud de devoluci[oó]n",
        ],
    )
    offset_amount = extract_labeled_amount(
        text_value,
        [
            "a compensar",
            "importe a compensar",
            "cuota a compensar",
        ],
    )
    filing_status = None
    normalized_amount = payable_amount
    if "sin actividad" in lower_text or "sin operaciones" in lower_text or "sin cuota" in lower_text:
        filing_status = "sin_actividad"
        normalized_amount = 0.0
    elif offset_amount is not None or "a compensar" in lower_text:
        filing_status = "a_compensar"
        normalized_amount = offset_amount if offset_amount is not None else 0.0
    elif refund_amount is not None or "a devolver" in lower_text:
        filing_status = "a_devolver"
        normalized_amount = refund_amount if refund_amount is not None else 0.0
    elif payable_amount is not None or "a ingresar" in lower_text:
        filing_status = "a_ingresar"

    return {
        "model_name": model_name,
        "tax_period": extract_model_period(text_value),
        "filing_status": filing_status,
        "amount": normalized_amount,
        "total_amount": normalized_amount,
        "payable_amount": payable_amount,
        "offset_amount": offset_amount,
        "refund_amount": refund_amount,
    }


def normalize_tax_period_key(value):
    if not value:
        return ""
    raw = str(value).strip().lower()
    if not raw:
        return ""
    raw = raw.replace("setiembre", "septiembre")
    quarter_match = re.search(r"t\s*([1-4])\s*(20\d{2})", raw, flags=re.IGNORECASE)
    if not quarter_match:
        quarter_match = re.search(
            r"([1-4])\s*(?:t|trimestre|trim)\s*(20\d{2})",
            raw,
            flags=re.IGNORECASE,
        )
    if quarter_match:
        return f"T{quarter_match.group(1)} {quarter_match.group(2)}"
    months_match = re.search(r"(3|9|11)\s*meses?\s*(20\d{2})", raw, flags=re.IGNORECASE)
    if months_match:
        return f"{months_match.group(1)} MESES {months_match.group(2)}"
    year_match = re.search(r"(20\d{2})", raw)
    month_names = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    for name in month_names:
        if name in raw and year_match:
            return f"{name.title()} {year_match.group(1)}"
    return raw.upper()


def _fetch_registered_tax_filings(conn, company_id):
    rows = conn.execute(
        select(processed_documents_table)
        .where(processed_documents_table.c.company_id == company_id)
        .where(processed_documents_table.c.detected_document_type == "tax_model")
        .where(processed_documents_table.c.registered_at.is_not(None))
        .order_by(processed_documents_table.c.registered_at.desc(), processed_documents_table.c.id.desc())
    ).mappings().all()
    lookup = {}
    for row in rows:
        data = effective_document_data(row)
        model_name = str(data.get("model_name") or "").strip()
        period_key = normalize_tax_period_key(data.get("tax_period"))
        if not model_name or not period_key:
            continue
        lookup[(model_name, period_key)] = {
            "document_id": row["id"],
            "filing_status": data.get("filing_status") or "a_ingresar",
            "amount": parse_amount(data.get("amount")) or parse_amount(data.get("total_amount")) or 0.0,
            "payable_amount": parse_amount(data.get("payable_amount")),
            "offset_amount": parse_amount(data.get("offset_amount")),
            "refund_amount": parse_amount(data.get("refund_amount")),
            "tax_period": data.get("tax_period"),
            "registered_at": row.get("registered_at"),
            "concept": data.get("concept"),
        }
    return lookup


def enrich_payroll_data(extracted, text_value, fallback_name=None):
    employee_name = extracted.get("employee_name") or extracted.get("provider_name")
    if not employee_name and fallback_name:
        clean_name = os.path.splitext(os.path.basename(fallback_name))[0]
        clean_name = re.sub(r"^\d{1,2}[._-]?\d{4}\s*", "", clean_name).strip()
        employee_name = clean_name or None
    gross_amount = (
        parse_amount(extracted.get("base_amount"))
        or extract_labeled_amount(text_value, ["total devengado", "devengos", "bruto", "salario bruto"])
    )
    net_amount = (
        parse_amount(extracted.get("payroll_net_amount"))
        or parse_amount(extracted.get("total_amount"))
        or extract_labeled_amount(text_value, ["liquido a percibir", "líquido a percibir", "neto", "total liquido"])
    )
    deductions_amount = (
        parse_amount(extracted.get("payroll_total_deductions_amount"))
        or extract_labeled_amount(text_value, ["total deducciones", "deducciones"])
    )
    employer_cost_amount = (
        parse_amount(extracted.get("payroll_employer_cost_amount"))
        or extract_labeled_amount(text_value, ["coste empresa", "total empresa", "coste total empresa"])
    )
    if gross_amount is None and net_amount is not None and deductions_amount is not None:
        gross_amount = round(net_amount + deductions_amount, 2)
    if deductions_amount is None and gross_amount is not None and net_amount is not None:
        deductions_amount = round(max(gross_amount - net_amount, 0), 2)
    extracted["employee_name"] = employee_name
    extracted["provider_name"] = employee_name
    extracted["payroll_period"] = extracted.get("payroll_period") or extract_payroll_period(text_value)
    extracted["base_amount"] = gross_amount
    extracted["payroll_net_amount"] = net_amount
    extracted["payroll_total_deductions_amount"] = deductions_amount
    extracted["payroll_employer_cost_amount"] = employer_cost_amount
    extracted["total_amount"] = net_amount if net_amount is not None else extracted.get("total_amount")
    extracted["concept"] = extracted.get("concept") or (
        f"Nómina - {employee_name}" if employee_name else "Nómina"
    )
    extracted["vat_deductible"] = False
    extracted["vat_amount"] = None
    extracted["vat_rate"] = None
    return extracted


def extract_confidence_and_fields(
    detected_type,
    extracted_data,
    text_value="",
):
    score = 0
    issue_type = None
    issue_description = None
    normalized = extracted_data or {}

    if detected_type in {"purchase_invoice", "sales_invoice", "receipt"}:
        if detected_type:
            score += 20
        if normalized.get("provider_name") or normalized.get("client_name"):
            score += 15
        if normalized.get("tax_id"):
            score += 15
        if normalized.get("invoice_date"):
            score += 15
        if normalized.get("invoice_number"):
            score += 15
        if normalized.get("base_amount") is not None:
            score += 10
        if normalized.get("vat_amount") is not None and normalized.get("total_amount") is not None:
            score += 10
    elif detected_type == "payroll":
        if detected_type:
            score += 30
        if normalized.get("employee_name") or normalized.get("provider_name"):
            score += 20
        if normalized.get("payroll_period"):
            score += 15
        if normalized.get("base_amount") is not None:
            score += 15
        if normalized.get("payroll_net_amount") is not None:
            score += 10
        if normalized.get("payroll_total_deductions_amount") is not None:
            score += 10
    elif detected_type == "tax_model":
        if detected_type:
            score += 35
        if normalized.get("model_name"):
            score += 20
        if normalized.get("filing_status"):
            score += 10
        if normalized.get("tax_period"):
            score += 15
        if normalized.get("amount") is not None or normalized.get("total_amount") is not None:
            score += 15
        if normalized.get("invoice_date") or normalized.get("payment_date"):
            score += 10
    else:
        if detected_type != "unknown":
            score += 35
        if text_value and len(text_value.strip()) > 60:
            score += 20
        if normalized.get("invoice_date") or normalized.get("payment_date"):
            score += 15
        if normalized.get("total_amount") is not None or normalized.get("amount") is not None:
            score += 15
        if normalized.get("provider_name") or normalized.get("client_name") or normalized.get("bank_name"):
            score += 15

    score = max(0, min(score, 100))
    if score >= 80:
        validation_status = "ready_to_register"
    else:
        validation_status = "needs_review"
        if score < 50:
            issue_type = "low_confidence"
            issue_description = "Confianza baja en la clasificación o en la extracción."
    return score, validation_status, issue_type, issue_description


def build_document_extracted_data(
    *,
    detected_type,
    text_value,
    filename,
    file_bytes,
    content_type,
    company_names,
    known_suppliers,
):
    extracted = {
        "filename": filename,
        "invoice_number": extract_invoice_number_from_text(text_value),
        "tax_id": extract_tax_id_from_text(text_value),
        "invoice_date": _extract_first_date(text_value),
        "payment_date": _find_payment_date_by_keywords(text_value),
        "payment_dates": _find_payment_dates_by_keywords(text_value, _extract_first_date(text_value)),
        "analysis_text": (text_value or "")[:12000],
    }
    ai_doc_type = {
        "purchase_invoice": "expense",
        "sales_invoice": "income",
        "payroll": "expense_payroll",
        "receipt": "expense_other",
        "social_security": "expense_other",
    }.get(detected_type)
    if ai_doc_type and file_bytes:
        ai_data = _analyze_invoice_with_timeout(
            file_bytes=file_bytes,
            filename=filename,
            stored_name=filename,
            mime_type=content_type or document_content_type_from_name(filename),
            document_type=ai_doc_type,
            company_names=company_names,
            known_suppliers=known_suppliers,
        )
        if isinstance(ai_data, dict):
            extracted.update({key: value for key, value in ai_data.items() if value not in (None, "", [])})
    if detected_type == "payroll":
        extracted = enrich_payroll_data(extracted, text_value, filename)
    elif detected_type == "loan_document":
        installments = parse_loan_installments_from_text(text_value or "")
        if not installments and text_value:
            installments = extract_loan_schedule(text_value) or []
        extracted["installments"] = installments
        extracted["bank_name"] = " ".join(re.findall(r"(?:banco|bank|caixa|bbva|santander|sabadell)", text_value, flags=re.IGNORECASE)[:2]) or None
        if installments:
            first_item = installments[0]
            extracted["payment_date"] = first_item.get("payment_date")
            extracted["amount"] = first_item.get("total_amount")
    elif detected_type == "bank_statement":
        extracted["movement_lines"] = [line.strip() for line in (text_value or "").splitlines() if line.strip()][:100]
    elif detected_type == "tax_model":
        extracted.update(
            {
                key: value
                for key, value in extract_tax_model_data(text_value).items()
                if value not in (None, "", [])
            }
        )
    return extracted


def effective_document_data(row):
    corrected = json_loads_dict(row.get("corrected_data_json"))
    if corrected:
        return corrected
    return json_loads_dict(row.get("original_extracted_data_json"))


def serialize_document_batch(row):
    return {
        "id": row["id"],
        "companyId": row["company_id"],
        "uploadedByUserId": row["uploaded_by_user_id"],
        "period": row["period"],
        "totalDocuments": int(row.get("total_documents") or 0),
        "processedDocuments": int(row.get("processed_documents") or 0),
        "readyDocuments": int(row.get("ready_documents") or 0),
        "reviewDocuments": int(row.get("review_documents") or 0),
        "duplicateDocuments": int(row.get("duplicate_documents") or 0),
        "failedDocuments": int(row.get("failed_documents") or 0),
        "status": row.get("status"),
        "createdAt": row.get("created_at"),
        "updatedAt": row.get("updated_at"),
    }


def get_company_names_for_analysis(conn, company_id):
    if not conn or not company_id:
        return []
    row = conn.execute(
        select(companies_table.c.display_name, companies_table.c.legal_name).where(
            companies_table.c.id == company_id
        )
    ).mappings().first()
    if not row:
        return []
    return [row.get("display_name"), row.get("legal_name")]


def build_processed_document_payload(
    *,
    file_bytes,
    filename,
    content_type,
    company_id,
    data_owner_id,
    current_user_id,
    batch_id,
    period,
    company_names,
    known_suppliers,
):
    uploaded_at = datetime.utcnow().isoformat()
    content_hash = hashlib.sha256(file_bytes).hexdigest()
    stored_path, file_url = persist_document_bytes(
        company_id,
        batch_id,
        filename,
        file_bytes,
        content_type=content_type or document_content_type_from_name(filename),
    )
    text_value = extract_text_from_document_bytes(file_bytes, filename)
    detected_type = detect_document_type(text_value, filename, company_names)
    extracted_data = build_document_extracted_data(
        detected_type=detected_type,
        text_value=text_value,
        filename=filename,
        file_bytes=file_bytes,
        content_type=content_type,
        company_names=company_names,
        known_suppliers=known_suppliers,
    )
    confidence_score, validation_status, issue_type, issue_description = extract_confidence_and_fields(
        detected_type,
        extracted_data,
        text_value=text_value,
    )
    return {
        "company_id": company_id,
        "uploaded_by_user_id": current_user_id,
        "source_batch_id": batch_id,
        "original_filename": filename,
        "storage_path": stored_path,
        "file_url": file_url,
        "file_type": document_file_type_from_name(filename),
        "uploaded_at": uploaded_at,
        "processing_status": "processed",
        "detected_document_type": detected_type,
        "confidence_score": confidence_score,
        "extracted_data_json": json_dumps(extracted_data),
        "original_extracted_data_json": json_dumps(extracted_data),
        "corrected_data_json": None,
        "validation_status": validation_status,
        "issue_type": issue_type,
        "issue_description": issue_description,
        "linked_accounting_record_id": None,
        "linked_accounting_record_type": None,
        "approved_at": None,
        "approved_by_user_id": None,
        "rejected_at": None,
        "rejected_by_user_id": None,
        "registered_at": None,
        "registered_by_user_id": None,
        "period": period,
        "extracted_text": text_value[:12000] if text_value else None,
        "duplicate_of_document_id": None,
        "content_hash": content_hash,
        "audit_log_json": append_document_audit(
            None,
            "uploaded",
            current_user_id,
            {"filename": filename, "detected_type": detected_type},
        ),
        "created_at": uploaded_at,
        "updated_at": uploaded_at,
    }


def document_belongs_to_company(row, company_id):
    return row and int(row["company_id"]) == int(company_id)


def get_processed_document_for_company(conn, document_id, company_id):
    row = conn.execute(
        select(processed_documents_table).where(processed_documents_table.c.id == document_id)
    ).mappings().first()
    if not document_belongs_to_company(row, company_id):
        return None
    return row


def update_processed_document_record(conn, document_id, **values):
    values["updated_at"] = datetime.utcnow().isoformat()
    conn.execute(
        processed_documents_table.update()
        .where(processed_documents_table.c.id == document_id)
        .values(**values)
    )


def register_processed_document(conn, row, data_owner_id, current_user_id):
    doc_type = row.get("detected_document_type")
    data = effective_document_data(row)
    now = datetime.utcnow().isoformat()
    filename = row.get("original_filename") or "Documento"
    invoice_date = normalize_date(data.get("invoice_date")) or date.today().isoformat()
    payment_dates = parse_payment_dates(data.get("payment_dates"))
    payment_date = compute_payment_date(
        invoice_date,
        data.get("payment_date") or (payment_dates[0] if payment_dates else None),
    )

    if doc_type == "purchase_invoice":
        vat_breakdown = parse_vat_breakdown(data.get("vat_breakdown"))
        vat_breakdown_json = json.dumps(vat_breakdown) if vat_breakdown else None
        vat_rate = infer_vat_rate_from_breakdown(vat_breakdown) if vat_breakdown else data.get("vat_rate")
        if vat_rate is None:
            vat_rate = 0
        try:
            vat_rate = int(vat_rate)
        except (TypeError, ValueError):
            vat_rate = 0
        base_amount, vat_amount, total_amount = normalize_vat_amounts(
            parse_amount(data.get("base_amount")),
            vat_rate,
            parse_amount(data.get("vat_amount")),
            parse_amount(data.get("total_amount")),
        )
        supplier = (data.get("provider_name") or data.get("supplier") or "").strip() or "Proveedor pendiente"
        expense_category = "non_deductible" if data.get("vat_deductible") is False else "with_invoice"
        expense_profile = derive_invoice_profile(
            expense_category,
            data.get("vat_deductible") is not False,
            vat_amount,
            parse_amount(data.get("withholding_amount")) or 0.0,
        )
        withholding_amount = parse_amount(data.get("withholding_amount")) or 0.0
        result = conn.execute(
            invoices_table.insert().values(
                user_id=data_owner_id,
                company_id=row["company_id"],
                original_filename=filename,
                stored_filename=row.get("storage_path"),
                invoice_date=invoice_date,
                supplier=supplier,
                base_amount=base_amount or 0.0,
                vat_deductible=data.get("vat_deductible") is not False,
                vat_rate=vat_rate,
                vat_amount=vat_amount,
                total_amount=total_amount or 0.0,
                vat_breakdown=vat_breakdown_json,
                withholding_amount=withholding_amount,
                payment_date=payment_date,
                payment_dates=serialize_payment_dates(payment_dates),
                ocr_text=row.get("extracted_text"),
                extraction_source="document_center",
                confidence_score=row.get("confidence_score"),
                expense_category=expense_category,
                created_at=now,
                **expense_profile,
            )
        )
        supplier and store_known_supplier(conn, data_owner_id, row["company_id"], supplier)
        return str(result.inserted_primary_key[0]), "invoice"

    if doc_type == "sales_invoice":
        vat_breakdown = parse_vat_breakdown(data.get("vat_breakdown"))
        vat_breakdown_json = json.dumps(vat_breakdown) if vat_breakdown else None
        vat_rate = infer_vat_rate_from_breakdown(vat_breakdown) if vat_breakdown else data.get("vat_rate")
        if vat_rate is None:
            vat_rate = 0
        try:
            vat_rate = int(vat_rate)
        except (TypeError, ValueError):
            vat_rate = 0
        base_amount, vat_amount, total_amount = normalize_vat_amounts(
            parse_amount(data.get("base_amount")),
            vat_rate,
            parse_amount(data.get("vat_amount")),
            parse_amount(data.get("total_amount")),
        )
        client = (data.get("client_name") or data.get("client") or "").strip() or "Cliente pendiente"
        result = conn.execute(
            income_invoices_table.insert().values(
                user_id=data_owner_id,
                company_id=row["company_id"],
                original_filename=filename,
                stored_filename=row.get("storage_path"),
                invoice_date=invoice_date,
                client=client,
                base_amount=base_amount or 0.0,
                vat_rate=vat_rate,
                vat_amount=vat_amount,
                total_amount=total_amount or 0.0,
                vat_breakdown=vat_breakdown_json,
                payment_date=payment_date,
                payment_dates=serialize_payment_dates(payment_dates),
                ocr_text=row.get("extracted_text"),
                extraction_source="document_center",
                confidence_score=row.get("confidence_score"),
                created_at=now,
            )
        )
        return str(result.inserted_primary_key[0]), "income_invoice"

    if doc_type in {"receipt", "payroll", "social_security"}:
        if doc_type == "payroll":
            expense_type = "nomina"
            concept = (data.get("concept") or data.get("employee_name") or filename).strip()
        elif doc_type == "social_security":
            expense_type = "seguridad_social"
            concept = (data.get("concept") or "Seguridad Social").strip()
        else:
            expense_type = "otro"
            concept = (data.get("concept") or data.get("provider_name") or filename).strip()
        amount = parse_amount(data.get("total_amount")) or parse_amount(data.get("amount")) or 0.0
        withholding_amount = parse_amount(data.get("withholding_amount")) or 0.0
        vat_deductible = bool(data.get("vat_deductible")) if doc_type == "receipt" else False
        vat_rate = None
        vat_amount = None
        base_amount = parse_amount(data.get("base_amount"))
        if vat_deductible:
            try:
                vat_rate = int(data.get("vat_rate") or 0)
            except (TypeError, ValueError):
                vat_rate = 0
            if base_amount is None:
                base_amount = round(amount / (1 + vat_rate / 100), 2) if vat_rate >= 0 else amount
            vat_amount = parse_amount(data.get("vat_amount"))
            if vat_amount is None:
                vat_amount = round(amount - base_amount, 2)
        if base_amount is None:
            base_amount = amount
        expense_profile = derive_no_invoice_profile(
            expense_type,
            vat_deductible,
            withholding_amount,
        )
        result = conn.execute(
            no_invoice_table.insert().values(
                user_id=data_owner_id,
                company_id=row["company_id"],
                expense_date=invoice_date,
                payment_date=payment_date or invoice_date,
                payment_dates=serialize_payment_dates(payment_dates or [payment_date or invoice_date]),
                concept=concept,
                amount=amount,
                interest_amount=None,
                vat_deductible=vat_deductible,
                vat_rate=vat_rate,
                vat_amount=vat_amount,
                base_amount=base_amount,
                withholding_amount=withholding_amount,
                payroll_employee_name=(data.get("employee_name") or "").strip() or None,
                payroll_period=(data.get("payroll_period") or row.get("period") or "").strip() or None,
                payroll_net_amount=parse_amount(data.get("payroll_net_amount")),
                payroll_total_deductions_amount=parse_amount(data.get("payroll_total_deductions_amount")),
                payroll_employer_cost_amount=parse_amount(data.get("payroll_employer_cost_amount")),
                expense_type=expense_type,
                deductible=True,
                created_at=now,
                **expense_profile,
            )
        )
        return str(result.inserted_primary_key[0]), "no_invoice_expense"

    if doc_type == "loan_document":
        installments = data.get("installments") or []
        if not installments:
            raise ValueError("No se han detectado cuotas válidas en el documento del préstamo.")
        inserted_ids = []
        concept = (data.get("concept") or "Préstamo bancario").strip()
        for installment in installments:
            total_amount = parse_amount(installment.get("total_amount"))
            interest_amount = parse_amount(installment.get("interest_amount"))
            principal_amount = parse_amount(installment.get("principal_amount"))
            if total_amount is None or interest_amount is None:
                continue
            if principal_amount is None:
                principal_amount = round(total_amount - interest_amount, 2)
            result = conn.execute(
                loan_installments_table.insert().values(
                    user_id=data_owner_id,
                    company_id=row["company_id"],
                    bank_name=(installment.get("bank_name") or data.get("bank_name") or "").strip() or None,
                    concept=concept,
                    payment_date=normalize_date(installment.get("payment_date")),
                    payment_completed_dates=None,
                    total_amount=total_amount,
                    interest_amount=interest_amount,
                    principal_amount=principal_amount,
                    created_at=now,
                )
            )
            inserted_ids.append(str(result.inserted_primary_key[0]))
        if not inserted_ids:
            raise ValueError("No se han podido registrar cuotas del préstamo.")
        return ",".join(inserted_ids), "loan_installment"

    if doc_type == "tax_model":
        return str(row["id"]), "tax_filing"

    return None, "archive"


def document_duplicate_match(conn, company_id, detected_type, extracted_data):
    provider_or_client = (
        extracted_data.get("provider_name")
        or extracted_data.get("client_name")
        or extracted_data.get("supplier")
        or extracted_data.get("client")
    )
    invoice_number = extracted_data.get("invoice_number")
    tax_id = extracted_data.get("tax_id")
    invoice_date = extracted_data.get("invoice_date")
    total_amount = extracted_data.get("total_amount")
    rows = conn.execute(
        select(
            processed_documents_table.c.id,
            processed_documents_table.c.detected_document_type,
            processed_documents_table.c.original_extracted_data_json,
            processed_documents_table.c.corrected_data_json,
            processed_documents_table.c.validation_status,
        ).where(processed_documents_table.c.company_id == company_id)
    ).mappings().all()
    normalized_name = normalize_entity_name(provider_or_client or "")
    for row in rows:
        if row["detected_document_type"] != detected_type:
            continue
        comparison = json_loads_dict(row.get("corrected_data_json")) or json_loads_dict(
            row.get("original_extracted_data_json")
        )
        if (
            normalize_entity_name(
                comparison.get("provider_name")
                or comparison.get("client_name")
                or comparison.get("supplier")
                or comparison.get("client")
                or ""
            )
            == normalized_name
            and (comparison.get("tax_id") or None) == (tax_id or None)
            and (comparison.get("invoice_number") or None) == (invoice_number or None)
            and (comparison.get("invoice_date") or None) == (invoice_date or None)
            and round(float(comparison.get("total_amount") or 0), 2)
            == round(float(total_amount or 0), 2)
        ):
            return row["id"]
    return None


def refresh_document_batch_counters(conn, batch_id):
    rows = conn.execute(
        select(
            processed_documents_table.c.processing_status,
            processed_documents_table.c.validation_status,
        ).where(processed_documents_table.c.source_batch_id == batch_id)
    ).mappings().all()
    totals = {
        "total_documents": len(rows),
        "processed_documents": sum(1 for row in rows if row["processing_status"] == "processed"),
        "ready_documents": sum(1 for row in rows if row["validation_status"] == "ready_to_register"),
        "review_documents": sum(1 for row in rows if row["validation_status"] == "needs_review"),
        "duplicate_documents": sum(1 for row in rows if row["validation_status"] == "duplicate"),
        "failed_documents": sum(1 for row in rows if row["processing_status"] == "failed"),
    }
    totals["status"] = "processed" if totals["processed_documents"] + totals["failed_documents"] >= totals["total_documents"] else "processing"
    totals["updated_at"] = datetime.utcnow().isoformat()
    conn.execute(
        document_batches_table.update()
        .where(document_batches_table.c.id == batch_id)
        .values(**totals)
    )


def serialize_processed_document(row):
    original_data = json_loads_dict(row.get("original_extracted_data_json"))
    corrected_data = json_loads_dict(row.get("corrected_data_json"))
    effective_data = corrected_data or original_data
    return {
        "id": row["id"],
        "companyId": row["company_id"],
        "sourceBatchId": row["source_batch_id"],
        "originalFileName": row["original_filename"],
        "fileType": row.get("file_type"),
        "uploadedAt": row.get("uploaded_at"),
        "processingStatus": row.get("processing_status"),
        "detectedDocumentType": row.get("detected_document_type"),
        "confidenceScore": float(row["confidence_score"]) if row.get("confidence_score") is not None else None,
        "validationStatus": row.get("validation_status"),
        "issueType": row.get("issue_type"),
        "issueDescription": row.get("issue_description"),
        "period": row.get("period"),
        "linkedAccountingRecordId": row.get("linked_accounting_record_id"),
        "linkedAccountingRecordType": row.get("linked_accounting_record_type"),
        "approvedAt": row.get("approved_at"),
        "rejectedAt": row.get("rejected_at"),
        "registeredAt": row.get("registered_at"),
        "duplicateOfDocumentId": row.get("duplicate_of_document_id"),
        "fileUrl": row.get("file_url"),
        "extractedText": row.get("extracted_text"),
        "originalExtractedData": original_data,
        "correctedData": corrected_data,
        "effectiveData": effective_data,
        "auditLog": json_loads_list(row.get("audit_log_json")),
        "counterparty": effective_data.get("provider_name")
        or effective_data.get("client_name")
        or effective_data.get("employee_name")
        or effective_data.get("bank_name"),
        "invoiceDate": effective_data.get("invoice_date"),
        "baseAmount": effective_data.get("base_amount"),
        "vatAmount": effective_data.get("vat_amount"),
        "totalAmount": effective_data.get("total_amount") or effective_data.get("amount"),
    }


def replace_payment_date(values, previous_date, next_date):
    previous = normalize_date(previous_date)
    updated = normalize_date(next_date)
    dates = parse_payment_dates(values)
    if previous:
        dates = [item for item in dates if item != previous]
    if updated:
        dates.append(updated)
    return sorted(set(dates))


def is_trackable_payment_type(item_type):
    return item_type in {"expense", "no_invoice", "loan_installment"}


def resolve_payment_status(item_type, payment_date, completed_dates, today_iso=None):
    if not is_trackable_payment_type(item_type):
        return None
    normalized_date = normalize_date(payment_date)
    if not normalized_date:
        return None
    completed = set(parse_payment_dates(completed_dates))
    if normalized_date in completed:
        return "paid"
    today_value = today_iso or date.today().isoformat()
    if normalized_date < today_value:
        return "overdue"
    if normalized_date == today_value:
        return "due_today"
    return "pending"


def parse_tax_model_targets(raw_value):
    if not raw_value:
        return []
    if isinstance(raw_value, list):
        values = raw_value
    else:
        try:
            parsed = json.loads(raw_value) if isinstance(raw_value, str) else raw_value
        except (TypeError, json.JSONDecodeError):
            parsed = raw_value
        values = parsed if isinstance(parsed, list) else [parsed]
    normalized = []
    for item in values:
        if item is None:
            continue
        value = str(item).strip().upper()
        if value:
            normalized.append(value)
    return sorted(set(normalized))


def serialize_tax_model_targets(targets):
    return json.dumps(parse_tax_model_targets(targets))


def derive_invoice_profile(
    expense_category,
    vat_deductible=True,
    vat_amount=None,
    withholding_amount=0.0,
):
    category = (expense_category or "with_invoice").strip()
    withholding_amount = float(withholding_amount or 0)
    if category == "non_deductible":
        targets = []
        subtype = "non_deductible_invoice"
        pnl_bucket = "non_deductible_expense"
    else:
        targets = ["303"] if bool(vat_deductible) else []
        if withholding_amount > 0:
            targets.extend(["111", "190"])
            subtype = "professional_invoice"
        else:
            subtype = "supplier_invoice"
        pnl_bucket = "operating_expense"
    return {
        "expense_family": "supplier",
        "expense_subtype": subtype,
        "pnl_bucket": pnl_bucket,
        "tax_model_targets": serialize_tax_model_targets(targets),
    }


def get_invoice_deductible_amount(row):
    if not row or row.get("expense_category") == "non_deductible":
        return 0.0
    if row.get("vat_deductible"):
        return float(row.get("base_amount") or 0)
    return float(row.get("total_amount") or 0)


def derive_no_invoice_profile(expense_type, vat_deductible=False, withholding_amount=0.0):
    targets = []
    vat_deductible = bool(vat_deductible)
    withholding_amount = float(withholding_amount or 0)

    if expense_type == "nomina":
        targets = ["111", "190"]
        family = "personnel"
        subtype = "payroll"
        pnl_bucket = "personnel_expense"
    elif expense_type == "seguridad_social":
        family = "personnel"
        subtype = "social_security"
        pnl_bucket = "personnel_expense"
    elif expense_type == "alquiler_local":
        family = "rent"
        subtype = "local_rent"
        pnl_bucket = "operating_expense"
        if vat_deductible:
            targets.append("303")
        if withholding_amount > 0:
            targets.extend(["115", "180"])
    elif expense_type == "alquiler_cabina":
        family = "rent"
        subtype = "booth_rent"
        pnl_bucket = "operating_expense"
        if vat_deductible:
            targets.append("303")
        if withholding_amount > 0:
            targets.extend(["115", "180"])
    elif expense_type == "prestamo":
        family = "financing"
        subtype = "loan_installment"
        pnl_bucket = "financial_expense"
    elif expense_type == "amortizacion":
        family = "adjustment"
        subtype = "amortization"
        pnl_bucket = "amortization_expense"
    elif expense_type == "kilometraje":
        family = "adjustment"
        subtype = "mileage"
        pnl_bucket = "operating_expense"
    else:
        family = "adjustment"
        subtype = "other_adjustment"
        pnl_bucket = "operating_expense"
        if vat_deductible:
            targets.append("303")
        if withholding_amount > 0:
            targets.extend(["111", "190"])

    return {
        "expense_family": family,
        "expense_subtype": subtype,
        "pnl_bucket": pnl_bucket,
        "tax_model_targets": serialize_tax_model_targets(targets),
    }


def get_effective_invoice_tax_targets(row):
    tax_targets = parse_tax_model_targets(row.get("tax_model_targets"))
    if tax_targets:
        return tax_targets
    profile = derive_invoice_profile(
        row.get("expense_category"),
        row.get("vat_deductible"),
        row.get("vat_amount"),
        row.get("withholding_amount"),
    )
    return parse_tax_model_targets(profile.get("tax_model_targets"))


def get_effective_no_invoice_tax_targets(row):
    tax_targets = parse_tax_model_targets(row.get("tax_model_targets"))
    inferred_targets = parse_tax_model_targets(
        derive_no_invoice_profile(
            row.get("expense_type"),
            row.get("vat_deductible"),
            row.get("withholding_amount"),
        ).get("tax_model_targets")
    )
    if not tax_targets:
        return inferred_targets
    if "115" in inferred_targets and "115" not in tax_targets:
        return inferred_targets
    if "111" in inferred_targets and "111" not in tax_targets:
        return inferred_targets
    return tax_targets


def get_no_invoice_deductible_amount(row):
    expense_type = row.get("expense_type")
    if row.get("expense_family") == "financing" or expense_type == "prestamo":
        return float(row.get("interest_amount") or 0)
    if row.get("vat_deductible"):
        return float(row.get("base_amount") or row.get("amount") or 0)
    if not row.get("deductible"):
        return 0.0
    return float(row.get("amount") or 0)


def parse_loan_date(value):
    normalized = normalize_date(value)
    if normalized:
        return normalized
    if not value:
        return None
    raw = str(value).strip()
    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", raw)
    if not match:
        return None
    day, month, year = match.groups()
    if len(year) == 2:
        year = f"20{year}"
    try:
        return date(int(year), int(month), int(day)).isoformat()
    except ValueError:
        return None


def parse_amount(value):
    if value is None:
        return None
    raw = str(value).strip()
    if raw == "":
        return None
    raw = raw.replace("€", "").replace("EUR", "").replace("euro", "")
    raw = raw.replace(" ", "")
    if raw.count(",") >= 1 and raw.count(".") >= 1:
        raw = raw.replace(".", "").replace(",", ".")
    elif raw.count(",") == 1 and raw.count(".") == 0:
        raw = raw.replace(",", ".")
    elif raw.count(".") >= 1 and raw.count(",") == 0:
        parts = raw.split(".")
        if len(parts[-1]) <= 2:
            raw = "".join(parts[:-1]) + "." + parts[-1]
        else:
            raw = raw.replace(".", "")
    try:
        return float(raw)
    except ValueError:
        return None


def _choose_total_interest(amounts):
    amounts = [amount for amount in amounts if amount is not None and amount >= 0]
    if len(amounts) < 2:
        return None, None
    if len(amounts) == 2:
        total = max(amounts)
        interest = min(amounts)
        return total, interest
    for total in amounts:
        for interest in amounts:
            if total <= interest:
                continue
            principal = round(total - interest, 2)
            for candidate in amounts:
                if candidate in (total, interest):
                    continue
                if abs(candidate - principal) <= max(0.05, principal * 0.01):
                    return total, interest
    sorted_amounts = sorted(amounts)
    total = sorted_amounts[-2] if len(sorted_amounts) >= 2 else sorted_amounts[-1]
    interest = sorted_amounts[0]
    if total <= interest:
        return None, None
    return total, interest


def parse_loan_installments_from_text(text):
    installments = []
    if not text:
        return installments
    bank_name = None
    for line in text.splitlines():
        lowered = line.lower()
        if any(keyword in lowered for keyword in ["banco", "entidad", "bank"]):
            cleaned = re.sub(r"^(banco|entidad|bank)\s*[:\-]\s*", "", line, flags=re.I)
            cleaned = cleaned.strip()
            if cleaned and len(cleaned) > 2:
                bank_name = cleaned
                break
    for line in text.splitlines():
        date_value = parse_loan_date(line)
        if not date_value:
            continue
        numbers = re.findall(r"\d{1,3}(?:[.\s]\d{3})*(?:,\d{2})|\d+[.,]\d{2}", line)
        amounts = [parse_amount(value) for value in numbers]
        total, interest = _choose_total_interest(amounts)
        if total is None or interest is None:
            continue
        principal = round(total - interest, 2)
        installments.append(
            {
                "payment_date": date_value,
                "bank_name": bank_name,
                "total_amount": round(total, 2),
                "interest_amount": round(interest, 2),
                "principal_amount": round(principal, 2),
            }
        )
    return installments


def parse_loan_installments_from_excel(file_bytes):
    installments = []
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return installments
    headers = [str(value).strip().lower() if value else "" for value in rows[0]]
    date_idx = None
    total_idx = None
    interest_idx = None
    principal_idx = None
    bank_idx = None
    for idx, header in enumerate(headers):
        if "fecha" in header:
            date_idx = idx
        if "interes" in header or "interés" in header:
            interest_idx = idx
        if "cuota" in header or "total" in header or "importe" in header:
            total_idx = idx
        if "principal" in header or "amort" in header:
            principal_idx = idx
        if "banco" in header or "entidad" in header or "bank" in header:
            bank_idx = idx

    data_rows = rows[1:] if any(headers) else rows
    for row in data_rows:
        if date_idx is None or date_idx >= len(row):
            continue
        payment_date = parse_loan_date(row[date_idx])
        if not payment_date:
            continue
        total_amount = None
        interest_amount = None
        principal_amount = None
        if total_idx is not None and total_idx < len(row):
            total_amount = parse_amount(row[total_idx])
        if interest_idx is not None and interest_idx < len(row):
            interest_amount = parse_amount(row[interest_idx])
        if principal_idx is not None and principal_idx < len(row):
            principal_amount = parse_amount(row[principal_idx])

        if total_amount is None and principal_amount is not None and interest_amount is not None:
            total_amount = principal_amount + interest_amount
        if interest_amount is None and total_amount is not None and principal_amount is not None:
            interest_amount = total_amount - principal_amount
        if principal_amount is None and total_amount is not None and interest_amount is not None:
            principal_amount = total_amount - interest_amount

        if total_amount is None or interest_amount is None or principal_amount is None:
            continue
        if interest_amount < 0 or total_amount < 0:
            continue
        bank_name = None
        if bank_idx is not None and bank_idx < len(row):
            raw_bank = row[bank_idx]
            if raw_bank:
                bank_name = str(raw_bank).strip()
        installments.append(
            {
                "payment_date": payment_date,
                "bank_name": bank_name,
                "total_amount": round(total_amount, 2),
                "interest_amount": round(interest_amount, 2),
                "principal_amount": round(principal_amount, 2),
            }
        )
    return installments


def vat_rate_to_str(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_vat_breakdown(raw_value):
    if not raw_value:
        return []
    if isinstance(raw_value, list):
        values = raw_value
    else:
        value = raw_value
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return []
            try:
                parsed = json.loads(value)
                if isinstance(parsed, list):
                    values = parsed
                else:
                    values = [parsed]
            except json.JSONDecodeError:
                return []
        else:
            values = [value]

    lines = []
    for entry in values:
        if not isinstance(entry, dict):
            continue
        rate_raw = str(entry.get("rate") or entry.get("vat_rate") or entry.get("vat") or "").strip()
        rate = None
        if rate_raw:
            try:
                rate = float(rate_raw)
            except ValueError:
                rate = None
        if rate is not None and rate < 0:
            continue
        base_amount = parse_amount(str(entry.get("base") or entry.get("base_amount") or ""))
        vat_amount = parse_amount(str(entry.get("vat_amount") or entry.get("iva") or ""))
        total_amount = parse_amount(str(entry.get("total") or entry.get("total_amount") or ""))
        if base_amount is None and total_amount is None and vat_amount is None:
            continue
        if base_amount is None and total_amount is not None and vat_amount is not None:
            base_amount = round(total_amount - vat_amount, 2)
        if base_amount is None and total_amount is not None and vat_amount is None and rate is not None:
            base_amount = round(total_amount / (1 + rate / 100), 2)
        if base_amount is not None and vat_amount is None and total_amount is not None:
            vat_amount = round(total_amount - base_amount, 2)
        if base_amount is not None and vat_amount is None and rate is not None:
            vat_amount = round(base_amount * (rate / 100), 2)
        if base_amount is not None and total_amount is None and vat_amount is not None:
            total_amount = round(base_amount + vat_amount, 2)
        lines.append(
            {
                "rate": round(rate, 2) if rate is not None else None,
                "base": base_amount,
                "vat_amount": vat_amount,
                "total": total_amount,
            }
        )
    return lines


def summarize_vat_breakdown(lines):
    if not lines:
        return None
    base_total = 0.0
    vat_total = 0.0
    total_total = 0.0
    for line in lines:
        base_total += float(line.get("base") or 0)
        vat_total += float(line.get("vat_amount") or 0)
        total_total += float(line.get("total") or 0)
    return round(base_total, 2), round(vat_total, 2), round(total_total, 2)


def infer_vat_rate_from_breakdown(lines):
    if not lines:
        return None
    inferred_rates = set()
    for line in lines:
        if not isinstance(line, dict):
            continue
        rate = line.get("rate")
        if rate is None:
            base_amount = parse_amount(str(line.get("base") or ""))
            vat_amount = parse_amount(str(line.get("vat_amount") or ""))
            if base_amount is None or vat_amount is None:
                continue
            if abs(vat_amount) <= 0.02 and base_amount is not None:
                rate = 0.0
            elif base_amount:
                rate = round((vat_amount / base_amount) * 100, 2)
        if rate is None:
            continue
        matched = None
        for allowed in (0, 4, 10, 21):
            if abs(float(rate) - allowed) <= 0.25:
                matched = allowed
                break
        if matched is not None:
            inferred_rates.add(matched)
    if len(inferred_rates) == 1:
        return int(next(iter(inferred_rates)))
    if len(inferred_rates) > 1:
        return -1
    return None


def store_known_supplier(conn, user_id, company_id, supplier):
    if not supplier:
        return
    normalized = supplier.strip()
    if not normalized:
        return
    existing = conn.execute(
        select(known_suppliers_table.c.id)
        .where(known_suppliers_table.c.user_id == user_id)
        .where(known_suppliers_table.c.company_id == company_id)
        .where(func.lower(known_suppliers_table.c.name) == normalized.lower())
    ).first()
    if existing:
        return
    conn.execute(
        known_suppliers_table.insert().values(
            user_id=user_id,
            company_id=company_id,
            name=normalized,
            tax_id=None,
            confirmed_at=datetime.utcnow().isoformat(),
        )
    )


def fetch_known_suppliers(conn, user_id, company_id):
    rows = conn.execute(
        select(known_suppliers_table.c.name)
        .where(known_suppliers_table.c.user_id == user_id)
        .where(known_suppliers_table.c.company_id == company_id)
    ).scalars().all()
    return [name for name in rows if name]


def _parse_period_params():
    year = request.args.get("year") or request.args.get("anio") or request.args.get("año")
    quarter = request.args.get("quarter")
    start_month = request.args.get("start_month")
    end_month = request.args.get("end_month")
    try:
        year = int(year)
    except (TypeError, ValueError):
        year = None
    try:
        quarter = int(quarter) if quarter else None
    except (TypeError, ValueError):
        quarter = None
    try:
        start_month = int(start_month) if start_month else None
    except (TypeError, ValueError):
        start_month = None
    try:
        end_month = int(end_month) if end_month else None
    except (TypeError, ValueError):
        end_month = None
    return year, quarter, start_month, end_month


def _get_months_for_period(year, quarter=None, start_month=None, end_month=None):
    if not year:
        return []
    if quarter in {1, 2, 3, 4}:
        start = (quarter - 1) * 3 + 1
        return [start, start + 1, start + 2]
    if start_month and end_month and 1 <= start_month <= 12 and 1 <= end_month <= 12:
        if start_month <= end_month:
            return list(range(start_month, end_month + 1))
        return list(range(start_month, 13)) + list(range(1, end_month + 1))
    return list(range(1, 13))


def parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    text_value = str(value).strip()
    if not text_value:
        return None
    if "T" in text_value:
        text_value = text_value.split("T", 1)[0]
    try:
        return date.fromisoformat(text_value)
    except ValueError:
        return None


def date_in_range(value, start_date, end_date):
    parsed_value = parse_iso_date(value)
    if not parsed_value:
        return False
    return start_date <= parsed_value <= end_date


def period_label_from_dates(start_date, end_date):
    if not start_date or not end_date:
        return ""
    if start_date == end_date:
        return start_date.strftime("%d/%m/%Y")
    return f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"


def normalize_export_decimal(value):
    amount = round(float(value or 0), 2)
    return f"{amount:.2f}"


def csv_decimal(value):
    return normalize_export_decimal(value).replace(".", ",")


def normalize_export_text(value):
    if value is None:
        return ""
    return str(value).strip()


def format_export_account(account_code, label):
    code = normalize_export_text(account_code)
    account_label = normalize_export_text(label)
    return f"{code} {account_label}".strip()


def suggest_expense_account(*, expense_type=None, expense_family=None, expense_subtype=None, pnl_bucket=None):
    normalized_type = (expense_type or "").strip().lower()
    normalized_family = (expense_family or "").strip().lower()
    normalized_subtype = (expense_subtype or "").strip().lower()
    normalized_bucket = (pnl_bucket or "").strip().lower()

    mapping = {
        "alquiler_local": ("621", "Arrendamientos y cánones"),
        "alquiler_cabina": ("621", "Arrendamientos y cánones"),
        "nomina": ("640", "Sueldos y salarios"),
        "seguridad_social": ("642", "Seguridad Social a cargo de la empresa"),
        "amortizacion": ("681", "Amortización del inmovilizado material"),
        "kilometraje": ("629", "Otros servicios"),
        "prestamo": ("662", "Intereses de deudas"),
    }
    if normalized_type in mapping:
        return mapping[normalized_type]
    if normalized_family == "rent":
        return ("621", "Arrendamientos y cánones")
    if normalized_family == "personnel":
        return ("640", "Gastos de personal")
    if normalized_family == "financing" or normalized_bucket == "financial_expense":
        return ("662", "Intereses de deudas")
    if normalized_subtype == "amortization" or normalized_bucket == "amortization_expense":
        return ("681", "Amortización del inmovilizado material")
    return ("629", "Otros servicios")


def build_export_row_filename(prefix, company_name, extension, start_date, end_date):
    safe_company = secure_filename(normalize_export_text(company_name) or "empresa")
    safe_period = f"{start_date.isoformat()}_{end_date.isoformat()}"
    return f"{prefix}_{safe_company}_{safe_period}.{extension}"


def build_accounting_export_range():
    start_date = parse_iso_date(
        request.args.get("start_date")
        or request.args.get("startDate")
    )
    end_date = parse_iso_date(
        request.args.get("end_date")
        or request.args.get("endDate")
    )
    if start_date and end_date:
        if start_date > end_date:
            raise ValueError("La fecha inicial no puede ser posterior a la final.")
        return start_date, end_date, period_label_from_dates(start_date, end_date)

    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    period = (request.args.get("period") or "monthly").strip().lower()

    if year and month and 1 <= month <= 12:
        if period == "quarterly":
            months = _quarter_months_for_month(month)
            start_date = date(year, months[0], 1)
            end_date = date(year, months[-1], calendar.monthrange(year, months[-1])[1])
            return start_date, end_date, _report_period_label(year, months, _quarter_number_for_month(month))
        start_date = date(year, month, 1)
        end_date = date(year, month, calendar.monthrange(year, month)[1])
        return start_date, end_date, _report_period_label(year, [month])

    parsed_year, quarter, start_month, end_month = _parse_period_params()
    if parsed_year:
        months = _get_months_for_period(parsed_year, quarter, start_month, end_month)
        if months:
            start_date = date(parsed_year, months[0], 1)
            end_date = date(parsed_year, months[-1], calendar.monthrange(parsed_year, months[-1])[1])
            return start_date, end_date, _report_period_label(parsed_year, months, quarter)

    today = date.today()
    start_date = date(today.year, today.month, 1)
    end_date = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    return start_date, end_date, _report_period_label(today.year, [today.month])


def filter_manual_billing_rows_by_date(rows, start_date, end_date):
    filtered_rows = []
    for row in rows:
        invoice_date = parse_iso_date(row.get("invoice_date"))
        if not invoice_date:
            try:
                invoice_date = date(int(row.get("anio")), int(row.get("mes")), 1)
            except (TypeError, ValueError):
                invoice_date = None
        if invoice_date and start_date <= invoice_date <= end_date:
            filtered_rows.append(row)
    return filtered_rows


def filter_processed_documents_by_date(rows, start_date, end_date):
    filtered_rows = []
    for row in rows:
        payload = effective_document_data(row)
        reference_date = (
            parse_iso_date(payload.get("invoice_date"))
            or parse_iso_date(payload.get("expense_date"))
            or parse_iso_date(payload.get("payment_date"))
            or parse_iso_date(row.get("registered_at"))
            or parse_iso_date(row.get("created_at"))
        )
        if reference_date and start_date <= reference_date <= end_date:
            filtered_rows.append(row)
    return filtered_rows


def load_accounting_export_source_data(conn, data_owner_id, company_id, start_date, end_date):
    start_iso = start_date.isoformat()
    end_iso = end_date.isoformat()
    purchase_rows = conn.execute(
        select(invoices_table)
        .where(invoices_table.c.user_id == data_owner_id)
        .where(invoices_table.c.company_id == company_id)
        .where(invoices_table.c.invoice_date.between(start_iso, end_iso))
        .order_by(invoices_table.c.invoice_date.asc(), invoices_table.c.id.asc())
    ).mappings().all()
    income_rows = conn.execute(
        select(income_invoices_table)
        .where(income_invoices_table.c.user_id == data_owner_id)
        .where(income_invoices_table.c.company_id == company_id)
        .where(income_invoices_table.c.invoice_date.between(start_iso, end_iso))
        .order_by(income_invoices_table.c.invoice_date.asc(), income_invoices_table.c.id.asc())
    ).mappings().all()
    no_invoice_rows = conn.execute(
        select(no_invoice_table)
        .where(no_invoice_table.c.user_id == data_owner_id)
        .where(no_invoice_table.c.company_id == company_id)
        .where(no_invoice_table.c.expense_date.between(start_iso, end_iso))
        .order_by(no_invoice_table.c.expense_date.asc(), no_invoice_table.c.id.asc())
    ).mappings().all()
    loan_rows = conn.execute(
        select(loan_installments_table)
        .where(loan_installments_table.c.user_id == data_owner_id)
        .where(loan_installments_table.c.company_id == company_id)
        .where(loan_installments_table.c.payment_date.between(start_iso, end_iso))
        .order_by(loan_installments_table.c.payment_date.asc(), loan_installments_table.c.id.asc())
    ).mappings().all()

    manual_sales_rows = conn.execute(
        select(facturacion_table)
        .where(facturacion_table.c.user_id == data_owner_id)
        .where(facturacion_table.c.company_id == company_id)
        .where(facturacion_table.c.anio.between(start_date.year, end_date.year))
        .order_by(facturacion_table.c.anio.asc(), facturacion_table.c.mes.asc(), facturacion_table.c.id.asc())
    ).mappings().all()
    document_rows = conn.execute(
        select(processed_documents_table)
        .where(processed_documents_table.c.company_id == company_id)
        .where(processed_documents_table.c.validation_status == "registered")
        .order_by(processed_documents_table.c.created_at.asc(), processed_documents_table.c.id.asc())
    ).mappings().all()

    return {
        "purchase_invoices": purchase_rows,
        "income_invoices": income_rows,
        "no_invoice_expenses": no_invoice_rows,
        "loan_installments": loan_rows,
        "manual_sales": filter_manual_billing_rows_by_date(manual_sales_rows, start_date, end_date),
        "documents": filter_processed_documents_by_date(document_rows, start_date, end_date),
    }


def build_purchase_export_rows(source_data):
    rows = []
    for row in source_data.get("purchase_invoices", []):
        account_code, account_label = suggest_expense_account(
            expense_family=row.get("expense_family"),
            expense_subtype=row.get("expense_subtype"),
            pnl_bucket=row.get("pnl_bucket"),
        )
        rows.append(
            {
                "fecha": row.get("invoice_date"),
                "documento_tipo": "Factura recibida",
                "origen_tipo": "purchase_invoice",
                "origen_id": row.get("id"),
                "contraparte": row.get("supplier"),
                "concepto": row.get("original_filename"),
                "base": float(row.get("base_amount") or 0),
                "iva": float(row.get("vat_amount") or 0),
                "retencion": 0.0,
                "total": float(row.get("total_amount") or 0),
                "iva_deducible": "Sí" if row.get("vat_deductible") else "No",
                "cuenta_sugerida": format_export_account(account_code, account_label),
                "familia": row.get("expense_family") or "",
                "subtipo": row.get("expense_subtype") or "",
                "bucket_pyg": row.get("pnl_bucket") or "",
                "modelos_fiscales": ", ".join(parse_tax_model_targets(row.get("tax_model_targets"))),
            }
        )
    for row in source_data.get("no_invoice_expenses", []):
        account_code, account_label = suggest_expense_account(
            expense_type=row.get("expense_type"),
            expense_family=row.get("expense_family"),
            expense_subtype=row.get("expense_subtype"),
            pnl_bucket=row.get("pnl_bucket"),
        )
        rows.append(
            {
                "fecha": row.get("expense_date"),
                "documento_tipo": (row.get("expense_type") or "otro").replace("_", " ").title(),
                "origen_tipo": "no_invoice_expense",
                "origen_id": row.get("id"),
                "contraparte": row.get("payroll_employee_name") or row.get("concept"),
                "concepto": row.get("concept"),
                "base": float(row.get("base_amount") or row.get("amount") or 0),
                "iva": float(row.get("vat_amount") or 0),
                "retencion": float(row.get("withholding_amount") or 0),
                "total": float(row.get("amount") or 0),
                "iva_deducible": "Sí" if row.get("vat_deductible") else "No",
                "cuenta_sugerida": format_export_account(account_code, account_label),
                "familia": row.get("expense_family") or "",
                "subtipo": row.get("expense_subtype") or "",
                "bucket_pyg": row.get("pnl_bucket") or "",
                "modelos_fiscales": ", ".join(parse_tax_model_targets(row.get("tax_model_targets"))),
            }
        )
    rows.sort(key=lambda item: (item.get("fecha") or "", str(item.get("origen_id") or "")))
    return rows


def build_sales_export_rows(source_data):
    rows = []
    for row in source_data.get("income_invoices", []):
        rows.append(
            {
                "fecha": row.get("invoice_date"),
                "documento_tipo": "Factura emitida",
                "origen_tipo": "income_invoice",
                "origen_id": row.get("id"),
                "cliente": row.get("client"),
                "concepto": row.get("original_filename"),
                "base": float(row.get("base_amount") or 0),
                "iva": float(row.get("vat_amount") or 0),
                "total": float(row.get("total_amount") or 0),
                "tipo_iva": row.get("vat_rate") if row.get("vat_rate") is not None else "",
                "vencimiento": row.get("payment_date") or "",
                "estado_pago": "Planificado" if row.get("payment_date") else "",
            }
        )
    for row in source_data.get("manual_sales", []):
        invoice_date = row.get("invoice_date") or f"{int(row.get('anio')):04d}-{int(row.get('mes')):02d}-01"
        base_amount = float(row.get("base_facturada") or 0)
        vat_amount = float(row.get("iva_repercutido") or 0)
        rows.append(
            {
                "fecha": invoice_date,
                "documento_tipo": "Registro manual",
                "origen_tipo": "manual_billing",
                "origen_id": row.get("id"),
                "cliente": "",
                "concepto": row.get("concept") or "Facturación manual",
                "base": base_amount,
                "iva": vat_amount,
                "total": float(row.get("total_amount") or (base_amount + vat_amount)),
                "tipo_iva": row.get("tipo_iva") if row.get("tipo_iva") is not None else "",
                "vencimiento": "",
                "estado_pago": "",
            }
        )
    rows.sort(key=lambda item: (item.get("fecha") or "", str(item.get("origen_id") or "")))
    return rows


def append_journal_lines(container, entry_key, entry_date, concept, source_type, source_id, counterparty, lines):
    for index, line in enumerate(lines, start=1):
        debit = round(float(line.get("debe") or 0), 2)
        credit = round(float(line.get("haber") or 0), 2)
        if abs(debit) < 0.005 and abs(credit) < 0.005:
            continue
        container.append(
            {
                "asiento_id": entry_key,
                "linea": index,
                "fecha": entry_date,
                "diario": line.get("diario") or "GENERAL",
                "concepto": concept,
                "cuenta": line.get("cuenta") or "",
                "descripcion_cuenta": line.get("descripcion_cuenta") or "",
                "debe": debit,
                "haber": credit,
                "tercero": counterparty or "",
                "documento_origen": f"{source_type}:{source_id}",
                "origen_tipo": source_type,
                "origen_id": source_id,
            }
        )


def build_journal_export_rows(source_data):
    rows = []
    for row in source_data.get("purchase_invoices", []):
        entry_key = f"PUR-{row.get('id')}"
        base_amount = round(float(row.get("base_amount") or 0), 2)
        vat_amount = round(float(row.get("vat_amount") or 0), 2)
        total_amount = round(float(row.get("total_amount") or 0), 2)
        withholding_amount = round(float(row.get("withholding_amount") or 0), 2)
        expense_account, expense_label = suggest_expense_account(
            expense_family=row.get("expense_family"),
            expense_subtype=row.get("expense_subtype"),
            pnl_bucket=row.get("pnl_bucket"),
        )
        lines = []
        if row.get("vat_deductible") and vat_amount > 0:
            lines.append({"cuenta": expense_account, "descripcion_cuenta": expense_label, "debe": base_amount})
            lines.append({"cuenta": "472", "descripcion_cuenta": "Hacienda Pública, IVA soportado", "debe": vat_amount})
        else:
            lines.append({"cuenta": expense_account, "descripcion_cuenta": expense_label, "debe": total_amount})
        if withholding_amount > 0:
            lines.append({"cuenta": "4751", "descripcion_cuenta": "Hacienda Pública acreedora por retenciones practicadas", "haber": withholding_amount})
        creditor_amount = round(total_amount - withholding_amount, 2)
        if creditor_amount > 0:
            lines.append({"cuenta": "410", "descripcion_cuenta": "Acreedores por prestaciones de servicios", "haber": creditor_amount})
        append_journal_lines(
            rows,
            entry_key,
            row.get("invoice_date"),
            f"Factura proveedor {row.get('supplier') or row.get('original_filename')}",
            "purchase_invoice",
            row.get("id"),
            row.get("supplier"),
            lines,
        )

    for row in source_data.get("income_invoices", []):
        entry_key = f"SAL-{row.get('id')}"
        base_amount = round(float(row.get("base_amount") or 0), 2)
        vat_amount = round(float(row.get("vat_amount") or 0), 2)
        total_amount = round(float(row.get("total_amount") or 0), 2)
        append_journal_lines(
            rows,
            entry_key,
            row.get("invoice_date"),
            f"Factura emitida {row.get('client') or row.get('original_filename')}",
            "income_invoice",
            row.get("id"),
            row.get("client"),
            [
                {"cuenta": "430", "descripcion_cuenta": "Clientes", "debe": total_amount},
                {"cuenta": "700", "descripcion_cuenta": "Ventas de mercaderías / servicios", "haber": base_amount},
                {"cuenta": "477", "descripcion_cuenta": "Hacienda Pública, IVA repercutido", "haber": vat_amount},
            ],
        )

    for row in source_data.get("manual_sales", []):
        entry_key = f"MAN-{row.get('id')}"
        invoice_date = row.get("invoice_date") or f"{int(row.get('anio')):04d}-{int(row.get('mes')):02d}-01"
        base_amount = round(float(row.get("base_facturada") or 0), 2)
        vat_amount = round(float(row.get("iva_repercutido") or 0), 2)
        total_amount = round(float(row.get("total_amount") or (base_amount + vat_amount)), 2)
        append_journal_lines(
            rows,
            entry_key,
            invoice_date,
            row.get("concept") or "Facturación manual",
            "manual_billing",
            row.get("id"),
            "",
            [
                {"cuenta": "430", "descripcion_cuenta": "Clientes", "debe": total_amount},
                {"cuenta": "700", "descripcion_cuenta": "Ventas de mercaderías / servicios", "haber": base_amount},
                {"cuenta": "477", "descripcion_cuenta": "Hacienda Pública, IVA repercutido", "haber": vat_amount},
            ],
        )

    for row in source_data.get("no_invoice_expenses", []):
        entry_key = f"EXP-{row.get('id')}"
        expense_type = (row.get("expense_type") or "").strip().lower()
        amount = round(float(row.get("amount") or 0), 2)
        base_amount = round(float(row.get("base_amount") or amount), 2)
        vat_amount = round(float(row.get("vat_amount") or 0), 2)
        withholding_amount = round(float(row.get("withholding_amount") or 0), 2)
        expense_account, expense_label = suggest_expense_account(
            expense_type=expense_type,
            expense_family=row.get("expense_family"),
            expense_subtype=row.get("expense_subtype"),
            pnl_bucket=row.get("pnl_bucket"),
        )
        concept = row.get("concept") or expense_type or "Gasto"
        counterparty = row.get("payroll_employee_name") or concept
        lines = []
        if expense_type == "nomina":
            gross_amount = round(float(row.get("base_amount") or row.get("amount") or 0), 2)
            net_amount = round(float(row.get("payroll_net_amount") or 0), 2)
            deductions_amount = round(float(row.get("payroll_total_deductions_amount") or 0), 2)
            employer_cost_amount = round(float(row.get("payroll_employer_cost_amount") or 0), 2)
            employer_social_security = round(max(employer_cost_amount - gross_amount, 0), 2)
            lines.append({"cuenta": "640", "descripcion_cuenta": "Sueldos y salarios", "debe": gross_amount})
            if employer_social_security > 0:
                lines.append({"cuenta": "642", "descripcion_cuenta": "Seguridad Social a cargo de la empresa", "debe": employer_social_security})
            if net_amount > 0:
                lines.append({"cuenta": "465", "descripcion_cuenta": "Remuneraciones pendientes de pago", "haber": net_amount})
            if deductions_amount > 0:
                lines.append({"cuenta": "476", "descripcion_cuenta": "Organismos de la Seguridad Social acreedores", "haber": deductions_amount})
            journal_balance = round(
                sum(float(line.get("debe") or 0) for line in lines)
                - sum(float(line.get("haber") or 0) for line in lines),
                2,
            )
            if abs(journal_balance) >= 0.01:
                lines.append(
                    {
                        "cuenta": "410",
                        "descripcion_cuenta": "Acreedores varios",
                        "haber": journal_balance if journal_balance > 0 else 0,
                        "debe": abs(journal_balance) if journal_balance < 0 else 0,
                    }
                )
        elif expense_type == "seguridad_social":
            lines = [
                {"cuenta": "642", "descripcion_cuenta": "Seguridad Social a cargo de la empresa", "debe": amount},
                {"cuenta": "476", "descripcion_cuenta": "Organismos de la Seguridad Social acreedores", "haber": amount},
            ]
        elif expense_type == "prestamo":
            interest_amount = round(float(row.get("interest_amount") or 0), 2)
            principal_amount = round(max(amount - interest_amount, 0), 2)
            lines = [
                {"cuenta": "520", "descripcion_cuenta": "Deudas a corto plazo con entidades de crédito", "debe": principal_amount},
                {"cuenta": "662", "descripcion_cuenta": "Intereses de deudas", "debe": interest_amount},
                {"cuenta": "572", "descripcion_cuenta": "Bancos e instituciones de crédito c/c vista", "haber": amount},
            ]
        else:
            if row.get("vat_deductible") and vat_amount > 0:
                lines.append({"cuenta": expense_account, "descripcion_cuenta": expense_label, "debe": base_amount})
                lines.append({"cuenta": "472", "descripcion_cuenta": "Hacienda Pública, IVA soportado", "debe": vat_amount})
            else:
                lines.append({"cuenta": expense_account, "descripcion_cuenta": expense_label, "debe": amount})
            if withholding_amount > 0:
                lines.append({"cuenta": "4751", "descripcion_cuenta": "Hacienda Pública acreedora por retenciones practicadas", "haber": withholding_amount})
            creditor_amount = round(amount - withholding_amount, 2)
            if creditor_amount > 0:
                lines.append({"cuenta": "410", "descripcion_cuenta": "Acreedores por prestaciones de servicios", "haber": creditor_amount})
        append_journal_lines(rows, entry_key, row.get("expense_date"), concept, "no_invoice_expense", row.get("id"), counterparty, lines)

    for row in source_data.get("loan_installments", []):
        entry_key = f"LOA-{row.get('id')}"
        total_amount = round(float(row.get("total_amount") or 0), 2)
        interest_amount = round(float(row.get("interest_amount") or 0), 2)
        principal_amount = round(float(row.get("principal_amount") or max(total_amount - interest_amount, 0)), 2)
        append_journal_lines(
            rows,
            entry_key,
            row.get("payment_date"),
            row.get("concept") or "Cuota de préstamo",
            "loan_installment",
            row.get("id"),
            row.get("bank_name") or "",
            [
                {"cuenta": "520", "descripcion_cuenta": "Deudas a corto plazo con entidades de crédito", "debe": principal_amount},
                {"cuenta": "662", "descripcion_cuenta": "Intereses de deudas", "debe": interest_amount},
                {"cuenta": "572", "descripcion_cuenta": "Bancos e instituciones de crédito c/c vista", "haber": total_amount},
            ],
        )

    rows.sort(key=lambda item: (item.get("fecha") or "", item.get("asiento_id") or "", int(item.get("linea") or 0)))
    return rows


def build_document_manifest_rows(source_data):
    rows = []
    for row in source_data.get("documents", []):
        payload = effective_document_data(row)
        rows.append(
            {
                "document_id": row.get("id"),
                "archivo": row.get("original_filename"),
                "tipo_detectado": row.get("detected_document_type"),
                "estado": row.get("validation_status"),
                "referencia_contable_tipo": row.get("linked_accounting_record_type") or "",
                "referencia_contable_id": row.get("linked_accounting_record_id") or "",
                "fecha_referencia": payload.get("invoice_date")
                or payload.get("expense_date")
                or payload.get("payment_date")
                or (normalize_export_text(row.get("registered_at")).split("T", 1)[0]),
                "contraparte": payload.get("provider_name")
                or payload.get("counterparty_name")
                or payload.get("employee_name")
                or payload.get("concept")
                or "",
                "importe_total": float(
                    payload.get("total_amount")
                    or payload.get("amount")
                    or payload.get("payable_amount")
                    or 0
                ),
                "periodo": row.get("period") or "",
            }
        )
    return rows


def build_export_matrix(rows, ordered_columns):
    matrix = [ordered_columns]
    for row in rows:
        matrix.append([row.get(column, "") for column in ordered_columns])
    return matrix


def write_csv_export(rows, ordered_columns):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(ordered_columns)
    for row in rows:
        serialized = []
        for column in ordered_columns:
            value = row.get(column, "")
            if isinstance(value, bool):
                serialized.append("Sí" if value else "No")
            elif isinstance(value, (int, float)):
                serialized.append(csv_decimal(value))
            else:
                serialized.append(normalize_export_text(value))
        writer.writerow(serialized)
    return output.getvalue().encode("utf-8-sig")


def write_xlsx_export(rows, ordered_columns, sheet_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = (sheet_name or "Export")[:31]
    worksheet.append(ordered_columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in ordered_columns])
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 36)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def load_document_binary(row):
    storage_path = row.get("storage_path")
    file_url = row.get("file_url")
    if storage_path and os.path.exists(storage_path):
        with open(storage_path, "rb") as file_handle:
            return file_handle.read()
    if file_url and os.path.exists(file_url):
        with open(file_url, "rb") as file_handle:
            return file_handle.read()
    if file_url and file_url.startswith(("http://", "https://")):
        try:
            response = httpx.get(file_url, timeout=30.0)
            response.raise_for_status()
            return response.content
        except Exception:
            app.logger.warning("No se pudo descargar el documento %s desde %s", row.get("id"), file_url)
    return None


def build_accounting_export_package(source_data, metadata_payload):
    purchase_rows = build_purchase_export_rows(source_data)
    sales_rows = build_sales_export_rows(source_data)
    journal_rows = build_journal_export_rows(source_data)
    manifest_rows = build_document_manifest_rows(source_data)
    package_stream = io.BytesIO()
    with zipfile.ZipFile(package_stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "compras.csv",
            write_csv_export(
                purchase_rows,
                [
                    "fecha",
                    "documento_tipo",
                    "origen_tipo",
                    "origen_id",
                    "contraparte",
                    "concepto",
                    "base",
                    "iva",
                    "retencion",
                    "total",
                    "iva_deducible",
                    "cuenta_sugerida",
                    "familia",
                    "subtipo",
                    "bucket_pyg",
                    "modelos_fiscales",
                ],
            ),
        )
        archive.writestr(
            "ventas.csv",
            write_csv_export(
                sales_rows,
                [
                    "fecha",
                    "documento_tipo",
                    "origen_tipo",
                    "origen_id",
                    "cliente",
                    "concepto",
                    "base",
                    "iva",
                    "total",
                    "tipo_iva",
                    "vencimiento",
                    "estado_pago",
                ],
            ),
        )
        archive.writestr(
            "asientos.csv",
            write_csv_export(
                journal_rows,
                [
                    "asiento_id",
                    "linea",
                    "fecha",
                    "diario",
                    "concepto",
                    "cuenta",
                    "descripcion_cuenta",
                    "debe",
                    "haber",
                    "tercero",
                    "documento_origen",
                    "origen_tipo",
                    "origen_id",
                ],
            ),
        )
        archive.writestr(
            "manifest_documental.csv",
            write_csv_export(
                manifest_rows,
                [
                    "document_id",
                    "archivo",
                    "tipo_detectado",
                    "estado",
                    "referencia_contable_tipo",
                    "referencia_contable_id",
                    "fecha_referencia",
                    "contraparte",
                    "importe_total",
                    "periodo",
                ],
            ),
        )
        archive.writestr(
            "manifest.json",
            json.dumps(
                {
                    **metadata_payload,
                    "purchase_rows": len(purchase_rows),
                    "sales_rows": len(sales_rows),
                    "journal_rows": len(journal_rows),
                    "document_rows": len(manifest_rows),
                },
                ensure_ascii=False,
                indent=2,
            ),
        )
        for row in source_data.get("documents", []):
            file_bytes = load_document_binary(row)
            if not file_bytes:
                continue
            safe_name = secure_filename(row.get("original_filename") or f"documento_{row.get('id')}")
            archive.writestr(f"documentos/{int(row.get('id')):05d}_{safe_name}", file_bytes)
    package_stream.seek(0)
    return package_stream


def export_response_from_rows(rows, ordered_columns, *, export_format, sheet_name, download_name):
    if export_format == "xlsx":
        payload = write_xlsx_export(rows, ordered_columns, sheet_name)
        return send_file(
            io.BytesIO(payload),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )
    payload = write_csv_export(rows, ordered_columns)
    return send_file(
        io.BytesIO(payload),
        mimetype="text/csv",
        as_attachment=True,
        download_name=download_name,
    )


def _quarter_number_for_month(month):
    return ((int(month) - 1) // 3) + 1


def _quarter_months_for_month(month):
    start = (_quarter_number_for_month(month) - 1) * 3 + 1
    return [start, start + 1, start + 2]


def _periods_for_months(year, months):
    return [(int(year), int(month)) for month in months]


def _periods_for_previous_month(month, year):
    previous_month = month - 1
    previous_year = year
    if previous_month <= 0:
        previous_month = 12
        previous_year -= 1
    return [(previous_year, previous_month)]


def _quarter_periods_due_in_month(month, year):
    if month not in {1, 4, 7, 10}:
        return [], None
    if month == 1:
        target_year = year - 1
        target_months = [10, 11, 12]
    else:
        target_year = year
        target_months = [month - 3, month - 2, month - 1]
    quarter = _quarter_number_for_month(target_months[-1])
    return _periods_for_months(target_year, target_months), f"T{quarter} {target_year}"


def _model_202_periods_for_due_month(month, year):
    if month == 4:
        return _periods_for_months(year, [1, 2, 3]), "3 meses"
    if month == 10:
        return _periods_for_months(year, list(range(1, 10))), "9 meses"
    if month == 12:
        return _periods_for_months(year, list(range(1, 12))), "11 meses"
    return [], None


def _format_period_label(periods):
    if not periods:
        return ""
    if len(periods) == 1:
        target_year, target_month = periods[0]
        return f"{MONTH_LABELS_ES[target_month]} {target_year}"
    years = {item[0] for item in periods}
    months = [item[1] for item in periods]
    if len(periods) == 3 and len(years) == 1 and months == _quarter_months_for_month(months[-1]):
        quarter = _quarter_number_for_month(months[-1])
        return f"T{quarter} {next(iter(years))}"
    if len(years) == 1:
        return f"Ene-{MONTH_LABELS_ES[months[-1]][:3]} {next(iter(years))}"
    first_year, first_month = periods[0]
    last_year, last_month = periods[-1]
    return f"{MONTH_LABELS_ES[first_month][:3]} {first_year} - {MONTH_LABELS_ES[last_month][:3]} {last_year}"


def _fetch_company_fiscal_profile(conn, company_id):
    return conn.execute(
        select(
            companies_table.c.id,
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.company_type,
            companies_table.c.vat_regime,
            companies_table.c.tax_periodicity,
            companies_table.c.files_model_303,
            companies_table.c.files_model_111,
            companies_table.c.files_model_115,
            companies_table.c.files_model_130,
            companies_table.c.files_model_202,
        ).where(companies_table.c.id == company_id)
    ).mappings().first()


def _split_amount_across_dates(total_amount, payment_dates):
    normalized_dates = parse_payment_dates(payment_dates)
    if not normalized_dates:
        return []
    total_amount = round(float(total_amount or 0), 2)
    base_share = round(total_amount / len(normalized_dates), 2)
    scheduled = [base_share for _ in normalized_dates]
    diff = round(total_amount - sum(scheduled), 2)
    index = 0
    while abs(diff) >= 0.01 and scheduled:
        step = 0.01 if diff > 0 else -0.01
        scheduled[index] = round(scheduled[index] + step, 2)
        diff = round(diff - step, 2)
        index = (index + 1) % len(scheduled)
    return list(zip(normalized_dates, scheduled))


def _sum_amount_for_period_dates(total_amount, payment_dates, period_prefixes):
    scheduled = _split_amount_across_dates(total_amount, payment_dates)
    if not scheduled:
        return 0.0
    return round(
        sum(
            amount
            for scheduled_date, amount in scheduled
            if any(str(scheduled_date).startswith(prefix) for prefix in period_prefixes)
        ),
        2,
    )


def _build_financial_metrics(user_id, company_id, periods, conn):
    period_prefixes = [f"{period_year}-{period_month:02d}" for period_year, period_month in periods]
    daily_expense_totals = {}
    supplier_totals = {}

    metrics = {
        "income_base": 0.0,
        "income_vat": 0.0,
        "income_gross_total": 0.0,
        "expense_base": 0.0,
        "expense_vat": 0.0,
        "expense_gross_total": 0.0,
        "withholding_111": 0.0,
        "withholding_115": 0.0,
        "invoice_expenses": 0.0,
        "payroll_expenses": 0.0,
        "other_operating_expenses": 0.0,
        "amortization_expenses": 0.0,
        "loan_interest": 0.0,
        "supplier_totals": supplier_totals,
        "daily_expense_totals": daily_expense_totals,
    }

    for period_year, period_month in periods:
        prefix = f"{period_year}-{period_month:02d}"
        invoice_rows = conn.execute(
            select(
                invoices_table.c.invoice_date,
                invoices_table.c.payment_date,
                invoices_table.c.payment_dates,
                invoices_table.c.supplier,
                invoices_table.c.base_amount,
                invoices_table.c.total_amount,
                invoices_table.c.vat_deductible,
                invoices_table.c.vat_amount,
                invoices_table.c.withholding_amount,
                invoices_table.c.expense_category,
                invoices_table.c.tax_model_targets,
            )
            .where(invoices_table.c.user_id == user_id)
            .where(invoices_table.c.company_id == company_id)
            .where(invoices_table.c.invoice_date.like(f"{prefix}%"))
        ).mappings().all()
        for row in invoice_rows:
            gross_total = float(row.get("total_amount") or 0)
            deductible_amount = get_invoice_deductible_amount(row)
            withholding_amount = float(row.get("withholding_amount") or 0)
            tax_targets = get_effective_invoice_tax_targets(row)
            payment_dates = parse_payment_dates(row.get("payment_dates"))
            if not payment_dates:
                fallback = normalize_date(row.get("payment_date")) or normalize_date(
                    row.get("invoice_date")
                )
                payment_dates = [fallback] if fallback else []
            metrics["expense_gross_total"] += gross_total
            metrics["expense_base"] += deductible_amount
            metrics["invoice_expenses"] += deductible_amount
            if row.get("expense_category") != "non_deductible" and row.get("vat_deductible") is not False:
                metrics["expense_vat"] += float(row.get("vat_amount") or 0)
            if withholding_amount > 0 and payment_dates:
                amount_in_selected_period = _sum_amount_for_period_dates(
                    withholding_amount,
                    payment_dates,
                    period_prefixes,
                )
                if amount_in_selected_period > 0:
                    if "115" in tax_targets:
                        metrics["withholding_115"] += amount_in_selected_period
                    elif "111" in tax_targets:
                        metrics["withholding_111"] += amount_in_selected_period
            supplier_name = (row.get("supplier") or "Sin proveedor").strip() or "Sin proveedor"
            supplier_totals[supplier_name] = supplier_totals.get(supplier_name, 0.0) + gross_total
            invoice_date = normalize_date(row.get("invoice_date"))
            if invoice_date and invoice_date.startswith(prefix):
                invoice_day = int(invoice_date[-2:])
                daily_expense_totals[invoice_day] = daily_expense_totals.get(invoice_day, 0.0) + gross_total

        no_invoice_rows = conn.execute(
            select(
                no_invoice_table.c.amount,
                no_invoice_table.c.interest_amount,
                no_invoice_table.c.vat_deductible,
                no_invoice_table.c.vat_amount,
                no_invoice_table.c.base_amount,
                no_invoice_table.c.expense_type,
                no_invoice_table.c.deductible,
                no_invoice_table.c.withholding_amount,
                no_invoice_table.c.expense_family,
                no_invoice_table.c.expense_subtype,
                no_invoice_table.c.pnl_bucket,
                no_invoice_table.c.tax_model_targets,
                no_invoice_table.c.expense_date,
                no_invoice_table.c.payment_date,
                no_invoice_table.c.payment_dates,
                no_invoice_table.c.concept,
            )
            .where(no_invoice_table.c.user_id == user_id)
            .where(
                (no_invoice_table.c.company_id == company_id)
                | (no_invoice_table.c.company_id.is_(None))
            )
            .where(no_invoice_table.c.expense_date.like(f"{prefix}%"))
        ).mappings().all()
        for row in no_invoice_rows:
            gross_total = float(row.get("amount") or 0)
            deductible_amount = get_no_invoice_deductible_amount(row)
            expense_type = row.get("expense_type")
            tax_targets = get_effective_no_invoice_tax_targets(row)
            withholding_amount = float(row.get("withholding_amount") or 0)
            payment_dates = parse_payment_dates(row.get("payment_dates"))
            if not payment_dates:
                fallback_date = normalize_date(row.get("payment_date")) or normalize_date(row.get("expense_date"))
                payment_dates = [fallback_date] if fallback_date else []
            if withholding_amount > 0:
                amount_in_selected_period = _sum_amount_for_period_dates(
                    withholding_amount,
                    payment_dates,
                    period_prefixes,
                )
                if amount_in_selected_period > 0:
                    if "115" in tax_targets:
                        metrics["withholding_115"] += amount_in_selected_period
                    elif "111" in tax_targets:
                        metrics["withholding_111"] += amount_in_selected_period

            metrics["expense_gross_total"] += gross_total
            metrics["expense_base"] += deductible_amount
            if row.get("vat_deductible"):
                metrics["expense_vat"] += float(row.get("vat_amount") or 0)

            if expense_type in {"nomina", "seguridad_social"}:
                metrics["payroll_expenses"] += deductible_amount
            elif expense_type == "amortizacion":
                metrics["amortization_expenses"] += deductible_amount
            elif expense_type == "prestamo":
                metrics["loan_interest"] += deductible_amount
            else:
                metrics["other_operating_expenses"] += deductible_amount

            expense_date = normalize_date(row.get("expense_date"))
            if expense_date and expense_date.startswith(prefix):
                expense_day = int(expense_date[-2:])
                daily_expense_totals[expense_day] = daily_expense_totals.get(expense_day, 0.0) + gross_total
            supplier_name = (row.get("concept") or expense_type or "Sin concepto").strip() or "Sin concepto"
            supplier_totals[supplier_name] = supplier_totals.get(supplier_name, 0.0) + gross_total

        loan_rows = conn.execute(
            select(loan_installments_table.c.interest_amount)
            .where(loan_installments_table.c.user_id == user_id)
            .where(loan_installments_table.c.company_id == company_id)
            .where(loan_installments_table.c.payment_date.like(f"{prefix}%"))
        ).mappings().all()
        for row in loan_rows:
            interest_amount = float(row.get("interest_amount") or 0)
            metrics["expense_base"] += interest_amount
            metrics["loan_interest"] += interest_amount

        income_invoice_rows = conn.execute(
            select(income_invoices_table.c.base_amount, income_invoices_table.c.vat_amount)
            .where(income_invoices_table.c.user_id == user_id)
            .where(income_invoices_table.c.company_id == company_id)
            .where(income_invoices_table.c.invoice_date.like(f"{prefix}%"))
        ).mappings().all()
        for row in income_invoice_rows:
            base_amount = float(row.get("base_amount") or 0)
            vat_amount = float(row.get("vat_amount") or 0)
            metrics["income_base"] += base_amount
            metrics["income_vat"] += vat_amount
            metrics["income_gross_total"] += base_amount + vat_amount

        billing_rows = conn.execute(
            select(
                facturacion_table.c.base_facturada,
                facturacion_table.c.iva_repercutido,
            )
            .where(facturacion_table.c.user_id == user_id)
            .where(facturacion_table.c.company_id == company_id)
            .where(facturacion_table.c.anio == period_year)
            .where(facturacion_table.c.mes == period_month)
        ).mappings().all()
        for row in billing_rows:
            base_amount = float(row.get("base_facturada") or 0)
            vat_amount = float(row.get("iva_repercutido") or 0)
            metrics["income_base"] += base_amount
            metrics["income_vat"] += vat_amount
            metrics["income_gross_total"] += base_amount + vat_amount

    metrics["income_base"] = round(metrics["income_base"], 2)
    metrics["income_vat"] = round(metrics["income_vat"], 2)
    metrics["income_gross_total"] = round(metrics["income_gross_total"], 2)
    metrics["expense_base"] = round(metrics["expense_base"], 2)
    metrics["expense_vat"] = round(metrics["expense_vat"], 2)
    metrics["expense_gross_total"] = round(metrics["expense_gross_total"], 2)
    metrics["invoice_expenses"] = round(metrics["invoice_expenses"], 2)
    metrics["payroll_expenses"] = round(metrics["payroll_expenses"], 2)
    metrics["other_operating_expenses"] = round(metrics["other_operating_expenses"], 2)
    metrics["amortization_expenses"] = round(metrics["amortization_expenses"], 2)
    metrics["loan_interest"] = round(metrics["loan_interest"], 2)
    metrics["withholding_111"] = round(metrics["withholding_111"], 2)
    metrics["withholding_115"] = round(metrics["withholding_115"], 2)
    metrics["net_result"] = round(metrics["income_base"] - metrics["expense_base"], 2)
    metrics["vat_result"] = round(metrics["income_vat"] - metrics["expense_vat"], 2)
    metrics["model_130_estimate"] = round(max(metrics["net_result"], 0) * 0.20, 2)
    metrics["corporate_tax_estimate"] = round(max(metrics["net_result"], 0) * 0.25, 2)
    metrics["model_202_estimate"] = round(max(metrics["net_result"], 0) * 0.18, 2)
    metrics["supplier_totals"] = {
        key: round(value, 2) for key, value in sorted(supplier_totals.items(), key=lambda item: item[0])
    }
    metrics["daily_expense_totals"] = {
        day: round(value, 2) for day, value in sorted(daily_expense_totals.items())
    }
    return metrics


def _build_tax_model_metrics(user_id, company_id, periods, conn):
    metrics = _build_financial_metrics(user_id, company_id, periods, conn)
    return {
        "income_base": metrics["income_base"],
        "income_vat": metrics["income_vat"],
        "expense_base": metrics["expense_base"],
        "expense_vat": metrics["expense_vat"],
        "net_result": metrics["net_result"],
        "vat_result": metrics["vat_result"],
        "withholding_111": metrics["withholding_111"],
        "withholding_115": metrics["withholding_115"],
        "model_130_estimate": metrics["model_130_estimate"],
        "corporate_tax_estimate": metrics["corporate_tax_estimate"],
        "model_202_estimate": metrics["model_202_estimate"],
    }


def _build_fiscal_model_rows(conn, user_id, company_id, selected_month, selected_year, selected_period):
    company = _fetch_company_fiscal_profile(conn, company_id)
    if not company:
        return []

    selected_months = (
        _quarter_months_for_month(selected_month)
        if selected_period == "quarterly"
        else [selected_month]
    )
    selected_periods = _periods_for_months(selected_year, selected_months)
    ytd_periods = _periods_for_months(selected_year, range(1, selected_month + 1))
    selected_metrics = _build_tax_model_metrics(user_id, company_id, selected_periods, conn)
    ytd_metrics = _build_tax_model_metrics(user_id, company_id, ytd_periods, conn)

    selected_label = _format_period_label(selected_periods)
    annual_label = f"Acumulado anual hasta {MONTH_LABELS_ES[selected_month]} {selected_year}"
    periodicity_label = "Mensual" if company.get("tax_periodicity") == "monthly" else "Trimestral"

    rows = []
    if company.get("files_model_303") is not False and company.get("vat_regime") != "exempt":
        rows.append(
            {
                "model": "303",
                "periodicity": periodicity_label,
                "status": f"Liquidación estimada {selected_label}",
                "amount": selected_metrics["vat_result"],
            }
        )
        rows.append(
            {
                "model": "390",
                "periodicity": "Anual",
                "status": annual_label,
                "amount": ytd_metrics["vat_result"],
            }
        )

    if company.get("files_model_111"):
        rows.append(
            {
                "model": "111",
                "periodicity": periodicity_label,
                "status": f"Retenciones acumuladas {selected_label}",
                "amount": selected_metrics["withholding_111"],
            }
        )
        rows.append(
            {
                "model": "190",
                "periodicity": "Anual",
                "status": annual_label,
                "amount": ytd_metrics["withholding_111"],
            }
        )

    if company.get("files_model_115"):
        rows.append(
            {
                "model": "115",
                "periodicity": periodicity_label,
                "status": f"Retenciones alquiler acumuladas {selected_label}",
                "amount": selected_metrics["withholding_115"],
            }
        )
        rows.append(
            {
                "model": "180",
                "periodicity": "Anual",
                "status": annual_label,
                "amount": ytd_metrics["withholding_115"],
            }
        )

    if company.get("company_type") == "individual" and company.get("files_model_130"):
        rows.append(
            {
                "model": "130",
                "periodicity": "Trimestral",
                "status": f"Pago fraccionado estimado {selected_label}",
                "amount": selected_metrics["model_130_estimate"],
            }
        )

    if company.get("company_type") == "company":
        rows.append(
            {
                "model": "200",
                "periodicity": "Anual",
                "status": annual_label,
                "amount": ytd_metrics["corporate_tax_estimate"],
            }
        )
        if company.get("files_model_202"):
            rows.append(
                {
                    "model": "202",
                    "periodicity": "Abr / Oct / Dic",
                    "status": f"Pago fraccionado estimado a {MONTH_LABELS_ES[selected_month]} {selected_year}",
                    "amount": ytd_metrics["model_202_estimate"],
                }
            )

    return rows


def _build_fiscal_calendar_items(conn, user_id, company_id, target_month, target_year):
    company = _fetch_company_fiscal_profile(conn, company_id)
    if not company:
        return []

    items = []
    due_date = date(target_year, target_month, 20).isoformat()
    filings_lookup = _fetch_registered_tax_filings(conn, company_id)

    def add_item(model, label, amount, period_key):
        amount = round(float(amount or 0), 2)
        filing = filings_lookup.get((str(model), normalize_tax_period_key(period_key)))
        if filing:
            filing_status = filing.get("filing_status") or "a_ingresar"
            if filing_status == "sin_actividad":
                amount = 0.0
                calendar_impact = "informational"
            elif filing_status == "a_compensar":
                amount = round(float(filing.get("offset_amount") or filing.get("amount") or 0), 2)
                calendar_impact = "credit"
            elif filing_status == "a_devolver":
                amount = round(float(filing.get("refund_amount") or filing.get("amount") or 0), 2)
                calendar_impact = "refund"
            else:
                amount = round(
                    float(filing.get("payable_amount") or filing.get("amount") or amount or 0),
                    2,
                )
                calendar_impact = "payment"
        else:
            filing_status = "estimated"
            calendar_impact = "payment"
        if amount <= 0 and not filing:
            return
        items.append(
            {
                "id": f"{model}-{target_year}-{target_month}",
                "counterparty": "AEAT",
                "concept": label,
                "payment_date": due_date,
                "payment_dates": [due_date],
                "invoice_date": due_date,
                "base_amount": amount,
                "vat_rate": None,
                "vat_amount": None,
                "total_amount": amount,
                "amount": amount,
                "type": "tax_obligation",
                "tax_model": model,
                "tax_period": period_key,
                "filing_status": filing_status,
                "calendar_impact": calendar_impact,
                "filing_document_id": filing.get("document_id") if filing else None,
                "filing_registered_at": filing.get("registered_at") if filing else None,
                "is_filed": bool(filing),
            }
        )

    if company.get("tax_periodicity") == "monthly":
        monthly_periods = _periods_for_previous_month(target_month, target_year)
        monthly_metrics = _build_tax_model_metrics(user_id, company_id, monthly_periods, conn)
        monthly_label = _format_period_label(monthly_periods)
        if company.get("files_model_303") is not False and company.get("vat_regime") != "exempt":
            add_item("303", f"Modelo 303 · {monthly_label}", monthly_metrics["vat_result"], monthly_label)
        if company.get("files_model_111"):
            add_item("111", f"Modelo 111 · {monthly_label}", monthly_metrics["withholding_111"], monthly_label)
        if company.get("files_model_115"):
            add_item("115", f"Modelo 115 · {monthly_label}", monthly_metrics["withholding_115"], monthly_label)

    quarter_periods, quarter_label = _quarter_periods_due_in_month(target_month, target_year)
    if quarter_periods and quarter_label:
        quarter_metrics = _build_tax_model_metrics(user_id, company_id, quarter_periods, conn)
        if company.get("tax_periodicity") != "monthly":
            if company.get("files_model_303") is not False and company.get("vat_regime") != "exempt":
                add_item("303", f"Modelo 303 · {quarter_label}", quarter_metrics["vat_result"], quarter_label)
            if company.get("files_model_111"):
                add_item("111", f"Modelo 111 · {quarter_label}", quarter_metrics["withholding_111"], quarter_label)
            if company.get("files_model_115"):
                add_item("115", f"Modelo 115 · {quarter_label}", quarter_metrics["withholding_115"], quarter_label)
        if company.get("company_type") == "individual" and company.get("files_model_130"):
            add_item("130", f"Modelo 130 · {quarter_label}", quarter_metrics["model_130_estimate"], quarter_label)

    if company.get("company_type") == "company" and company.get("files_model_202"):
        model_202_periods, model_202_label = _model_202_periods_for_due_month(target_month, target_year)
        if model_202_periods and model_202_label:
            model_202_metrics = _build_tax_model_metrics(user_id, company_id, model_202_periods, conn)
            add_item(
                "202",
                f"Modelo 202 · {model_202_label}",
                model_202_metrics["model_202_estimate"],
                f"{model_202_label} {target_year}",
            )

    return items


def _build_report_totals(user_id, company_id, months, year):
    with engine.connect() as conn:
        metrics = _build_financial_metrics(
            user_id,
            company_id,
            _periods_for_months(year, months),
            conn,
        )

    return {
        "income_base": metrics["income_base"],
        "income_vat": metrics["income_vat"],
        "expense_base": metrics["expense_base"],
        "expense_vat": metrics["expense_vat"],
        "net_result": metrics["net_result"],
        "vat_result": metrics["vat_result"],
    }


def _report_period_label(year, months, quarter=None):
    if quarter in {1, 2, 3, 4}:
        return f"T{quarter} {year}"
    if months:
        if len(months) == 12:
            return f"Año {year}"
        if len(months) == 1:
            return f"{calendar.month_name[months[0]]} {year}"
        return f"{calendar.month_name[months[0]]} - {calendar.month_name[months[-1]]} {year}"
    return str(year)


def _empty_extracted(analysis_status="ok"):
    return {
        "analysis_status": analysis_status,
        "provider_name": None,
        "invoice_date": None,
        "payment_dates": [],
        "payment_date": None,
        "base_amount": None,
        "vat_rate": None,
        "vat_amount": None,
        "total_amount": None,
        "analysis_text": "",
        "validation": {"is_consistent": None, "difference": None},
    }


def _pdf_has_text(data: bytes, min_chars: int = 100) -> bool:
    try:
        with fitz.open(stream=data, filetype="pdf") as doc:
            total = 0
            for idx, page in enumerate(doc):
                text = page.get_text("text") or ""
                total += len(text.strip())
                if total >= min_chars:
                    return True
                if idx >= 1:
                    break
    except Exception as exc:
        app.logger.warning("No se pudo verificar texto embebido en PDF: %s", exc)
        return False
    return False


def _analysis_worker(
    file_bytes,
    filename,
    mime_type,
    document_type,
    company_names,
    known_suppliers,
    queue,
):
    try:
        result = analyze_invoice(
            file_bytes=file_bytes,
            filename=filename,
            mime_type=mime_type,
            document_type=document_type,
            company_names=company_names,
            known_suppliers=known_suppliers,
        )
        queue.put(result)
    except Exception as exc:
        queue.put({"__error__": str(exc)})


def _analyze_invoice_with_timeout(
    file_bytes,
    filename,
    stored_name,
    mime_type,
    document_type="expense",
    company_names=None,
    known_suppliers=None,
    fallback_status=None,
):
    ctx = mp.get_context("spawn")
    queue = ctx.Queue(1)
    process = ctx.Process(
        target=_analysis_worker,
        args=(
            file_bytes,
            filename,
            mime_type,
            document_type,
            company_names or [],
            known_suppliers or [],
            queue,
        ),
    )
    process.start()
    process.join(ANALYSIS_TIMEOUT_SECONDS)

    if process.is_alive():
        process.terminate()
        process.join()
        app.logger.warning(
            "Timeout analizando %s (> %ss). Se pasa a modo manual.",
            stored_name,
            ANALYSIS_TIMEOUT_SECONDS,
        )
        return _empty_extracted(fallback_status or "ok")

    if process.exitcode != 0:
        app.logger.warning(
            "Analisis fallido para %s (exitcode %s). Se pasa a modo manual.",
            stored_name,
            process.exitcode,
        )
        return _empty_extracted(fallback_status or "ok")

    try:
        result = queue.get_nowait()
    except Exception:
        app.logger.warning(
            "Analisis sin resultado para %s. Se pasa a modo manual.",
            stored_name,
        )
        return _empty_extracted(fallback_status or "ok")

    if not isinstance(result, dict) or result.get("__error__"):
        app.logger.warning(
            "Analisis con error para %s. Se pasa a modo manual.",
            stored_name,
        )
        return _empty_extracted(fallback_status or "ok")

    return result


def _normalize_email(value):
    return (value or "").strip().lower()


def send_email(to_email, subject, html_content, reply_to=None, from_email=None):
    if not RESEND_API_KEY:
        app.logger.warning("RESEND_API_KEY no configurada. Email no enviado.")
        return False
    sender = (from_email or APP_FROM_EMAIL or "").strip()
    sender_lower = sender.lower()
    if (
        not sender
        or sender_lower == "soporte@ledged.app"
        or sender_lower.endswith("<soporte@ledged.app>")
    ):
        app.logger.warning("APP_FROM_EMAIL inválido para envíos automáticos.")
        return False
    payload = {
        "from": sender,
        "to": [to_email],
        "subject": subject,
        "html": html_content,
    }
    if reply_to:
        payload["reply_to"] = reply_to
    try:
        response = httpx.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
            json=payload,
            timeout=20.0,
        )
        response.raise_for_status()
        return True
    except Exception:
        app.logger.exception("Error enviando email a %s", to_email)
        return False


def get_agency_email_for_user(user_id):
    with engine.connect() as conn:
        agency_id = conn.execute(
            select(users_table.c.agency_id).where(users_table.c.id == user_id)
        ).scalar_one_or_none()
        target_id = agency_id or user_id
        email = conn.execute(
            select(users_table.c.email).where(users_table.c.id == target_id)
        ).scalar_one_or_none()
    return email


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        invited = (request.args.get("invited") or "").strip().lower()
        message = "Cuenta activada. Ya puedes iniciar sesión." if invited == "1" else None
        return render_template("login.html", message=message)
    payload = request.form or request.get_json(silent=True) or {}
    email = _normalize_email(payload.get("email"))
    password = payload.get("password") or ""
    if not email or not password:
        return render_template("login.html", error="Email y contraseña obligatorios.")
    with engine.connect() as conn:
        row = conn.execute(
            select(
                users_table.c.id,
                users_table.c.email,
                users_table.c.password_hash,
                users_table.c.role,
                users_table.c.plan,
                users_table.c.is_active,
            ).where(users_table.c.email == email)
        ).mappings().first()
    if not row or not row["is_active"]:
        return render_template("login.html", error="Credenciales inválidas.")
    if not check_password_hash(row["password_hash"], password):
        return render_template("login.html", error="Credenciales inválidas.")
    session["user_id"] = row["id"]
    if row["role"] in {"agency", "staff"}:
        agency_id = row["id"] if row["role"] == "agency" else row.get("agency_id")
        if agency_id:
            with engine.begin() as conn:
                conn.execute(
                    agencies_table.update()
                    .where(agencies_table.c.id == agency_id)
                    .values(last_login_at=datetime.utcnow().isoformat())
                )
    return redirect(url_for("app_home"))


@app.route("/register", methods=["GET", "POST"])
def register():
    message = "Altas cerradas temporalmente. El acceso se concede solo de forma manual o por invitacion."
    if request.method == "POST":
        if request.is_json:
            return jsonify({"ok": False, "errors": [message]}), 403
        return render_template("register.html", registration_closed=True, error=message), 403
    return render_template("register.html", registration_closed=True), 403


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/reset", methods=["GET", "POST"])
@app.route("/reset-password", methods=["GET", "POST"])
def reset_password_request():
    if request.method == "GET":
        return render_template("reset_request.html")
    payload = request.form or request.get_json(silent=True) or {}
    email = _normalize_email(payload.get("email"))
    if not email:
        return render_template("reset_request.html", error="Email obligatorio.")
    with engine.connect() as conn:
        row = conn.execute(
            select(users_table.c.id, users_table.c.email)
            .where(users_table.c.email == email)
            .where(users_table.c.is_active.is_(True))
        ).mappings().first()
    if row:
        token = secrets.token_urlsafe(32)
        expires_at = (datetime.utcnow() + timedelta(hours=1)).isoformat()
        with engine.begin() as conn:
            conn.execute(
                password_resets_table.insert().values(
                    user_id=row["id"],
                    token=token,
                    expires_at=expires_at,
                    used_at=None,
                )
            )
        reset_link = url_for("reset_password", token=token, _external=True)
        html = f"""
        <p>Has solicitado restablecer tu contraseña.</p>
        <p>Enlace válido durante 1 hora:</p>
        <p><a href="{reset_link}">Restablecer contraseña</a></p>
        """
        reply_to = get_agency_email_for_user(row["id"])
        send_email(
            email,
            "Restablece tu contraseña",
            html,
            reply_to=reply_to,
            from_email=get_access_from_email(),
        )
    return render_template(
        "reset_request.html",
        message="Si el email existe, recibirás un enlace de recuperación.",
    )


@app.route("/reset/<token>", methods=["GET", "POST"])
@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    if request.method == "GET":
        return render_template("reset_password.html", token=token)
    payload = request.form or request.get_json(silent=True) or {}
    password = payload.get("password") or ""
    if len(password) < 8:
        return render_template(
            "reset_password.html",
            token=token,
            error="La contraseña debe tener al menos 8 caracteres.",
        )
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        reset_row = conn.execute(
            select(
                password_resets_table.c.id,
                password_resets_table.c.user_id,
                password_resets_table.c.expires_at,
                password_resets_table.c.used_at,
            ).where(password_resets_table.c.token == token)
        ).mappings().first()
        if (
            not reset_row
            or reset_row["used_at"]
            or reset_row["expires_at"] < now
        ):
            return render_template(
                "reset_password.html",
                token=token,
                error="El enlace no es válido o ha caducado.",
            )
        conn.execute(
            users_table.update()
            .where(users_table.c.id == reset_row["user_id"])
            .values(password_hash=generate_password_hash(password))
        )
        conn.execute(
            password_resets_table.update()
            .where(password_resets_table.c.id == reset_row["id"])
            .values(used_at=now)
        )
    return redirect(url_for("login"))


@app.route("/invite/<token>", methods=["GET", "POST"])
def accept_invitation(token):
    now = datetime.utcnow().isoformat()
    with engine.connect() as conn:
        invitation = conn.execute(
            select(user_invitations_table).where(user_invitations_table.c.token == token)
        ).mappings().first()
    if (
        not invitation
        or invitation.get("accepted_at")
        or invitation.get("expires_at", "") < now
    ):
        return render_template(
            "invite_accept.html",
            invitation=None,
            error="La invitación no es válida o ha caducado.",
        )

    if request.method == "GET":
        return render_template("invite_accept.html", invitation=invitation)

    password = request.form.get("password") or ""
    if len(password) < 8:
        return render_template(
            "invite_accept.html",
            invitation=invitation,
            error="La contraseña debe tener al menos 8 caracteres.",
        )

    created_at = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        fresh_invitation = conn.execute(
            select(user_invitations_table).where(user_invitations_table.c.token == token)
        ).mappings().first()
        if (
            not fresh_invitation
            or fresh_invitation.get("accepted_at")
            or fresh_invitation.get("expires_at", "") < created_at
        ):
            return render_template(
                "invite_accept.html",
                invitation=None,
                error="La invitación no es válida o ha caducado.",
            )

        existing_user = conn.execute(
            select(users_table.c.id).where(users_table.c.email == fresh_invitation["email"])
        ).first()
        if existing_user:
            conn.execute(
                user_invitations_table.update()
                .where(user_invitations_table.c.id == fresh_invitation["id"])
                .values(accepted_at=created_at)
            )
            return redirect(url_for("login", invited="1"))

        if fresh_invitation["role"] == "agency":
            plan = (fresh_invitation.get("plan") or "starter").strip().lower()
            trial_ends = (datetime.utcnow() + timedelta(days=14)).isoformat()
            result = conn.execute(
                users_table.insert().values(
                    email=fresh_invitation["email"],
                    password_hash=generate_password_hash(password),
                    role="agency",
                    plan="trial",
                    agency_id=None,
                    created_at=created_at,
                    is_active=True,
                )
            )
            new_user_id = result.inserted_primary_key[0]
            conn.execute(
                users_table.update()
                .where(users_table.c.id == new_user_id)
                .values(agency_id=new_user_id)
            )
            conn.execute(
                agencies_table.insert().values(
                    id=new_user_id,
                    name=fresh_invitation.get("name") or fresh_invitation["email"],
                    email=fresh_invitation["email"],
                    phone=None,
                    plan=plan if plan in {"starter", "pro", "advanced"} else "starter",
                    status="trial",
                    stripe_customer_id=None,
                    stripe_subscription_id=None,
                    stripe_price_id=None,
                    stripe_subscription_status=None,
                    stripe_current_period_end=None,
                    trial_ends_at=trial_ends,
                    created_at=created_at,
                    last_login_at=None,
                )
            )
        else:
            agency_id = fresh_invitation.get("agency_id")
            agency_plan = conn.execute(
                select(users_table.c.plan).where(users_table.c.id == agency_id)
            ).scalar_one_or_none()
            agency_limits = get_agency_limits_and_usage(conn, agency_id)
            if agency_limits["usage"]["staff"] >= agency_limits["limits"]["staff"]:
                return render_template(
                    "invite_accept.html",
                    invitation=fresh_invitation,
                    error=get_limit_error("staff", agency_limits["limits"]["staff"]),
                )
            conn.execute(
                users_table.insert().values(
                    email=fresh_invitation["email"],
                    password_hash=generate_password_hash(password),
                    role="staff",
                    plan=agency_plan or "trial",
                    agency_id=agency_id,
                    created_at=created_at,
                    is_active=True,
                )
            )

        conn.execute(
            user_invitations_table.update()
            .where(user_invitations_table.c.id == fresh_invitation["id"])
            .values(accepted_at=created_at)
        )
    return redirect(url_for("login", invited="1"))


@app.route("/")
def landing():
    if g.current_user:
        return redirect(url_for("app_home"))
    return render_template("landing.html")


@app.route("/landing")
def landing_alias():
    if g.current_user:
        return redirect(url_for("app_home"))
    return render_template("landing.html")


@app.route("/app")
def app_home():
    return render_template(
        "index.html",
        user=g.current_user,
        account_context=get_account_context_for_user(g.current_user),
        billing_context=get_billing_context_for_user(g.current_user),
        billing_message=get_billing_message(),
    )


@app.route("/billing/checkout", methods=["POST"])
def start_stripe_checkout():
    user = g.current_user or {}
    if user.get("role") != "agency":
        return redirect(url_for("app_home"))
    if not stripe_is_configured():
        return redirect(url_for("app_home", billing_error="not_configured"))
    plan = (request.form.get("plan") or "").strip().lower()
    billing_period = (request.form.get("billing_period") or "monthly").strip().lower()
    if plan not in {"starter", "pro", "advanced"}:
        return redirect(url_for("app_home", billing_error="checkout_failed"))
    if billing_period not in {"monthly", "annual"}:
        return redirect(url_for("app_home", billing_error="checkout_failed"))
    price_id = get_stripe_price_id(plan, billing_period)
    if not price_id:
        return redirect(
            url_for(
                "app_home",
                billing_error="annual_not_configured"
                if billing_period == "annual"
                else "checkout_failed",
            )
        )
    agency = get_agency_row_for_user(user)
    if not agency:
        return redirect(url_for("app_home", billing_error="checkout_failed"))
    customer_id = agency.get("stripe_customer_id")
    try:
        if agency.get("stripe_subscription_id") and customer_id and stripe_customer_exists(customer_id):
            return redirect(url_for("open_stripe_portal"))
        if customer_id and not stripe_customer_exists(customer_id):
            with engine.begin() as conn:
                clear_agency_stripe_state(conn, agency["id"])
            agency = get_agency_row_for_user(user)
            customer_id = None
        if not customer_id:
            customer = stripe.Customer.create(
                email=agency.get("email") or user.get("email"),
                name=agency.get("name") or agency.get("email") or user.get("email"),
                metadata={"agency_id": str(agency["id"])},
            )
            customer_id = customer.get("id")
            with engine.begin() as conn:
                sync_agency_billing_state(
                    conn,
                    agency["id"],
                    stripe_customer_id=customer_id,
                )
        checkout_kwargs = {
            "mode": "subscription",
            "customer": customer_id,
            "client_reference_id": str(agency["id"]),
            "line_items": [{"price": price_id, "quantity": 1}],
            "metadata": {
                "agency_id": str(agency["id"]),
                "plan": plan,
                "billing_period": billing_period,
            },
            "subscription_data": {
                "metadata": {
                    "agency_id": str(agency["id"]),
                    "plan": plan,
                    "billing_period": billing_period,
                }
            },
            "success_url": f"{get_app_base_url()}{url_for('app_home')}?billing=success",
            "cancel_url": f"{get_app_base_url()}{url_for('app_home')}?billing=cancelled",
        }
        discounts = get_default_stripe_discounts() if billing_period == "monthly" else []
        if discounts:
            checkout_kwargs["discounts"] = discounts
        else:
            checkout_kwargs["allow_promotion_codes"] = True
        checkout_session = stripe.checkout.Session.create(**checkout_kwargs)
    except Exception:
        app.logger.exception("Stripe checkout creation failed")
        return redirect(url_for("app_home", billing_error="checkout_failed"))
    return redirect(checkout_session.get("url"), code=303)


@app.route("/billing/portal", methods=["POST", "GET"])
def open_stripe_portal():
    user = g.current_user or {}
    if user.get("role") != "agency":
        return redirect(url_for("app_home"))
    if not stripe_is_configured():
        return redirect(url_for("app_home", billing_error="not_configured"))
    agency = get_agency_row_for_user(user)
    if not agency or not agency.get("stripe_customer_id"):
        return redirect(url_for("app_home", billing_error="not_available"))
    try:
        if not stripe_customer_exists(agency["stripe_customer_id"]):
            with engine.begin() as conn:
                clear_agency_stripe_state(conn, agency["id"])
            return redirect(url_for("app_home", billing_error="not_available"))
        portal_session = stripe.billing_portal.Session.create(
            customer=agency["stripe_customer_id"],
            return_url=f"{get_app_base_url()}{url_for('app_home')}",
        )
    except Exception:
        app.logger.exception("Stripe billing portal creation failed")
        return redirect(url_for("app_home", billing_error="portal_failed"))
    return redirect(portal_session.get("url"), code=303)


@app.route("/api/document-center/batches")
def list_document_batches():
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    with engine.connect() as conn:
        rows = conn.execute(
            select(document_batches_table)
            .where(document_batches_table.c.company_id == company_id)
            .order_by(document_batches_table.c.created_at.desc(), document_batches_table.c.id.desc())
            .limit(25)
        ).mappings().all()

    return jsonify({"ok": True, "batches": [serialize_document_batch(row) for row in rows]})


@app.route("/api/document-center/documents")
def list_document_center_documents():
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    batch_id = request.args.get("batch_id", type=int)
    validation_status = (request.args.get("validation_status") or "").strip().lower()
    query = (
        select(processed_documents_table)
        .where(processed_documents_table.c.company_id == company_id)
        .order_by(processed_documents_table.c.created_at.desc(), processed_documents_table.c.id.desc())
    )
    if batch_id:
        query = query.where(processed_documents_table.c.source_batch_id == batch_id)
    if validation_status:
        query = query.where(processed_documents_table.c.validation_status == validation_status)

    with engine.connect() as conn:
        rows = conn.execute(query).mappings().all()

    return jsonify({"ok": True, "documents": [serialize_processed_document(row) for row in rows]})


@app.route("/api/document-center/documents/<int:document_id>")
def get_document_center_document(document_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    with engine.connect() as conn:
        row = get_processed_document_for_company(conn, document_id, company_id)
    if not row:
        return jsonify({"ok": False, "errors": ["Documento no encontrado."]}), 404

    return jsonify({"ok": True, "document": serialize_processed_document(row)})


@app.route("/api/document-center/upload", methods=["POST"])
def upload_document_center_batch():
    data_owner_id = get_data_owner_id()
    current_user_id = get_current_user_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    files = request.files.getlist("files")
    if not files:
        return jsonify({"ok": False, "errors": ["No se recibieron archivos."]}), 400

    period = normalize_period_label(request.form.get("period"))
    created_at = datetime.utcnow().isoformat()
    inserted_documents = []

    with engine.begin() as conn:
        company_names = get_company_names_for_analysis(conn, company_id)
        known_suppliers = fetch_known_suppliers(conn, data_owner_id, company_id)
        batch_result = conn.execute(
            document_batches_table.insert().values(
                company_id=company_id,
                uploaded_by_user_id=current_user_id,
                period=period,
                total_documents=0,
                processed_documents=0,
                ready_documents=0,
                review_documents=0,
                duplicate_documents=0,
                failed_documents=0,
                status="processing",
                created_at=created_at,
                updated_at=created_at,
            )
        )
        batch_id = batch_result.inserted_primary_key[0]

        for item in iter_uploaded_documents(files):
            payload = build_processed_document_payload(
                file_bytes=item["bytes"],
                filename=item["filename"],
                content_type=item.get("content_type") or document_content_type_from_name(item["filename"]),
                company_id=company_id,
                data_owner_id=data_owner_id,
                current_user_id=current_user_id,
                batch_id=batch_id,
                period=period,
                company_names=company_names,
                known_suppliers=known_suppliers,
            )
            duplicate_row = conn.execute(
                select(processed_documents_table.c.id)
                .where(processed_documents_table.c.company_id == company_id)
                .where(processed_documents_table.c.content_hash == payload["content_hash"])
                .limit(1)
            ).first()
            duplicate_id = duplicate_row[0] if duplicate_row else document_duplicate_match(
                conn,
                company_id,
                payload["detected_document_type"],
                json_loads_dict(payload["original_extracted_data_json"]),
            )
            if duplicate_id:
                payload["validation_status"] = "duplicate"
                payload["issue_type"] = "duplicate"
                payload["issue_description"] = "Documento duplicado detectado."
                payload["duplicate_of_document_id"] = duplicate_id
                payload["audit_log_json"] = append_document_audit(
                    payload["audit_log_json"],
                    "duplicate_detected",
                    current_user_id,
                    {"duplicate_of_document_id": duplicate_id},
                )
            result = conn.execute(processed_documents_table.insert().values(**payload))
            inserted_documents.append(result.inserted_primary_key[0])

        refresh_document_batch_counters(conn, batch_id)
        batch_row = conn.execute(
            select(document_batches_table).where(document_batches_table.c.id == batch_id)
        ).mappings().first()
        document_rows = conn.execute(
            select(processed_documents_table)
            .where(processed_documents_table.c.source_batch_id == batch_id)
            .order_by(processed_documents_table.c.id.desc())
        ).mappings().all()

    return jsonify(
        {
            "ok": True,
            "batch": serialize_document_batch(batch_row),
            "documents": [serialize_processed_document(row) for row in document_rows],
            "count": len(inserted_documents),
        }
    )


@app.route("/api/document-center/documents/<int:document_id>", methods=["PUT"])
def update_document_center_document(document_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    current_user_id = get_current_user_id()
    payload = request.get_json(silent=True) or {}
    corrected_data = payload.get("corrected_data") or payload.get("correctedData") or {}
    if not isinstance(corrected_data, dict):
        return jsonify({"ok": False, "errors": ["Formato de corrección inválido."]}), 400
    detected_type = (payload.get("detected_document_type") or payload.get("detectedDocumentType") or "").strip()

    with engine.begin() as conn:
        row = get_processed_document_for_company(conn, document_id, company_id)
        if not row:
            return jsonify({"ok": False, "errors": ["Documento no encontrado."]}), 404
        effective_type = detected_type or row.get("detected_document_type") or "unknown"
        confidence_score, validation_status, issue_type, issue_description = extract_confidence_and_fields(
            effective_type,
            corrected_data,
            text_value=row.get("extracted_text") or "",
        )
        if row.get("duplicate_of_document_id"):
            validation_status = "duplicate"
            issue_type = "duplicate"
            issue_description = "Documento duplicado detectado."
        update_processed_document_record(
            conn,
            document_id,
            detected_document_type=effective_type,
            corrected_data_json=json_dumps(corrected_data),
            confidence_score=confidence_score,
            validation_status=validation_status,
            issue_type=issue_type,
            issue_description=issue_description,
            audit_log_json=append_document_audit(
                row.get("audit_log_json"),
                "edited",
                current_user_id,
                {"detected_type": effective_type},
            ),
        )
        refresh_document_batch_counters(conn, row["source_batch_id"])
        updated = get_processed_document_for_company(conn, document_id, company_id)

    return jsonify({"ok": True, "document": serialize_processed_document(updated)})


@app.route("/api/document-center/documents/<int:document_id>/approve", methods=["POST"])
def approve_document_center_document(document_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    current_user_id = get_current_user_id()

    with engine.begin() as conn:
        row = get_processed_document_for_company(conn, document_id, company_id)
        if not row:
            return jsonify({"ok": False, "errors": ["Documento no encontrado."]}), 404
        if row.get("duplicate_of_document_id"):
            return jsonify({"ok": False, "errors": ["No puedes aprobar un documento duplicado."]}), 400
        update_processed_document_record(
            conn,
            document_id,
            validation_status="ready_to_register",
            approved_at=datetime.utcnow().isoformat(),
            approved_by_user_id=current_user_id,
            rejected_at=None,
            rejected_by_user_id=None,
            issue_type=None,
            issue_description=None,
            audit_log_json=append_document_audit(row.get("audit_log_json"), "approved", current_user_id),
        )
        refresh_document_batch_counters(conn, row["source_batch_id"])
        updated = get_processed_document_for_company(conn, document_id, company_id)

    return jsonify({"ok": True, "document": serialize_processed_document(updated)})


@app.route("/api/document-center/documents/<int:document_id>/reject", methods=["POST"])
def reject_document_center_document(document_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    current_user_id = get_current_user_id()
    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "Documento rechazado manualmente.").strip()

    with engine.begin() as conn:
        row = get_processed_document_for_company(conn, document_id, company_id)
        if not row:
            return jsonify({"ok": False, "errors": ["Documento no encontrado."]}), 404
        update_processed_document_record(
            conn,
            document_id,
            validation_status="rejected",
            issue_type="manual_rejection",
            issue_description=reason,
            rejected_at=datetime.utcnow().isoformat(),
            rejected_by_user_id=current_user_id,
            audit_log_json=append_document_audit(
                row.get("audit_log_json"),
                "rejected",
                current_user_id,
                {"reason": reason},
            ),
        )
        refresh_document_batch_counters(conn, row["source_batch_id"])
        updated = get_processed_document_for_company(conn, document_id, company_id)

    return jsonify({"ok": True, "document": serialize_processed_document(updated)})


@app.route("/api/document-center/documents/<int:document_id>/register", methods=["POST"])
def register_document_center_document(document_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    data_owner_id = get_data_owner_id()
    current_user_id = get_current_user_id()

    with engine.begin() as conn:
        row = get_processed_document_for_company(conn, document_id, company_id)
        if not row:
            return jsonify({"ok": False, "errors": ["Documento no encontrado."]}), 404
        if row.get("registered_at"):
            return jsonify({"ok": False, "errors": ["El documento ya está registrado."]}), 400
        if row.get("validation_status") == "duplicate":
            return jsonify({"ok": False, "errors": ["No se puede registrar un documento duplicado."]}), 400
        try:
            record_id, record_type = register_processed_document(
                conn,
                row,
                data_owner_id,
                current_user_id,
            )
        except ValueError as exc:
            return jsonify({"ok": False, "errors": [str(exc)]}), 400
        update_processed_document_record(
            conn,
            document_id,
            validation_status="registered",
            linked_accounting_record_id=record_id,
            linked_accounting_record_type=record_type,
            registered_at=datetime.utcnow().isoformat(),
            registered_by_user_id=current_user_id,
            approved_at=row.get("approved_at") or datetime.utcnow().isoformat(),
            approved_by_user_id=row.get("approved_by_user_id") or current_user_id,
            audit_log_json=append_document_audit(
                row.get("audit_log_json"),
                "registered",
                current_user_id,
                {"record_id": record_id, "record_type": record_type},
            ),
        )
        refresh_document_batch_counters(conn, row["source_batch_id"])
        updated = get_processed_document_for_company(conn, document_id, company_id)

    return jsonify({"ok": True, "document": serialize_processed_document(updated)})


@app.route("/api/document-center/batches/<int:batch_id>/register-ready", methods=["POST"])
def register_ready_document_batch(batch_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    data_owner_id = get_data_owner_id()
    current_user_id = get_current_user_id()
    registered = 0
    errors = []

    with engine.begin() as conn:
        batch = conn.execute(
            select(document_batches_table)
            .where(document_batches_table.c.id == batch_id)
            .where(document_batches_table.c.company_id == company_id)
        ).mappings().first()
        if not batch:
            return jsonify({"ok": False, "errors": ["Lote no encontrado."]}), 404
        rows = conn.execute(
            select(processed_documents_table)
            .where(processed_documents_table.c.source_batch_id == batch_id)
            .where(processed_documents_table.c.validation_status == "ready_to_register")
            .where(processed_documents_table.c.registered_at.is_(None))
        ).mappings().all()
        for row in rows:
            try:
                record_id, record_type = register_processed_document(
                    conn,
                    row,
                    data_owner_id,
                    current_user_id,
                )
            except ValueError as exc:
                errors.append(f"{row['original_filename']}: {exc}")
                continue
            update_processed_document_record(
                conn,
                row["id"],
                validation_status="registered",
                linked_accounting_record_id=record_id,
                linked_accounting_record_type=record_type,
                registered_at=datetime.utcnow().isoformat(),
                registered_by_user_id=current_user_id,
                approved_at=row.get("approved_at") or datetime.utcnow().isoformat(),
                approved_by_user_id=row.get("approved_by_user_id") or current_user_id,
                audit_log_json=append_document_audit(
                    row.get("audit_log_json"),
                    "registered",
                    current_user_id,
                    {"record_id": record_id, "record_type": record_type},
                ),
            )
            registered += 1
        refresh_document_batch_counters(conn, batch_id)
        updated_batch = conn.execute(
            select(document_batches_table).where(document_batches_table.c.id == batch_id)
        ).mappings().first()

    return jsonify({"ok": True, "registered": registered, "errors": errors, "batch": serialize_document_batch(updated_batch)})


@app.route("/api/document-center/documents/<int:document_id>/download")
def download_document_center_document(document_id):
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    with engine.connect() as conn:
        row = get_processed_document_for_company(conn, document_id, company_id)
    if not row:
        return jsonify({"ok": False, "errors": ["Documento no encontrado."]}), 404

    storage_path = row.get("storage_path")
    file_url = row.get("file_url")
    if storage_path and os.path.exists(storage_path):
        return send_file(storage_path, as_attachment=False, download_name=row.get("original_filename"))
    if file_url and file_url.startswith(("http://", "https://")):
        return redirect(file_url)
    if file_url and os.path.exists(file_url):
        return send_file(file_url, as_attachment=False, download_name=row.get("original_filename"))
    return jsonify({"ok": False, "errors": ["Archivo no disponible."]}), 404


@app.route("/api/account", methods=["PUT"])
def update_account():
    current_user = g.current_user or {}
    user_id = current_user.get("id")
    if not user_id:
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403

    payload = request.get_json(silent=True) or {}
    email = _normalize_email(payload.get("email"))
    agency_name = (payload.get("agency_name") or payload.get("agencyName") or "").strip()
    phone = (payload.get("phone") or "").strip()
    current_password = payload.get("current_password") or payload.get("currentPassword") or ""
    new_password = payload.get("new_password") or payload.get("newPassword") or ""

    role = current_user.get("role")
    errors = []
    if not email:
        errors.append("Email obligatorio.")
    if role == "agency" and not agency_name:
        errors.append("Nombre de la gestoría obligatorio.")
    if new_password and len(new_password) < 8:
        errors.append("La nueva contraseña debe tener al menos 8 caracteres.")
    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    with engine.begin() as conn:
        user_row = conn.execute(
            select(
                users_table.c.id,
                users_table.c.email,
                users_table.c.password_hash,
            ).where(users_table.c.id == user_id)
        ).mappings().first()
        if not user_row:
            return jsonify({"ok": False, "errors": ["Usuario no encontrado."]}), 404

        email_changed = email != (user_row.get("email") or "").strip().lower()
        password_changed = bool(new_password)
        if email_changed or password_changed:
            if not current_password:
                return (
                    jsonify(
                        {
                            "ok": False,
                            "errors": ["Debes indicar tu contraseña actual para guardar estos cambios."],
                        }
                    ),
                    400,
                )
            if not check_password_hash(user_row["password_hash"], current_password):
                return jsonify({"ok": False, "errors": ["La contraseña actual no es correcta."]}), 400

        if email_changed:
            existing = conn.execute(
                select(users_table.c.id)
                .where(users_table.c.email == email)
                .where(users_table.c.id != user_id)
            ).first()
            if existing:
                return jsonify({"ok": False, "errors": ["Ese email ya está en uso."]}), 400

        user_updates = {}
        if email_changed:
            user_updates["email"] = email
        if password_changed:
            user_updates["password_hash"] = generate_password_hash(new_password)
        if user_updates:
            conn.execute(
                users_table.update().where(users_table.c.id == user_id).values(**user_updates)
            )

        stripe_sync = None
        if role == "agency":
            agency_row = conn.execute(
                select(
                    agencies_table.c.id,
                    agencies_table.c.email,
                    agencies_table.c.name,
                    agencies_table.c.phone,
                    agencies_table.c.stripe_customer_id,
                ).where(agencies_table.c.id == get_agency_id_for_user(current_user))
            ).mappings().first()
            if not agency_row:
                return jsonify({"ok": False, "errors": ["Gestoría no encontrada."]}), 404
            agency_updates = {
                "name": agency_name,
                "phone": phone or None,
            }
            if email_changed:
                agency_updates["email"] = email
            conn.execute(
                agencies_table.update()
                .where(agencies_table.c.id == agency_row["id"])
                .values(**agency_updates)
            )
            stripe_sync = {
                "customer_id": agency_row.get("stripe_customer_id"),
                "email": email if email_changed else agency_row.get("email"),
                "name": agency_name,
                "phone": phone or None,
            }

    if stripe_sync and stripe and stripe_sync.get("customer_id"):
        try:
            stripe.Customer.modify(
                stripe_sync["customer_id"],
                email=stripe_sync.get("email"),
                name=stripe_sync.get("name"),
                phone=stripe_sync.get("phone"),
            )
        except Exception:
            app.logger.warning(
                "No se pudo sincronizar el cliente de Stripe para la gestoría %s",
                get_agency_id_for_user(current_user),
            )

    refreshed_user = get_current_user()
    return jsonify(
        {
            "ok": True,
            "account": get_account_context_for_user(refreshed_user),
        }
    )


@app.route("/api/stripe/webhook", methods=["POST"])
def stripe_webhook():
    if not stripe_is_configured():
        return jsonify({"ok": False, "error": "stripe_not_configured"}), 503
    payload = request.get_data(as_text=False)
    signature = request.headers.get("Stripe-Signature", "")
    try:
        event = stripe.Webhook.construct_event(
            payload=payload,
            sig_header=signature,
            secret=STRIPE_WEBHOOK_SECRET,
        )
    except ValueError:
        return jsonify({"ok": False, "error": "invalid_payload"}), 400
    except stripe.error.SignatureVerificationError:
        return jsonify({"ok": False, "error": "invalid_signature"}), 400

    event_type = event.get("type")
    data_object = (event.get("data") or {}).get("object") or {}

    try:
        with engine.begin() as conn:
            if event_type == "checkout.session.completed":
                agency_id = (
                    (data_object.get("metadata") or {}).get("agency_id")
                    or data_object.get("client_reference_id")
                )
                if agency_id:
                    sync_agency_billing_state(
                        conn,
                        int(agency_id),
                        plan=(data_object.get("metadata") or {}).get("plan") or "starter",
                        app_status="active",
                        stripe_customer_id=data_object.get("customer"),
                        stripe_subscription_id=data_object.get("subscription"),
                    )
            elif event_type in {
                "customer.subscription.created",
                "customer.subscription.updated",
                "customer.subscription.deleted",
            }:
                metadata = data_object.get("metadata") or {}
                agency_id = metadata.get("agency_id")
                price_id = ""
                items = ((data_object.get("items") or {}).get("data") or [])
                if items:
                    price_id = ((items[0].get("price") or {}).get("id") or "").strip()
                price_mapping = get_plan_from_price_id(price_id)
                plan = metadata.get("plan") or price_mapping.get("plan") or "starter"
                current_period_end = data_object.get("current_period_end")
                resolved_status = map_stripe_subscription_status(data_object.get("status"))
                if not agency_id:
                    lookup = conn.execute(
                        select(agencies_table.c.id).where(
                            (agencies_table.c.stripe_customer_id == data_object.get("customer"))
                            | (
                                agencies_table.c.stripe_subscription_id
                                == data_object.get("id")
                            )
                        )
                    ).first()
                    agency_id = lookup[0] if lookup else None
                if agency_id:
                    sync_agency_billing_state(
                        conn,
                        int(agency_id),
                        plan=plan,
                        app_status=resolved_status,
                        stripe_customer_id=data_object.get("customer"),
                        stripe_subscription_id=data_object.get("id"),
                        stripe_price_id=price_id or None,
                        stripe_subscription_status=data_object.get("status"),
                        stripe_current_period_end=(
                            datetime.utcfromtimestamp(current_period_end).isoformat()
                            if current_period_end
                            else None
                        ),
                    )
    except Exception:
        app.logger.exception("Stripe webhook processing failed")
        return jsonify({"ok": False, "error": "webhook_processing_failed"}), 500

    return jsonify({"ok": True})


@app.route("/admin")
@require_owner
def admin_dashboard():
    admin_message = (request.args.get("message") or "").strip().lower()
    admin_error = (request.args.get("error") or "").strip().lower()
    with engine.connect() as conn:
        agencies = conn.execute(select(agencies_table)).mappings().all()
        company_counts = {
            row["agency_id"]: row["count"]
            for row in conn.execute(
                select(
                    companies_table.c.agency_id,
                    func.count().label("count"),
                ).group_by(companies_table.c.agency_id)
            ).mappings().all()
        }
        user_counts = {
            row["agency_id"]: row["count"]
            for row in conn.execute(
                select(
                    users_table.c.agency_id,
                    func.count().label("count"),
                )
                .where(users_table.c.role.in_(["agency", "staff"]))
                .group_by(users_table.c.agency_id)
            ).mappings().all()
        }

    view_rows = []
    for agency in agencies:
        stripe_status = (
            "Con suscripción"
            if agency.get("stripe_subscription_id")
            else "Sin suscripción"
        )
        view_rows.append(
            {
                "id": agency["id"],
                "name": agency["name"],
                "email": agency["email"],
                "plan": agency["plan"],
                "status": agency["status"],
                "company_count": company_counts.get(agency["id"], 0),
                "user_count": user_counts.get(agency["id"], 0),
                "last_login": agency.get("last_login_at"),
                "stripe_status": stripe_status,
            }
        )
    view_rows.sort(key=lambda item: item["name"].lower())
    flash = None
    if admin_message == "agency_created":
        flash = {
            "type": "success",
            "text": "Gestoría creada correctamente. Ya puedes entrar con esa cuenta.",
        }
    elif admin_message == "agency_invited":
        flash = {
            "type": "success",
            "text": "Invitación enviada a la gestoría. El alta se completará cuando acepte el email.",
        }
    elif admin_error == "missing_fields":
        flash = {
            "type": "warning",
            "text": "Nombre, email y contraseña son obligatorios.",
        }
    elif admin_error == "short_password":
        flash = {
            "type": "warning",
            "text": "La contraseña debe tener al menos 8 caracteres.",
        }
    elif admin_error == "email_exists":
        flash = {
            "type": "warning",
            "text": "Ese email ya está registrado.",
        }
    elif admin_error == "invite_exists":
        flash = {
            "type": "warning",
            "text": "Ya existe una invitación pendiente para ese email.",
        }
    elif admin_error == "invite_send_failed":
        flash = {
            "type": "warning",
            "text": "La invitación se creó pero el email no pudo enviarse. Revisa la configuración de Resend.",
        }
    elif admin_message == "stripe_reset":
        flash = {
            "type": "success",
            "text": "La vinculación de Stripe se ha limpiado para esa gestoría.",
        }
    return render_template("admin.html", agencies=view_rows, flash=flash)


@app.route("/admin/agency/create", methods=["POST"])
@require_owner
def admin_create_agency():
    name = (request.form.get("name") or "").strip()
    email = _normalize_email(request.form.get("email"))
    password = request.form.get("password") or ""
    plan = (request.form.get("plan") or "starter").strip().lower()

    if not name or not email or not password:
        return redirect(url_for("admin_dashboard", error="missing_fields"))
    if len(password) < 8:
        return redirect(url_for("admin_dashboard", error="short_password"))
    if plan not in {"starter", "pro", "advanced"}:
        plan = "starter"

    created_at = datetime.utcnow().isoformat()
    trial_ends = (datetime.utcnow() + timedelta(days=14)).isoformat()
    with engine.begin() as conn:
        exists = conn.execute(
            select(users_table.c.id).where(users_table.c.email == email)
        ).first()
        if exists:
            return redirect(url_for("admin_dashboard", error="email_exists"))

        result = conn.execute(
            users_table.insert().values(
                email=email,
                password_hash=generate_password_hash(password),
                role="agency",
                plan="trial",
                agency_id=None,
                created_at=created_at,
                is_active=True,
            )
        )
        agency_user_id = result.inserted_primary_key[0]
        conn.execute(
            users_table.update()
            .where(users_table.c.id == agency_user_id)
            .values(agency_id=agency_user_id)
        )
        conn.execute(
            agencies_table.insert().values(
                id=agency_user_id,
                name=name,
                email=email,
                phone=None,
                plan=plan,
                status="trial",
                stripe_customer_id=None,
                stripe_subscription_id=None,
                stripe_price_id=None,
                stripe_subscription_status=None,
                stripe_current_period_end=None,
                trial_ends_at=trial_ends,
                created_at=created_at,
                last_login_at=None,
            )
        )
    return redirect(url_for("admin_dashboard", message="agency_created"))


@app.route("/admin/agency/invite", methods=["POST"])
@require_owner
def admin_invite_agency():
    name = (request.form.get("name") or "").strip()
    email = _normalize_email(request.form.get("email"))
    plan = (request.form.get("plan") or "starter").strip().lower()
    if not name or not email:
        return redirect(url_for("admin_dashboard", error="missing_fields"))
    if plan not in {"starter", "pro", "advanced"}:
        plan = "starter"
    result = create_user_invitation(
        email=email,
        role="agency",
        created_by_user_id=int(g.current_user["id"]),
        name=name,
        plan=plan,
    )
    if not result.get("ok"):
        errors = result.get("errors") or []
        if any("pendiente" in error.lower() for error in errors):
            return redirect(url_for("admin_dashboard", error="invite_exists"))
        return redirect(url_for("admin_dashboard", error="email_exists"))
    sent = send_user_invitation_email(
        email=email,
        role="agency",
        token=result["token"],
        sender_name=g.current_user.get("email"),
        agency_name=name,
        reply_to=g.current_user.get("email"),
    )
    if not sent:
        with engine.begin() as conn:
            conn.execute(
                user_invitations_table.delete().where(
                    user_invitations_table.c.token == result["token"]
                )
            )
        return redirect(url_for("admin_dashboard", error="invite_send_failed"))
    return redirect(url_for("admin_dashboard", message="agency_invited"))


@app.route("/admin/agency/<int:agency_id>/stripe/reset", methods=["POST"])
@require_owner
def admin_reset_agency_stripe(agency_id):
    with engine.begin() as conn:
        exists = conn.execute(
            select(agencies_table.c.id).where(agencies_table.c.id == agency_id)
        ).first()
        if not exists:
            return jsonify({"ok": False, "errors": ["Gestoría no encontrada."]}), 404
        clear_agency_stripe_state(conn, agency_id)
    return redirect(url_for("admin_dashboard", message="stripe_reset"))


@app.route("/admin/agency/<int:agency_id>/plan", methods=["POST"])
@require_owner
def admin_update_plan(agency_id):
    plan = (request.form.get("plan") or "").strip().lower()
    if plan not in {"starter", "pro", "advanced"}:
        return jsonify({"ok": False, "errors": ["Plan inválido."]}), 400
    with engine.begin() as conn:
        result = conn.execute(
            agencies_table.update()
            .where(agencies_table.c.id == agency_id)
            .values(plan=plan)
        )
    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Gestoría no encontrada."]}), 404
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/agency/<int:agency_id>/status", methods=["POST"])
@require_owner
def admin_update_status(agency_id):
    status = (request.form.get("status") or "").strip().lower()
    if status not in {"trial", "active", "suspended"}:
        return jsonify({"ok": False, "errors": ["Estado inválido."]}), 400
    with engine.begin() as conn:
        result = conn.execute(
            agencies_table.update()
            .where(agencies_table.c.id == agency_id)
            .values(status=status)
        )
        if result.rowcount:
            conn.execute(
                users_table.update()
                .where(users_table.c.agency_id == agency_id)
                .values(is_active=status != "suspended")
            )
    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Gestoría no encontrada."]}), 404
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/agency/<int:agency_id>/trial", methods=["POST"])
@require_owner
def admin_reset_trial(agency_id):
    trial_ends = (datetime.utcnow() + timedelta(days=14)).isoformat()
    with engine.begin() as conn:
        result = conn.execute(
            agencies_table.update()
            .where(agencies_table.c.id == agency_id)
            .values(status="trial", trial_ends_at=trial_ends)
        )
    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Gestoría no encontrada."]}), 404
    return redirect(url_for("admin_dashboard"))


@app.route("/aviso-legal")
def legal_notice():
    return render_template("aviso_legal.html")


@app.route("/privacidad")
def privacy():
    return render_template("privacidad.html")


@app.route("/cookies")
def cookies():
    return render_template("cookies.html")


@app.route("/terminos")
def terms():
    return render_template("terminos.html")


@app.route("/api/companies")
def list_companies():
    user_id = get_current_user_id()
    user_role = (g.current_user or {}).get("role")
    base_query = select(
        companies_table.c.id,
        companies_table.c.display_name,
        companies_table.c.legal_name,
        companies_table.c.tax_id,
        companies_table.c.company_type,
        companies_table.c.email,
        companies_table.c.phone,
        companies_table.c.assigned_user_id,
        companies_table.c.vat_regime,
        companies_table.c.tax_periodicity,
        companies_table.c.files_model_303,
        companies_table.c.files_model_111,
        companies_table.c.files_model_115,
        companies_table.c.files_model_130,
        companies_table.c.files_model_202,
        companies_table.c.balance_manual_data,
    )
    if user_role == "staff":
        base_query = base_query.where(companies_table.c.assigned_user_id == user_id)
    elif user_role == "owner":
        base_query = base_query
    else:
        base_query = base_query.where(companies_table.c.agency_id == user_id)
    with engine.connect() as conn:
        rows = conn.execute(base_query).mappings().all()
    companies = [
        {
            "id": row["id"],
            "display_name": row["display_name"],
            "legal_name": row["legal_name"],
            "tax_id": row["tax_id"],
            "company_type": row["company_type"],
            "email": row["email"],
            "phone": row["phone"],
            "assigned_user_id": row["assigned_user_id"],
            "vat_regime": row.get("vat_regime") or "general",
            "tax_periodicity": row.get("tax_periodicity") or "quarterly",
            "files_model_303": bool(row.get("files_model_303")) if row.get("files_model_303") is not None else True,
            "files_model_111": bool(row.get("files_model_111")) if row.get("files_model_111") is not None else False,
            "files_model_115": bool(row.get("files_model_115")) if row.get("files_model_115") is not None else False,
            "files_model_130": bool(row.get("files_model_130")) if row.get("files_model_130") is not None else False,
            "files_model_202": bool(row.get("files_model_202")) if row.get("files_model_202") is not None else False,
            "balance_manual_data": row.get("balance_manual_data"),
        }
        for row in rows
    ]
    return jsonify({"companies": companies})


@app.route("/api/companies", methods=["POST"])
def create_company():
    user_id = get_current_user_id()
    if (g.current_user or {}).get("role") == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    payload = request.get_json(silent=True) or {}

    display_name = (payload.get("display_name") or payload.get("displayName") or "").strip()
    legal_name = (payload.get("legal_name") or payload.get("legalName") or "").strip()
    tax_id = (payload.get("tax_id") or payload.get("taxId") or "").strip().upper()
    company_type = payload.get("company_type") or payload.get("companyType") or ""
    email = (payload.get("email") or "").strip()
    phone = (payload.get("phone") or "").strip()
    assigned_user_id = payload.get("assigned_user_id") or payload.get("assignedUserId")
    vat_regime = (payload.get("vat_regime") or payload.get("vatRegime") or "general").strip()
    tax_periodicity = (payload.get("tax_periodicity") or payload.get("taxPeriodicity") or "quarterly").strip()
    files_model_303 = payload.get("files_model_303")
    files_model_111 = payload.get("files_model_111")
    files_model_115 = payload.get("files_model_115")
    files_model_130 = payload.get("files_model_130")
    files_model_202 = payload.get("files_model_202")

    errors = []
    if not display_name:
        errors.append("Nombre comercial obligatorio.")
    if not legal_name:
        errors.append("Razón social obligatoria.")
    if company_type not in {"individual", "company"}:
        errors.append("Tipo de empresa inválido.")
    if not validate_tax_id(tax_id, company_type):
        errors.append("CIF/NIF inválido.")
    if vat_regime not in {"general", "exempt", "prorata"}:
        errors.append("Régimen de IVA inválido.")
    if tax_periodicity not in {"quarterly", "monthly"}:
        errors.append("Periodicidad fiscal inválida.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    with engine.connect() as conn:
        existing_count = conn.execute(
            select(func.count())
            .select_from(companies_table)
            .where(companies_table.c.agency_id == user_id)
        ).scalar_one()
        exists = conn.execute(
            select(companies_table.c.id)
            .where(companies_table.c.agency_id == user_id)
            .where(companies_table.c.tax_id == tax_id)
        ).first()
    if exists:
        return jsonify({"ok": False, "errors": ["Ya existe una empresa con ese CIF/NIF."]}), 400

    with engine.connect() as conn:
        agency_limits = get_agency_limits_and_usage(conn, user_id)
    if agency_limits["usage"]["companies"] >= agency_limits["limits"]["companies"]:
        return (
            jsonify(
                {
                    "ok": False,
                    "errors": [
                        get_limit_error("companies", agency_limits["limits"]["companies"])
                    ],
                }
            ),
            403,
        )

    assigned_user_id = resolve_assigned_staff(user_id, assigned_user_id)
    files_model_303 = files_model_303 not in (False, "false", "False", 0, "0", None) if files_model_303 is not None else vat_regime != "exempt"
    files_model_111 = files_model_111 in (True, "true", "True", 1, "1")
    files_model_115 = files_model_115 in (True, "true", "True", 1, "1")
    files_model_130 = files_model_130 in (True, "true", "True", 1, "1")
    files_model_202 = files_model_202 in (True, "true", "True", 1, "1")

    created_at = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        result = conn.execute(
            companies_table.insert().values(
                user_id=user_id,
                agency_id=user_id,
                display_name=display_name,
                legal_name=legal_name,
                tax_id=tax_id,
                company_type=company_type,
                email=email,
                phone=phone,
                assigned_user_id=assigned_user_id,
                vat_regime=vat_regime,
                tax_periodicity=tax_periodicity,
                files_model_303=files_model_303,
                files_model_111=files_model_111,
                files_model_115=files_model_115,
                files_model_130=files_model_130,
                files_model_202=files_model_202,
                created_at=created_at,
            )
        )
        new_id = result.inserted_primary_key[0]
        if existing_count == 0:
            conn.execute(
                invoices_table.update()
                .where(invoices_table.c.user_id == user_id)
                .where(invoices_table.c.company_id.is_(None))
                .values(company_id=new_id)
            )
            conn.execute(
                no_invoice_table.update()
                .where(no_invoice_table.c.user_id == user_id)
                .where(no_invoice_table.c.company_id.is_(None))
                .values(company_id=new_id)
            )
            conn.execute(
                facturacion_table.update()
                .where(facturacion_table.c.user_id == user_id)
                .where(facturacion_table.c.company_id.is_(None))
                .values(company_id=new_id)
            )
            conn.execute(
                income_invoices_table.update()
                .where(income_invoices_table.c.user_id == user_id)
                .where(income_invoices_table.c.company_id.is_(None))
                .values(company_id=new_id)
            )

    return jsonify({"ok": True, "id": new_id})


@app.route("/api/companies/<int:company_id>", methods=["PUT"])
def update_company(company_id):
    user_id = get_current_user_id()
    if (g.current_user or {}).get("role") == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    payload = request.get_json(silent=True) or {}

    display_name = (payload.get("display_name") or payload.get("displayName") or "").strip()
    legal_name = (payload.get("legal_name") or payload.get("legalName") or "").strip()
    tax_id = (payload.get("tax_id") or payload.get("taxId") or "").strip().upper()
    company_type = payload.get("company_type") or payload.get("companyType") or ""
    email = (payload.get("email") or "").strip()
    phone = (payload.get("phone") or "").strip()
    assigned_user_id = payload.get("assigned_user_id") or payload.get("assignedUserId")
    vat_regime = (payload.get("vat_regime") or payload.get("vatRegime") or "general").strip()
    tax_periodicity = (payload.get("tax_periodicity") or payload.get("taxPeriodicity") or "quarterly").strip()
    files_model_303 = payload.get("files_model_303")
    files_model_111 = payload.get("files_model_111")
    files_model_115 = payload.get("files_model_115")
    files_model_130 = payload.get("files_model_130")
    files_model_202 = payload.get("files_model_202")

    errors = []
    if not display_name:
        errors.append("Nombre comercial obligatorio.")
    if not legal_name:
        errors.append("Razón social obligatoria.")
    if company_type not in {"individual", "company"}:
        errors.append("Tipo de empresa inválido.")
    if not validate_tax_id(tax_id, company_type):
        errors.append("CIF/NIF inválido.")
    if vat_regime not in {"general", "exempt", "prorata"}:
        errors.append("Régimen de IVA inválido.")
    if tax_periodicity not in {"quarterly", "monthly"}:
        errors.append("Periodicidad fiscal inválida.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    with engine.connect() as conn:
        exists = conn.execute(
            select(companies_table.c.id)
            .where(companies_table.c.agency_id == user_id)
            .where(companies_table.c.tax_id == tax_id)
            .where(companies_table.c.id != company_id)
        ).first()
    if exists:
        return jsonify({"ok": False, "errors": ["Ya existe una empresa con ese CIF/NIF."]}), 400

    assigned_user_id = resolve_assigned_staff(user_id, assigned_user_id)
    files_model_303 = files_model_303 not in (False, "false", "False", 0, "0", None) if files_model_303 is not None else vat_regime != "exempt"
    files_model_111 = files_model_111 in (True, "true", "True", 1, "1")
    files_model_115 = files_model_115 in (True, "true", "True", 1, "1")
    files_model_130 = files_model_130 in (True, "true", "True", 1, "1")
    files_model_202 = files_model_202 in (True, "true", "True", 1, "1")

    with engine.begin() as conn:
        result = conn.execute(
            companies_table.update()
            .where(companies_table.c.id == company_id)
            .where(companies_table.c.agency_id == user_id)
            .values(
                display_name=display_name,
                legal_name=legal_name,
                tax_id=tax_id,
                company_type=company_type,
                email=email,
                phone=phone,
                assigned_user_id=assigned_user_id,
                vat_regime=vat_regime,
                tax_periodicity=tax_periodicity,
                files_model_303=files_model_303,
                files_model_111=files_model_111,
                files_model_115=files_model_115,
                files_model_130=files_model_130,
                files_model_202=files_model_202,
            )
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404

    return jsonify({"ok": True})


@app.route("/api/companies/<int:company_id>/balance-manual-data", methods=["PUT"])
def update_company_balance_manual_data(company_id):
    user_id = get_current_user_id()
    user_role = (g.current_user or {}).get("role")
    if user_role == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403

    payload = request.get_json(silent=True) or {}
    raw_data = payload.get("balance_manual_data") or payload.get("balanceManualData") or {}
    if not isinstance(raw_data, dict):
        return jsonify({"ok": False, "errors": ["Formato de balance inválido."]}), 400

    clean_data = {}
    for key, value in raw_data.items():
        if value in (None, "", False):
            continue
        try:
            clean_value = round(float(value), 2)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "errors": [f"Valor inválido para {key}."]}), 400
        clean_data[str(key)] = clean_value

    with engine.begin() as conn:
        query = (
            companies_table.update()
            .where(companies_table.c.id == company_id)
            .values(balance_manual_data=json.dumps(clean_data) if clean_data else None)
        )
        if user_role != "owner":
            query = query.where(companies_table.c.agency_id == user_id)
        result = conn.execute(query)

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404

    return jsonify({"ok": True, "balance_manual_data": clean_data})


@app.route("/api/companies/<int:company_id>", methods=["DELETE"])
def delete_company(company_id):
    user_id = get_current_user_id()
    if (g.current_user or {}).get("role") == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    with engine.begin() as conn:
        result = conn.execute(
            companies_table.delete()
            .where(companies_table.c.id == company_id)
            .where(companies_table.c.agency_id == user_id)
        )
    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
    return jsonify({"ok": True})


@app.route("/api/staff")
def list_staff():
    user_id = get_current_user_id()
    role = (g.current_user or {}).get("role")
    if role == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    with engine.connect() as conn:
        rows = conn.execute(
            select(
                users_table.c.id,
                users_table.c.email,
                users_table.c.role,
                users_table.c.is_active,
            )
            .where(users_table.c.agency_id == user_id)
            .where(users_table.c.role == "staff")
        ).mappings().all()
    staff = [
        {
            "id": row["id"],
            "email": row["email"],
            "is_active": bool(row["is_active"]),
        }
        for row in rows
    ]
    return jsonify({"staff": staff})


@app.route("/api/staff", methods=["POST"])
def create_staff():
    user_id = get_current_user_id()
    role = (g.current_user or {}).get("role")
    if role == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    payload = request.get_json(silent=True) or {}
    email = _normalize_email(payload.get("email"))
    password = payload.get("password") or ""
    if not email or not password:
        return jsonify({"ok": False, "errors": ["Email y contraseña obligatorios."]}), 400
    if len(password) < 8:
        return jsonify({"ok": False, "errors": ["La contraseña debe tener al menos 8 caracteres."]}), 400

    with engine.begin() as conn:
        exists = conn.execute(
            select(users_table.c.id).where(users_table.c.email == email)
        ).first()
        if exists:
            return jsonify({"ok": False, "errors": ["El email ya está registrado."]}), 400
        agency_limits = get_agency_limits_and_usage(conn, user_id)
        if agency_limits["usage"]["staff"] >= agency_limits["limits"]["staff"]:
            return (
                jsonify(
                    {
                        "ok": False,
                        "errors": [get_limit_error("staff", agency_limits["limits"]["staff"])],
                    }
                ),
                403,
            )
        agency_plan = conn.execute(
            select(users_table.c.plan).where(users_table.c.id == user_id)
        ).scalar_one_or_none()
        result = conn.execute(
            users_table.insert().values(
                email=email,
                password_hash=generate_password_hash(password),
                role="staff",
                plan=agency_plan or "trial",
                agency_id=user_id,
                created_at=datetime.utcnow().isoformat(),
                is_active=True,
            )
        )
        staff_id = result.inserted_primary_key[0]
    return jsonify({"ok": True, "id": staff_id})


@app.route("/api/staff/invite", methods=["POST"])
def invite_staff():
    user_id = get_current_user_id()
    role = (g.current_user or {}).get("role")
    if role == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    payload = request.get_json(silent=True) or {}
    email = _normalize_email(payload.get("email"))
    if not email:
        return jsonify({"ok": False, "errors": ["Email obligatorio."]}), 400
    with engine.connect() as conn:
        agency_limits = get_agency_limits_and_usage(conn, user_id)
    occupied_staff_slots = (
        agency_limits["usage"]["staff"] + agency_limits["usage"]["pending_staff_invitations"]
    )
    if occupied_staff_slots >= agency_limits["limits"]["staff"]:
        return (
            jsonify(
                {
                    "ok": False,
                    "errors": [get_limit_error("staff", agency_limits["limits"]["staff"])],
                }
            ),
            403,
        )

    sender_name = (g.current_user or {}).get("email")
    agency = get_agency_row_for_user(g.current_user or {})
    agency_name = agency.get("name") if agency else None
    result = create_user_invitation(
        email=email,
        role="staff",
        created_by_user_id=user_id,
        agency_id=get_agency_id_for_user(g.current_user or {}),
    )
    if not result.get("ok"):
        return jsonify(result), 400
    sent = send_user_invitation_email(
        email=email,
        role="staff",
        token=result["token"],
        sender_name=sender_name,
        agency_name=agency_name,
        reply_to=sender_name,
    )
    if not sent:
        with engine.begin() as conn:
            conn.execute(
                user_invitations_table.delete().where(
                    user_invitations_table.c.token == result["token"]
                )
            )
        return jsonify({"ok": False, "errors": ["No se pudo enviar la invitación."]}), 500
    return jsonify({"ok": True})


@app.route("/api/staff/<int:staff_id>", methods=["PUT"])
def update_staff(staff_id):
    user_id = get_current_user_id()
    role = (g.current_user or {}).get("role")
    if role == "staff":
        return jsonify({"ok": False, "errors": ["No autorizado."]}), 403
    payload = request.get_json(silent=True) or {}
    password = payload.get("password")
    is_active = payload.get("is_active")
    updates = {}
    if password:
        if len(password) < 8:
            return jsonify({"ok": False, "errors": ["La contraseña debe tener al menos 8 caracteres."]}), 400
        updates["password_hash"] = generate_password_hash(password)
    if is_active is not None:
        updates["is_active"] = bool(is_active)
    if not updates:
        return jsonify({"ok": False, "errors": ["Nada que actualizar."]}), 400
    with engine.begin() as conn:
        result = conn.execute(
            users_table.update()
            .where(users_table.c.id == staff_id)
            .where(users_table.c.agency_id == user_id)
            .where(users_table.c.role == "staff")
            .values(**updates)
        )
    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Usuario no encontrado."]}), 404
    return jsonify({"ok": True})


@app.route("/api/years")
def available_years():
    years = set()
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=False)
    if company_id is None:
        return jsonify({"years": [date.today().year]})
    with engine.connect() as conn:
        invoice_dates = conn.execute(
            select(invoices_table.c.invoice_date)
            .where(invoices_table.c.user_id == data_owner_id)
            .where(invoices_table.c.company_id == company_id)
        ).scalars().all()
        for value in invoice_dates:
            if value:
                try:
                    years.add(int(str(value)[:4]))
                except ValueError:
                    continue
        billing_years = conn.execute(
            select(facturacion_table.c.anio)
            .where(facturacion_table.c.user_id == data_owner_id)
            .where(facturacion_table.c.company_id == company_id)
        ).scalars().all()
        years.update(int(year) for year in billing_years if year)
        income_years = conn.execute(
            select(income_invoices_table.c.invoice_date)
            .where(income_invoices_table.c.user_id == data_owner_id)
            .where(income_invoices_table.c.company_id == company_id)
        ).scalars().all()
        for value in income_years:
            if value:
                try:
                    years.add(int(str(value)[:4]))
                except ValueError:
                    continue
        loan_years = conn.execute(
            select(loan_installments_table.c.payment_date)
            .where(loan_installments_table.c.user_id == data_owner_id)
            .where(loan_installments_table.c.company_id == company_id)
        ).scalars().all()
        for value in loan_years:
            if value:
                try:
                    years.add(int(str(value)[:4]))
                except ValueError:
                    continue

    if not years:
        years = {date.today().year}

    return jsonify({"years": sorted(years)})


@app.route("/api/upload", methods=["POST"])
def upload_invoices():
    data_owner_id = get_data_owner_id()
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        entries = payload.get("entries", [])
        if not entries:
            return jsonify({"ok": False, "errors": ["No se recibieron entradas."]}), 400
        company_id = get_company_id(required=True)
        if company_id is None:
            return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

        errors = []
        inserted = 0
        with engine.begin() as conn:
            for idx, entry in enumerate(entries):
                is_manual = any(
                    key in entry for key in ("supplier", "base", "vat", "total", "vatAmount")
                )
                original_name = entry.get("originalFilename") or ""
                invoice_date = entry.get("date") or date.today().isoformat()
                entry_company_id = entry.get("company_id") or entry.get("companyId")
                if entry_company_id:
                    try:
                        company_id = int(entry_company_id)
                    except (TypeError, ValueError):
                        errors.append("Empresa inválida.")
                        continue
                    if not is_company_accessible(company_id):
                        errors.append("Empresa inválida.")
                        continue
                supplier = (entry.get("supplier") or "").strip()
                base_amount = parse_amount(str(entry.get("base") or ""))
                vat_rate_raw = vat_rate_to_str(entry.get("vat"))
                vat_amount = parse_amount(str(entry.get("vatAmount") or ""))
                total_amount = parse_amount(str(entry.get("total") or ""))
                withholding_amount = parse_amount(
                    str(
                        entry.get("withholding_amount")
                        or entry.get("withholdingAmount")
                        or ""
                    )
                )
                is_rectificativa = bool(entry.get("isRectificativa"))
                vat_breakdown = parse_vat_breakdown(
                    entry.get("vatBreakdown") or entry.get("vat_breakdown")
                )
                vat_breakdown_json = json.dumps(vat_breakdown) if vat_breakdown else None
                payment_dates = parse_payment_dates(
                    entry.get("paymentDates") or entry.get("payment_dates")
                )
                payment_date = compute_payment_date(
                    invoice_date,
                    entry.get("paymentDate")
                    or entry.get("payment_date")
                    or (payment_dates[0] if payment_dates else None),
                )
                analysis_status = entry.get("analysisStatus") or entry.get("analysis_status") or "ok"
                extraction_source = (
                    entry.get("extractionSource") or entry.get("extraction_source")
                )
                confidence_score = entry.get("confidenceScore") or entry.get("confidence_score")
                expense_category = entry.get("expenseCategory") or "with_invoice"

                if not supplier:
                    app.logger.info(
                        "Proveedor vacío para %s. Se permite guardado manual.",
                        original_name,
                    )

                if vat_breakdown:
                    vat_rate_int = infer_vat_rate_from_breakdown(vat_breakdown)
                    summary = summarize_vat_breakdown(vat_breakdown)
                    if summary:
                        base_amount, vat_amount, total_amount = summary
                else:
                    try:
                        vat_rate_int = int(vat_rate_raw)
                    except ValueError:
                        errors.append(f"Tipo de IVA inválido para {original_name}.")
                        continue
                    if vat_rate_int not in {0, 4, 10, 21}:
                        errors.append(f"Tipo de IVA inválido para {original_name}.")
                        continue
                if expense_category not in {"with_invoice", "without_invoice", "non_deductible"}:
                    errors.append(f"Tipo de gasto inválido para {original_name}.")
                    continue
                if base_amount is None:
                    errors.append(f"Base imponible inválida para {original_name}.")
                    continue
                if total_amount is None:
                    errors.append(f"Total inválido para {original_name}.")
                    continue
                if withholding_amount is None:
                    withholding_amount = 0.0
                if withholding_amount < 0:
                    errors.append(f"Retención inválida para {original_name}.")
                    continue
                if withholding_amount > total_amount:
                    errors.append(
                        f"La retención no puede superar el total en {original_name}."
                    )
                    continue
                if (base_amount < 0 or total_amount < 0) and not is_rectificativa:
                    errors.append(f"Factura rectificativa no indicada en {original_name}.")
                    continue
                if supplier and is_supplier_same_as_company(supplier, company_id, conn):
                    errors.append(
                        f"El proveedor no puede ser la empresa activa ({original_name})."
                    )
                    continue

                if not is_manual and vat_rate_int is not None and vat_rate_int >= 0:
                    base_amount, vat_amount, total_amount = normalize_vat_amounts(
                        base_amount, vat_rate_int, vat_amount, total_amount
                    )
                vat_deductible = entry.get("vatDeductible")
                if vat_deductible is None:
                    vat_deductible = expense_category != "non_deductible"
                else:
                    vat_deductible = vat_deductible in (True, "true", "True", 1, "1")
                expense_profile = derive_invoice_profile(
                    expense_category,
                    vat_deductible,
                    vat_amount,
                    withholding_amount,
                )

                created_at = datetime.utcnow().isoformat()

                conn.execute(
                    invoices_table.insert().values(
                        user_id=data_owner_id,
                        company_id=company_id,
                        original_filename=original_name,
                        stored_filename="",
                        invoice_date=invoice_date,
                        supplier=supplier,
                        base_amount=base_amount,
                        vat_deductible=vat_deductible,
                        vat_rate=vat_rate_int,
                        vat_amount=vat_amount,
                        total_amount=total_amount,
                        vat_breakdown=vat_breakdown_json,
                        withholding_amount=withholding_amount,
                        payment_date=payment_date,
                        payment_dates=json.dumps(payment_dates) if payment_dates else None,
                        ocr_text=None,
                        extraction_source=extraction_source,
                        confidence_score=confidence_score,
                        expense_category=expense_category,
                        **expense_profile,
                        created_at=created_at,
                    )
                )
                inserted += 1
                if supplier and analysis_status != "low_quality_scan":
                    store_known_supplier(conn, data_owner_id, company_id, supplier)

        return jsonify({"ok": True, "inserted": inserted, "errors": errors})

    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    files = request.files.getlist("files")
    dates = request.form.getlist("date")
    suppliers = request.form.getlist("supplier")
    bases = request.form.getlist("base")
    vats = request.form.getlist("vat")
    vat_amounts = request.form.getlist("vatAmount")
    totals = request.form.getlist("total")
    payment_dates = request.form.getlist("paymentDate")

    if not files:
        return jsonify({"ok": False, "errors": ["No se recibieron archivos."]}), 400

    if not (
        len(files)
        == len(dates)
        == len(suppliers)
        == len(bases)
        == len(vats)
        == len(vat_amounts)
        == len(totals)
        == len(payment_dates)
    ):
        return (
            jsonify({"ok": False, "errors": ["Los datos no coinciden con los archivos."]}),
            400,
        )

    errors = []
    inserted = 0
    with engine.begin() as conn:
        for idx, file in enumerate(files):
            if not file or not file.filename:
                errors.append(f"Archivo vacío en posición {idx + 1}.")
                continue

            original_name = os.path.basename(file.filename)
            if not allowed_file(original_name):
                errors.append(f"Tipo de archivo no permitido: {original_name}")
                continue

            invoice_date = dates[idx] or date.today().isoformat()
            payment_dates_list = parse_payment_dates(payment_dates[idx] if payment_dates else None)
            payment_date = compute_payment_date(
                invoice_date,
                payment_dates[idx] if payment_dates else None,
            )
            supplier = suppliers[idx].strip() if suppliers[idx] else ""
            base_amount = parse_amount(bases[idx])
            vat_rate = vats[idx].strip() if vats[idx] else ""
            vat_amount = parse_amount(vat_amounts[idx])
            total_amount = parse_amount(totals[idx])
            is_rectificativa = base_amount is not None and base_amount < 0 or total_amount is not None and total_amount < 0

            if not supplier:
                app.logger.info("Proveedor vacío para %s. Se permite guardado manual.", original_name)
            try:
                vat_rate_int = int(vat_rate)
            except ValueError:
                errors.append(f"Tipo de IVA inválido para {original_name}.")
                continue
            if vat_rate_int not in {0, 4, 10, 21}:
                errors.append(f"Tipo de IVA inválido para {original_name}.")
                continue
            if base_amount is None:
                errors.append(f"Base imponible inválida para {original_name}.")
                continue
            if total_amount is None:
                errors.append(f"Total inválido para {original_name}.")
                continue
            if (base_amount < 0 or total_amount < 0) and not is_rectificativa:
                errors.append(f"Factura rectificativa no indicada en {original_name}.")
                continue
            if supplier and is_supplier_same_as_company(supplier, company_id, conn):
                errors.append(
                    f"El proveedor no puede ser la empresa activa ({original_name})."
                )
                continue

            # Guardado manual: no recalcular ni aplicar heurísticas semánticas.
            created_at = datetime.utcnow().isoformat()
            expense_profile = derive_invoice_profile("with_invoice", True, vat_amount)

            conn.execute(
                invoices_table.insert().values(
                    user_id=data_owner_id,
                    company_id=company_id,
                    original_filename=original_name,
                    stored_filename="",
                    invoice_date=invoice_date,
                    supplier=supplier,
                    base_amount=base_amount,
                    vat_deductible=True,
                    vat_rate=vat_rate_int,
                    vat_amount=vat_amount,
                    total_amount=total_amount,
                    withholding_amount=0.0,
                    payment_date=payment_date,
                    payment_dates=json.dumps(payment_dates_list) if payment_dates_list else None,
                    ocr_text=None,
                    extraction_source=None,
                    confidence_score=None,
                    expense_category="with_invoice",
                    **expense_profile,
                    created_at=created_at,
                )
            )
            inserted += 1
            # Entrada manual (sin OCR/IA) no genera aprendizaje automático.

    return jsonify({"ok": True, "inserted": inserted, "errors": errors})


@app.route("/api/analyze-invoice", methods=["POST"])
def analyze_invoice_api():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"ok": False, "errors": ["Archivo no recibido."]}), 400

    original_name = os.path.basename(file.filename)
    if not allowed_file(original_name):
        return jsonify({"ok": False, "errors": ["Tipo de archivo no permitido."]}), 400

    document_type = request.form.get("document_type") or request.args.get("document_type") or "expense"
    company_id = get_company_id(required=False)
    company_names = []
    known_suppliers = []
    if company_id and is_company_accessible(company_id):
        with engine.connect() as conn:
            row = conn.execute(
                select(companies_table.c.display_name, companies_table.c.legal_name).where(
                    companies_table.c.id == company_id
                )
            ).mappings().first()
            if row:
                company_names = [row.get("display_name"), row.get("legal_name")]
            known_suppliers = fetch_known_suppliers(conn, get_data_owner_id(), company_id)
    app.logger.info("Solicitud de análisis recibida: %s (%s)", original_name, file.mimetype)
    file_bytes = file.read()
    fallback_status = None
    mime_lower = (file.mimetype or "").lower()
    if mime_lower.startswith("image/"):
        fallback_status = "low_quality_scan"
    elif mime_lower == "application/pdf":
        if not _pdf_has_text(file_bytes, min_chars=100):
            fallback_status = "low_quality_scan"

    extracted = _analyze_invoice_with_timeout(
        file_bytes,
        original_name,
        original_name,
        file.mimetype,
        document_type=document_type,
        company_names=company_names,
        known_suppliers=known_suppliers,
        fallback_status=fallback_status,
    )

    app.logger.info(
        "AI extracted for %s: provider=%s date=%s payment=%s base=%s vat_rate=%s vat_amount=%s total=%s",
        original_name,
        extracted.get("provider_name"),
        extracted.get("invoice_date"),
        extracted.get("payment_date"),
        extracted.get("base_amount"),
        extracted.get("vat_rate"),
        extracted.get("vat_amount"),
        extracted.get("total_amount"),
    )

    return jsonify(
        {
            "ok": True,
            "storedFilename": "__transient__",
            "originalFilename": original_name,
            "companyId": company_id,
            "transientProcessing": True,
            "extracted": extracted,
        }
    )


@app.route("/api/billing", methods=["POST"])
def create_billing():
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or request.form

    month = int(payload.get("month") or 0)
    year = int(payload.get("year") or 0)
    base_amount = parse_amount(str(payload.get("base") or ""))
    vat_rate_raw = vat_rate_to_str(payload.get("vat"))
    concept = (payload.get("concept") or "").strip()
    invoice_date = payload.get("invoice_date") or payload.get("date") or ""

    errors = []
    if month < 1 or month > 12:
        errors.append("Mes inválido.")
    if year < 2000:
        errors.append("Año inválido.")
    if invoice_date:
        normalized_date = normalize_date(invoice_date)
        if normalized_date is None:
            errors.append("Fecha inválida.")
        else:
            invoice_date = normalized_date
    if base_amount is None or base_amount < 0:
        errors.append("Base facturada inválida.")
    try:
        vat_rate = int(vat_rate_raw)
    except ValueError:
        vat_rate = None
    if vat_rate not in {0, 4, 10, 21}:
        errors.append("Tipo de IVA inválido.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    iva_repercutido = round(base_amount * (vat_rate / 100), 2)
    total_amount = round(base_amount + iva_repercutido, 2)
    if invoice_date:
        try:
            month = int(invoice_date[5:7])
            year = int(invoice_date[:4])
        except (TypeError, ValueError):
            pass

    with engine.begin() as conn:
        conn.execute(
            facturacion_table.insert().values(
                user_id=data_owner_id,
                company_id=company_id,
                mes=month,
                anio=year,
                invoice_date=invoice_date or None,
                concept=concept or None,
                base_facturada=base_amount,
                tipo_iva=vat_rate,
                iva_repercutido=iva_repercutido,
                total_amount=total_amount,
            )
        )

    return jsonify({"ok": True})


@app.route("/api/billing/summary")
def billing_summary():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                facturacion_table.c.tipo_iva,
                func.sum(facturacion_table.c.base_facturada).label("base_total"),
                func.sum(facturacion_table.c.iva_repercutido).label("vat_total"),
            )
            .where(
                facturacion_table.c.mes == month,
                facturacion_table.c.anio == year,
                facturacion_table.c.user_id == data_owner_id,
                facturacion_table.c.company_id == company_id,
            )
            .group_by(facturacion_table.c.tipo_iva)
        ).mappings().all()

    base_totals = {0: 0.0, 4: 0.0, 10: 0.0, 21: 0.0}
    vat_totals = {0: 0.0, 4: 0.0, 10: 0.0, 21: 0.0}

    for row in rows:
        vat_rate = int(row["tipo_iva"])
        base_totals[vat_rate] = float(row["base_total"] or 0)
        vat_totals[vat_rate] = float(row["vat_total"] or 0)

    total_vat = round(sum(vat_totals.values()), 2)

    return jsonify(
        {
            "baseTotals": {
                "0": round(base_totals[0], 2),
                "4": round(base_totals[4], 2),
                "10": round(base_totals[10], 2),
                "21": round(base_totals[21], 2),
            },
            "vatTotals": {
                "0": round(vat_totals[0], 2),
                "4": round(vat_totals[4], 2),
                "10": round(vat_totals[10], 2),
                "21": round(vat_totals[21], 2),
            },
            "totalVat": total_vat,
        }
    )


@app.route("/api/invoices")
def list_invoices():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    _, last_day = calendar.monthrange(year, month)
    start = date(year, month, 1).isoformat()
    end = date(year, month, last_day).isoformat()

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                invoices_table.c.id,
                invoices_table.c.invoice_date,
                invoices_table.c.supplier,
                invoices_table.c.base_amount,
                invoices_table.c.vat_deductible,
                invoices_table.c.vat_rate,
                invoices_table.c.vat_amount,
                invoices_table.c.total_amount,
                invoices_table.c.vat_breakdown,
                invoices_table.c.withholding_amount,
                invoices_table.c.payment_date,
                invoices_table.c.payment_dates,
                invoices_table.c.payment_completed_dates,
                invoices_table.c.extraction_source,
                invoices_table.c.confidence_score,
                invoices_table.c.original_filename,
                invoices_table.c.expense_category,
                invoices_table.c.expense_family,
                invoices_table.c.expense_subtype,
                invoices_table.c.pnl_bucket,
                invoices_table.c.tax_model_targets,
            )
            .where(invoices_table.c.user_id == data_owner_id)
            .where(invoices_table.c.company_id == company_id)
            .where(invoices_table.c.invoice_date.between(start, end))
            .order_by(invoices_table.c.invoice_date.desc(), invoices_table.c.id.desc())
        ).mappings().all()

    invoices = [
        {
            "id": row["id"],
            "invoice_date": row["invoice_date"],
            "payment_date": row["payment_date"]
            or compute_payment_date(row["invoice_date"], row["payment_date"]),
            "payment_dates": parse_payment_dates(row.get("payment_dates"))
            or ([row["payment_date"]] if row.get("payment_date") else []),
            "payment_completed_dates": parse_payment_dates(row.get("payment_completed_dates")),
            "supplier": row["supplier"],
            "base_amount": float(row["base_amount"]),
            "vat_deductible": bool(row.get("vat_deductible")) if row.get("vat_deductible") is not None else True,
            "vat_rate": int(row["vat_rate"]) if row["vat_rate"] is not None and row["vat_rate"] >= 0 else None,
            "vat_amount": float(row["vat_amount"]) if row["vat_amount"] is not None else None,
            "total_amount": float(row["total_amount"]),
            "vat_breakdown": row["vat_breakdown"],
            "withholding_amount": float(row["withholding_amount"] or 0),
            "extraction_source": row.get("extraction_source"),
            "confidence_score": float(row["confidence_score"]) if row["confidence_score"] is not None else None,
            "original_filename": row["original_filename"],
            "expense_category": row["expense_category"] or "with_invoice",
            "expense_family": row.get("expense_family"),
            "expense_subtype": row.get("expense_subtype"),
            "pnl_bucket": row.get("pnl_bucket"),
            "tax_model_targets": parse_tax_model_targets(row.get("tax_model_targets")),
        }
        for row in rows
    ]

    return jsonify({"invoices": invoices})


@app.route("/api/payments")
def list_payments():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    _, last_day = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, last_day)
    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    buffer_start = (year_start - timedelta(days=31)).isoformat()
    year_start_iso = year_start.isoformat()
    year_end_iso = year_end.isoformat()

    with engine.connect() as conn:
        company = _fetch_company_fiscal_profile(conn, company_id)
        expense_rows = conn.execute(
            select(
                invoices_table.c.id,
                invoices_table.c.invoice_date,
                invoices_table.c.payment_date,
                invoices_table.c.payment_dates,
                invoices_table.c.payment_completed_dates,
                invoices_table.c.supplier,
                invoices_table.c.base_amount,
                invoices_table.c.vat_rate,
                invoices_table.c.vat_amount,
                invoices_table.c.total_amount,
                invoices_table.c.original_filename,
                invoices_table.c.withholding_amount,
                invoices_table.c.expense_category,
                invoices_table.c.expense_family,
                invoices_table.c.expense_subtype,
                invoices_table.c.pnl_bucket,
                invoices_table.c.tax_model_targets,
            )
            .where(
                (
                    invoices_table.c.payment_date.between(year_start_iso, year_end_iso)
                )
                | (
                    invoices_table.c.payment_date.is_(None)
                    & invoices_table.c.invoice_date.between(buffer_start, year_end_iso)
                )
            )
            .where(invoices_table.c.user_id == data_owner_id)
            .where(invoices_table.c.company_id == company_id)
            .order_by(invoices_table.c.invoice_date.desc(), invoices_table.c.id.desc())
        ).mappings().all()

        no_invoice_rows = conn.execute(
            select(
                no_invoice_table.c.id,
                no_invoice_table.c.expense_date,
                no_invoice_table.c.payment_date,
                no_invoice_table.c.payment_dates,
                no_invoice_table.c.payment_completed_dates,
                no_invoice_table.c.concept,
                no_invoice_table.c.amount,
                no_invoice_table.c.interest_amount,
                no_invoice_table.c.vat_deductible,
                no_invoice_table.c.vat_rate,
                no_invoice_table.c.vat_amount,
                no_invoice_table.c.base_amount,
                no_invoice_table.c.withholding_amount,
                no_invoice_table.c.payroll_employee_name,
                no_invoice_table.c.payroll_period,
                no_invoice_table.c.payroll_net_amount,
                no_invoice_table.c.payroll_total_deductions_amount,
                no_invoice_table.c.payroll_employer_cost_amount,
                no_invoice_table.c.expense_type,
                no_invoice_table.c.deductible,
                no_invoice_table.c.payment_completed_dates,
                no_invoice_table.c.expense_family,
                no_invoice_table.c.expense_subtype,
                no_invoice_table.c.pnl_bucket,
                no_invoice_table.c.tax_model_targets,
            )
            .where(no_invoice_table.c.user_id == data_owner_id)
            .where(
                (no_invoice_table.c.company_id == company_id)
                | (no_invoice_table.c.company_id.is_(None))
            )
            .where(no_invoice_table.c.expense_date.between(buffer_start, year_end_iso))
            .order_by(no_invoice_table.c.expense_date.desc(), no_invoice_table.c.id.desc())
        ).mappings().all()

        loan_rows = conn.execute(
            select(
                loan_installments_table.c.id,
                loan_installments_table.c.payment_date,
                loan_installments_table.c.payment_completed_dates,
                loan_installments_table.c.bank_name,
                loan_installments_table.c.concept,
                loan_installments_table.c.total_amount,
                loan_installments_table.c.interest_amount,
                loan_installments_table.c.principal_amount,
            )
            .where(loan_installments_table.c.user_id == data_owner_id)
            .where(loan_installments_table.c.company_id == company_id)
            .where(
                loan_installments_table.c.payment_date.between(buffer_start, year_end_iso)
            )
            .order_by(
                loan_installments_table.c.payment_date.desc(),
                loan_installments_table.c.id.desc(),
            )
        ).mappings().all()

        income_rows = conn.execute(
            select(
                income_invoices_table.c.id,
                income_invoices_table.c.invoice_date,
                income_invoices_table.c.payment_date,
                income_invoices_table.c.payment_dates,
                income_invoices_table.c.payment_completed_dates,
                income_invoices_table.c.client,
                income_invoices_table.c.base_amount,
                income_invoices_table.c.vat_rate,
                income_invoices_table.c.vat_amount,
                income_invoices_table.c.total_amount,
                income_invoices_table.c.vat_breakdown,
                income_invoices_table.c.original_filename,
            )
            .where(
                (
                    income_invoices_table.c.payment_date.between(year_start_iso, year_end_iso)
                )
                | (
                    income_invoices_table.c.payment_date.is_(None)
                    & income_invoices_table.c.invoice_date.between(buffer_start, year_end_iso)
                )
            )
            .where(income_invoices_table.c.user_id == data_owner_id)
            .where(income_invoices_table.c.company_id == company_id)
            .order_by(income_invoices_table.c.invoice_date.desc(), income_invoices_table.c.id.desc())
        ).mappings().all()
        tax_obligation_rows = _build_fiscal_calendar_items(
            conn,
            data_owner_id,
            company_id,
            month,
            year,
        )

    items = []
    day_totals = {}
    today_iso = today.isoformat()
    today_pending = []
    overdue_pending_count = 0
    for row in expense_rows:
        payment_dates = parse_payment_dates(row.get("payment_dates"))
        completed_dates = parse_payment_dates(row.get("payment_completed_dates"))
        if not payment_dates:
            fallback = row["payment_date"] or compute_payment_date(row["invoice_date"], None)
            if fallback:
                payment_dates = [fallback]
        if not payment_dates:
            continue
        total_amount = float(row["total_amount"] or 0)
        withholding_amount = float(row.get("withholding_amount") or 0)
        payable_amount = max(round(total_amount - withholding_amount, 2), 0.0)
        split_count = len(payment_dates)
        base_amount = round(payable_amount / split_count, 2) if split_count else payable_amount
        amounts = [base_amount] * split_count
        if split_count > 1:
            amounts[-1] = round(payable_amount - base_amount * (split_count - 1), 2)
        for payment_date, amount in zip(payment_dates, amounts):
            try:
                payment_dt = date.fromisoformat(payment_date)
            except ValueError:
                continue
            status = resolve_payment_status("expense", payment_date, completed_dates, today_iso)
            if status == "due_today":
                today_pending.append(
                    {
                        "id": row["id"],
                        "type": "expense",
                        "counterparty": row["supplier"],
                        "concept": row["original_filename"],
                        "payment_date": payment_date,
                        "amount": amount,
                        "status": status,
                    }
                )
            elif status == "overdue":
                overdue_pending_count += 1
            if payment_dt < start or payment_dt > end:
                continue
            day = payment_dt.day
            day_totals[day] = round(day_totals.get(day, 0.0) + amount, 2)
            items.append(
                {
                    "id": row["id"],
                    "counterparty": row["supplier"],
                    "concept": row["original_filename"],
                    "payment_date": payment_date,
                    "payment_dates": payment_dates,
                    "payment_completed_dates": completed_dates,
                    "invoice_date": row["invoice_date"],
                    "base_amount": float(row["base_amount"] or 0),
                    "vat_rate": int(row["vat_rate"]) if row["vat_rate"] is not None and row["vat_rate"] >= 0 else None,
                    "vat_amount": float(row["vat_amount"] or 0)
                    if row["vat_amount"] is not None
                    else None,
                    "total_amount": total_amount,
                    "withholding_amount": withholding_amount,
                    "expense_category": row["expense_category"] or "with_invoice",
                    "expense_family": row.get("expense_family"),
                    "expense_subtype": row.get("expense_subtype"),
                    "pnl_bucket": row.get("pnl_bucket"),
                    "tax_model_targets": parse_tax_model_targets(row.get("tax_model_targets")),
                    "amount": amount,
                    "status": status,
                    "type": "expense",
                }
            )

    for row in no_invoice_rows:
        payment_dates = parse_payment_dates(row.get("payment_dates"))
        completed_dates = parse_payment_dates(row.get("payment_completed_dates"))
        if not payment_dates:
            fallback_payment_date = row.get("payment_date") or row.get("expense_date")
            if fallback_payment_date:
                payment_dates = [fallback_payment_date]
        if not payment_dates:
            continue
        withholding_amount = float(row.get("withholding_amount") or 0)
        gross_amount = float(row.get("amount") or 0)
        payroll_net_amount = float(row.get("payroll_net_amount") or 0)
        payable_amount = (
            payroll_net_amount
            if row.get("expense_type") == "nomina" and payroll_net_amount > 0
            else max(round(gross_amount - withholding_amount, 2), 0.0)
        )
        split_count = len(payment_dates)
        split_amount = round(payable_amount / split_count, 2)
        amounts = [split_amount] * split_count
        if split_count > 1:
            amounts[-1] = round(
                payable_amount - split_amount * (split_count - 1),
                2,
            )
        for payment_date, amount in zip(payment_dates, amounts):
            try:
                payment_dt = date.fromisoformat(payment_date)
            except ValueError:
                continue
            status = resolve_payment_status("no_invoice", payment_date, completed_dates, today_iso)
            if status == "due_today":
                today_pending.append(
                    {
                        "id": row["id"],
                        "type": "no_invoice",
                        "counterparty": row.get("concept"),
                        "concept": row.get("concept"),
                        "payment_date": payment_date,
                        "amount": amount,
                        "status": status,
                    }
                )
            elif status == "overdue":
                overdue_pending_count += 1
            if payment_dt < start or payment_dt > end:
                continue
            day = payment_dt.day
            day_totals[day] = round(day_totals.get(day, 0.0) + amount, 2)
            items.append(
                {
                    "id": row["id"],
                    "counterparty": row.get("concept"),
                    "concept": row.get("concept"),
                    "payment_date": payment_date,
                    "payment_dates": payment_dates,
                    "payment_completed_dates": completed_dates,
                    "invoice_date": row.get("expense_date"),
                    "base_amount": amount,
                    "vat_rate": 0,
                    "vat_amount": 0,
                    "total_amount": payable_amount,
                    "expense_category": "without_invoice",
                    "expense_type": row.get("expense_type"),
                    "interest_amount": float(row.get("interest_amount") or 0),
                    "withholding_amount": withholding_amount,
                    "payroll_employee_name": row.get("payroll_employee_name"),
                    "payroll_period": row.get("payroll_period"),
                    "payroll_net_amount": payroll_net_amount,
                    "payroll_total_deductions_amount": float(
                        row.get("payroll_total_deductions_amount") or 0
                    ),
                    "payroll_employer_cost_amount": float(
                        row.get("payroll_employer_cost_amount") or 0
                    ),
                    "vat_deductible": bool(row.get("vat_deductible"))
                    if row.get("vat_deductible") is not None
                    else False,
                    "vat_rate_no_invoice": int(row.get("vat_rate"))
                    if row.get("vat_rate") is not None
                    else None,
                    "vat_amount_no_invoice": float(row.get("vat_amount") or 0),
                    "base_amount_no_invoice": float(row.get("base_amount") or row.get("amount") or 0),
                    "deductible": bool(row.get("deductible")),
                    "expense_family": row.get("expense_family"),
                    "expense_subtype": row.get("expense_subtype"),
                    "pnl_bucket": row.get("pnl_bucket"),
                    "tax_model_targets": parse_tax_model_targets(row.get("tax_model_targets")),
                    "amount": amount,
                    "gross_amount": gross_amount,
                    "status": status,
                    "type": "no_invoice",
                }
            )

    for row in loan_rows:
        payment_date = row.get("payment_date")
        completed_dates = parse_payment_dates(row.get("payment_completed_dates"))
        if not payment_date:
            continue
        try:
            payment_dt = date.fromisoformat(payment_date)
        except ValueError:
            continue
        status = resolve_payment_status(
            "loan_installment", payment_date, completed_dates, today_iso
        )
        if status == "due_today":
            today_pending.append(
                {
                    "id": row["id"],
                    "type": "loan_installment",
                    "counterparty": row.get("bank_name") or row.get("concept"),
                    "concept": row.get("concept"),
                    "payment_date": payment_date,
                    "amount": float(row.get("total_amount") or 0),
                    "status": status,
                }
            )
        elif status == "overdue":
            overdue_pending_count += 1
        if payment_dt < start or payment_dt > end:
            continue
        day = payment_dt.day
        amount = float(row.get("total_amount") or 0)
        day_totals[day] = round(day_totals.get(day, 0.0) + amount, 2)
        items.append(
            {
                "id": row["id"],
                "counterparty": row.get("bank_name") or row.get("concept"),
                "concept": row.get("concept"),
                "bank_name": row.get("bank_name"),
                "payment_date": payment_date,
                "payment_dates": [payment_date],
                "payment_completed_dates": completed_dates,
                "invoice_date": payment_date,
                "base_amount": float(row.get("principal_amount") or 0),
                "vat_rate": 0,
                "vat_amount": 0,
                "total_amount": amount,
                "interest_amount": float(row.get("interest_amount") or 0),
                "principal_amount": float(row.get("principal_amount") or 0),
                "amount": amount,
                "status": status,
                "type": "loan_installment",
            }
        )

    for row in income_rows:
        payment_dates = parse_payment_dates(row.get("payment_dates"))
        if not payment_dates:
            fallback = row["payment_date"] or compute_payment_date(row["invoice_date"], None)
            if fallback:
                payment_dates = [fallback]
        if not payment_dates:
            continue
        total_amount = float(row["total_amount"] or 0)
        split_count = len(payment_dates)
        base_amount = round(total_amount / split_count, 2) if split_count else total_amount
        amounts = [base_amount] * split_count
        if split_count > 1:
            amounts[-1] = round(total_amount - base_amount * (split_count - 1), 2)
        for payment_date, amount in zip(payment_dates, amounts):
            try:
                payment_dt = date.fromisoformat(payment_date)
            except ValueError:
                continue
            if payment_dt < start or payment_dt > end:
                continue
            day = payment_dt.day
            day_totals[day] = round(day_totals.get(day, 0.0) + amount, 2)
            items.append(
                {
                    "id": row["id"],
                    "counterparty": row["client"],
                    "concept": row["original_filename"],
                    "payment_date": payment_date,
                    "payment_dates": payment_dates,
                    "payment_completed_dates": parse_payment_dates(
                        row.get("payment_completed_dates")
                    ),
                    "invoice_date": row["invoice_date"],
                    "base_amount": float(row["base_amount"] or 0),
                    "vat_rate": int(row["vat_rate"]) if row["vat_rate"] is not None and row["vat_rate"] >= 0 else None,
                    "vat_amount": float(row["vat_amount"] or 0)
                    if row["vat_amount"] is not None
                    else None,
                    "total_amount": total_amount,
                    "amount": amount,
                    "status": None,
                    "type": "income",
                }
            )

    for item in tax_obligation_rows:
        try:
            payment_dt = date.fromisoformat(item["payment_date"])
        except (TypeError, ValueError):
            continue
        if payment_dt < start or payment_dt > end:
            continue
        day = payment_dt.day
        amount = float(item.get("amount") or 0)
        day_totals[day] = round(day_totals.get(day, 0.0) + amount, 2)
        items.append(item)

    today_pending.sort(key=lambda item: (item["payment_date"], item.get("counterparty") or ""))
    items.sort(key=lambda item: (item.get("payment_date") or "", item.get("counterparty") or ""))

    return jsonify(
        {
            "items": items,
            "dayTotals": day_totals,
            "todayPending": today_pending,
            "overduePendingCount": overdue_pending_count,
        }
    )


@app.route("/api/fiscal-models/summary")
def fiscal_models_summary():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    period = (request.args.get("period") or "monthly").strip().lower()
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year
    if period not in {"monthly", "quarterly"}:
        period = "monthly"

    with engine.connect() as conn:
        company = _fetch_company_fiscal_profile(conn, company_id)
        if not company:
            return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
        rows = _build_fiscal_model_rows(
            conn,
            data_owner_id,
            company_id,
            month,
            year,
            period,
        )

    return jsonify({"ok": True, "rows": rows})


@app.route("/api/reports/quarterly")
def quarterly_report():
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    year, quarter, start_month, end_month = _parse_period_params()
    months = _get_months_for_period(year, quarter, start_month, end_month)
    if not months:
        return jsonify({"ok": False, "errors": ["Periodo inválido."]}), 400

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.tax_id,
            companies_table.c.company_type,
            companies_table.c.email,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(
            company_query
        ).mappings().first()
    if not company:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404

    totals = _build_report_totals(data_owner_id, company_id, months, year)
    period_label = _report_period_label(year, months, quarter)
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    html = f"""
    <html>
      <head>
        <meta charset="utf-8" />
        <title>Informe trimestral {period_label}</title>
        <style>
          body {{ font-family: Arial, sans-serif; margin: 32px; color: #1f2937; }}
          h1 {{ font-size: 20px; margin-bottom: 4px; }}
          p {{ margin: 4px 0; }}
          table {{ border-collapse: collapse; width: 100%; margin-top: 16px; }}
          th, td {{ border: 1px solid #e5e7eb; padding: 8px; text-align: left; }}
          th {{ background: #f9fafb; }}
        </style>
      </head>
      <body>
        <h1>Informe fiscal {period_label}</h1>
        <p><strong>Empresa:</strong> {company["display_name"]} ({company["legal_name"]})</p>
        <p><strong>CIF/NIF:</strong> {company["tax_id"]}</p>
        <p><strong>Generado:</strong> {generated_at}</p>
        <table>
          <tr><th>Concepto</th><th>Importe (€)</th></tr>
          <tr><td>Ingresos base</td><td>{totals["income_base"]:.2f}</td></tr>
          <tr><td>IVA repercutido</td><td>{totals["income_vat"]:.2f}</td></tr>
          <tr><td>Gastos deducibles</td><td>{totals["expense_base"]:.2f}</td></tr>
          <tr><td>IVA soportado</td><td>{totals["expense_vat"]:.2f}</td></tr>
          <tr><td>Resultado neto</td><td>{totals["net_result"]:.2f}</td></tr>
          <tr><td>Resultado IVA</td><td>{totals["vat_result"]:.2f}</td></tr>
        </table>
      </body>
    </html>
    """
    response = app.response_class(html, mimetype="text/html")
    filename = f"informe_{period_label.replace(' ', '_')}.html"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/api/reports/quarterly/email", methods=["POST"])
def quarterly_report_email():
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    payload = request.get_json(silent=True) or {}
    year = payload.get("year")
    quarter = payload.get("quarter")
    start_month = payload.get("start_month")
    end_month = payload.get("end_month")
    try:
        year = int(year)
    except (TypeError, ValueError):
        year = None
    try:
        quarter = int(quarter) if quarter else None
    except (TypeError, ValueError):
        quarter = None
    try:
        start_month = int(start_month) if start_month else None
    except (TypeError, ValueError):
        start_month = None
    try:
        end_month = int(end_month) if end_month else None
    except (TypeError, ValueError):
        end_month = None

    months = _get_months_for_period(year, quarter, start_month, end_month)
    if not months:
        return jsonify({"ok": False, "errors": ["Periodo inválido."]}), 400

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.tax_id,
            companies_table.c.company_type,
            companies_table.c.email,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(company_query).mappings().first()
        user = conn.execute(
            select(users_table.c.email).where(users_table.c.id == get_current_user_id())
        ).mappings().first()
    if not company:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
    if not company.get("email"):
        return jsonify({"ok": False, "errors": ["La empresa no tiene email."]}), 400

    totals = _build_report_totals(data_owner_id, company_id, months, year)
    period_label = _report_period_label(year, months, quarter)
    html = f"""
    <h2>Informe fiscal {period_label}</h2>
    <p><strong>Empresa:</strong> {company["display_name"]} ({company["legal_name"]})</p>
    <p><strong>CIF/NIF:</strong> {company["tax_id"]}</p>
    <ul>
      <li>Ingresos base: {totals["income_base"]:.2f} €</li>
      <li>IVA repercutido: {totals["income_vat"]:.2f} €</li>
      <li>Gastos deducibles: {totals["expense_base"]:.2f} €</li>
      <li>IVA soportado: {totals["expense_vat"]:.2f} €</li>
      <li>Resultado neto: {totals["net_result"]:.2f} €</li>
      <li>Resultado IVA: {totals["vat_result"]:.2f} €</li>
    </ul>
    """
    reply_to = user["email"] if user else None
    sent = send_email(
        company["email"],
        f"Informe fiscal {period_label}",
        html,
        reply_to=reply_to,
        from_email=get_reports_from_email(),
    )
    if not sent:
        return jsonify({"ok": False, "errors": ["No se pudo enviar el email."]}), 500
    return jsonify({"ok": True})


@app.route("/api/pnl/email", methods=["POST"])
def pnl_email():
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    tax_id = (payload.get("tax_id") or "").strip()
    period_label = (payload.get("period_label") or "").strip()
    lines = payload.get("lines") or []
    totals = payload.get("totals") or {}

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.tax_id,
            companies_table.c.email,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(company_query).mappings().first()
        user = conn.execute(
            select(users_table.c.email).where(users_table.c.id == get_current_user_id())
        ).mappings().first()
    if not company:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
    if not company.get("email"):
        return jsonify({"ok": False, "errors": ["La empresa no tiene email."]}), 400

    company_name = name or company.get("display_name") or company.get("legal_name") or ""
    company_tax = tax_id or company.get("tax_id") or ""

    rows_html = ""
    for item in lines:
        if not isinstance(item, dict):
            continue
        label = item.get("label") or item.get("id") or ""
        value = item.get("value")
        try:
            value = float(value)
            value_str = f"{value:.2f} €"
        except (TypeError, ValueError):
            value_str = str(value or "")
        rows_html += f"<tr><td>{label}</td><td style='text-align:right'>{value_str}</td></tr>"

    html = f"""
    <h2>Cuenta de pérdidas y ganancias (estimada)</h2>
    <p><strong>Empresa:</strong> {company_name}</p>
    <p><strong>CIF/NIF:</strong> {company_tax}</p>
    <p><strong>Periodo:</strong> {period_label or '-'}</p>
    <table style="border-collapse:collapse;width:100%;margin-top:12px">
      <thead>
        <tr>
          <th style="border:1px solid #e5e7eb;padding:8px;text-align:left">Concepto</th>
          <th style="border:1px solid #e5e7eb;padding:8px;text-align:right">Importe (€)</th>
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
    </table>
    <p><strong>Resultado de explotación:</strong> {totals.get("operating", "")}</p>
    <p><strong>Resultado financiero:</strong> {totals.get("financial", "")}</p>
    <p><strong>Resultado antes de impuestos:</strong> {totals.get("pretax", "")}</p>
    <p><strong>Resultado del ejercicio:</strong> {totals.get("net", "")}</p>
    """
    reply_to = user["email"] if user else None
    sent = send_email(
        company["email"],
        f"P&G {period_label}".strip(),
        html,
        reply_to=reply_to,
        from_email=get_reports_from_email(),
    )
    if not sent:
        return jsonify({"ok": False, "errors": ["No se pudo enviar el P&G."]}), 500
    return jsonify({"ok": True})


@app.route("/api/balance/email", methods=["POST"])
def balance_email():
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    tax_id = (payload.get("tax_id") or "").strip()
    period_label = (payload.get("period_label") or "").strip()
    lines = payload.get("lines") or []

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.tax_id,
            companies_table.c.email,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(company_query).mappings().first()
        user = conn.execute(
            select(users_table.c.email).where(users_table.c.id == get_current_user_id())
        ).mappings().first()
    if not company:
        return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
    if not company.get("email"):
        return jsonify({"ok": False, "errors": ["La empresa no tiene email."]}), 400

    company_name = name or company.get("display_name") or company.get("legal_name") or ""
    company_tax = tax_id or company.get("tax_id") or ""

    rows_html = ""
    for item in lines:
        if not isinstance(item, dict):
            continue
        label = item.get("label") or ""
        value = item.get("value")
        try:
            value_num = float(value)
            value = f"{value_num:.2f} €"
        except (TypeError, ValueError):
            value = value or ""
        rows_html += (
            f"<tr><td style='border:1px solid #e5e7eb;padding:8px'>{label}</td>"
            f"<td style='border:1px solid #e5e7eb;padding:8px;text-align:right'>{value}</td></tr>"
        )

    html = f"""
    <h2>Balance de situación (estimado)</h2>
    <p><strong>Empresa:</strong> {company_name}</p>
    <p><strong>CIF/NIF:</strong> {company_tax}</p>
    <p><strong>Periodo:</strong> {period_label or '-'}</p>
    <table style="border-collapse:collapse;width:100%;margin-top:12px">
      <tr>
        <th style="border:1px solid #e5e7eb;padding:8px;text-align:left">Concepto</th>
        <th style="border:1px solid #e5e7eb;padding:8px;text-align:right">Importe (€)</th>
      </tr>
      {rows_html}
    </table>
    """
    reply_to = user["email"] if user else None
    sent = send_email(
        company["email"],
        f"Balance de situación {period_label}".strip(),
        html,
        reply_to=reply_to,
        from_email=get_reports_from_email(),
    )
    if not sent:
        return jsonify({"ok": False, "errors": ["No se pudo enviar el balance."]}), 500
    return jsonify({"ok": True})


@app.route("/api/invoices/<int:invoice_id>", methods=["PUT"])
def update_invoice(invoice_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    payload = request.get_json(silent=True) or {}

    payment_only = payload.get("payment_only") or payload.get("paymentOnly")
    invoice_date = payload.get("invoice_date") or ""
    payment_dates_payload = payload.get("payment_dates") or payload.get("paymentDates")
    payment_dates = parse_payment_dates(payment_dates_payload)
    payment_date_input = (
        payload.get("payment_date")
        or payload.get("paymentDate")
        or (payment_dates[0] if payment_dates else None)
    )
    payment_date = compute_payment_date(invoice_date, payment_date_input)
    if payment_only:
        mark_paid = bool(payload.get("mark_paid") or payload.get("markPaid"))
        mark_unpaid = bool(payload.get("mark_unpaid") or payload.get("markUnpaid"))
        reference_date = normalize_date(
            payload.get("payment_reference_date") or payload.get("paymentReferenceDate")
        )
        if not payment_date_input and not mark_paid and not mark_unpaid:
            return jsonify({"ok": False, "errors": ["Fecha de pago obligatoria."]}), 400
        with engine.begin() as conn:
            current_row = conn.execute(
                select(
                    invoices_table.c.payment_date,
                    invoices_table.c.payment_dates,
                    invoices_table.c.payment_completed_dates,
                    invoices_table.c.invoice_date,
                )
                .where(invoices_table.c.id == invoice_id)
                .where(invoices_table.c.user_id == data_owner_id)
                .where(invoices_table.c.company_id == company_id)
            ).mappings().first()
            if not current_row:
                return jsonify({"ok": False, "errors": ["Factura no encontrada."]}), 404
            existing_dates = parse_payment_dates(current_row.get("payment_dates"))
            if not existing_dates:
                fallback = current_row.get("payment_date") or compute_payment_date(
                    current_row.get("invoice_date"), None
                )
                if fallback:
                    existing_dates = [fallback]
            updated_dates = payment_dates if payment_dates_payload is not None else existing_dates
            if payment_date_input and reference_date and reference_date != normalize_date(payment_date_input):
                updated_dates = replace_payment_date(updated_dates, reference_date, payment_date_input)
            elif payment_date_input and not updated_dates:
                updated_dates = [normalize_date(payment_date_input)]
            completed_dates = parse_payment_dates(current_row.get("payment_completed_dates"))
            if payment_date_input and reference_date and reference_date != normalize_date(payment_date_input):
                completed_dates = replace_payment_date(
                    completed_dates, reference_date, payment_date_input
                )
            if mark_paid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    None,
                    reference_date or normalize_date(payment_date_input),
                )
            if mark_unpaid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    reference_date or normalize_date(payment_date_input),
                    None,
                )
            updates = {
                "payment_date": updated_dates[0] if updated_dates else normalize_date(payment_date_input),
                "payment_dates": serialize_payment_dates(updated_dates),
                "payment_completed_dates": serialize_payment_dates(completed_dates),
            }
            result = conn.execute(
                invoices_table.update()
                .where(invoices_table.c.id == invoice_id)
                .where(invoices_table.c.user_id == data_owner_id)
                .where(invoices_table.c.company_id == company_id)
                .values(**updates)
            )
        return jsonify({"ok": True})
    supplier = (payload.get("supplier") or "").strip()
    base_amount = parse_amount(str(payload.get("base_amount") or ""))
    vat_rate_raw = vat_rate_to_str(payload.get("vat_rate"))
    vat_amount = parse_amount(str(payload.get("vat_amount") or ""))
    total_amount = parse_amount(str(payload.get("total_amount") or ""))
    withholding_amount = parse_amount(
        str(payload.get("withholding_amount") or payload.get("withholdingAmount") or "")
    )
    is_rectificativa = bool(payload.get("is_rectificativa") or payload.get("isRectificativa"))
    is_rectificativa = bool(payload.get("is_rectificativa") or payload.get("isRectificativa"))
    vat_breakdown = parse_vat_breakdown(
        payload.get("vat_breakdown") or payload.get("vatBreakdown")
    )
    vat_breakdown_json = json.dumps(vat_breakdown) if vat_breakdown else None
    expense_category = payload.get("expense_category") or "with_invoice"
    vat_deductible = payload.get("vat_deductible")

    errors = []
    if not invoice_date:
        errors.append("Fecha obligatoria.")
    if not supplier:
        app.logger.info("Proveedor vacío en actualización de factura %s.", invoice_id)
    if supplier and company_id:
        with engine.connect() as conn:
            if is_supplier_same_as_company(supplier, company_id, conn):
                errors.append("El proveedor no puede ser la empresa activa.")
    if base_amount is None and total_amount is None:
        errors.append("Base imponible o total obligatorio.")
    if base_amount is not None and total_amount is not None:
        if (base_amount < 0 or total_amount < 0) and not is_rectificativa:
            errors.append("Factura rectificativa no indicada.")
    if vat_breakdown:
        vat_rate = infer_vat_rate_from_breakdown(vat_breakdown)
        summary = summarize_vat_breakdown(vat_breakdown)
        if summary:
            base_amount, vat_amount, total_amount = summary
    else:
        try:
            vat_rate = int(vat_rate_raw)
        except ValueError:
            vat_rate = None
        if vat_rate not in {0, 4, 10, 21}:
            errors.append("Tipo de IVA inválido.")
    if expense_category not in {"with_invoice", "without_invoice", "non_deductible"}:
        errors.append("Tipo de gasto inválido.")
    if vat_deductible is None:
        vat_deductible = expense_category != "non_deductible"
    else:
        vat_deductible = vat_deductible in (True, "true", "True", 1, "1")
    if expense_category == "non_deductible":
        vat_deductible = False
    if withholding_amount is None:
        withholding_amount = 0.0
    if withholding_amount < 0:
        errors.append("Retención inválida.")
    elif total_amount is not None and withholding_amount > total_amount:
        errors.append("La retención no puede superar el total.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    if vat_rate is not None and vat_rate >= 0:
        base_amount, vat_amount, total_amount = normalize_vat_amounts(
            base_amount, vat_rate, vat_amount, total_amount
        )
    expense_profile = derive_invoice_profile(
        expense_category, vat_deductible, vat_amount, withholding_amount
    )

    updates = {
        "invoice_date": invoice_date,
        "supplier": supplier,
        "base_amount": base_amount,
        "vat_deductible": vat_deductible,
        "vat_rate": vat_rate,
        "vat_amount": vat_amount,
        "total_amount": total_amount,
        "vat_breakdown": vat_breakdown_json,
        "withholding_amount": withholding_amount,
        "payment_date": payment_date,
        "expense_category": expense_category,
        **expense_profile,
    }
    if payment_dates_payload is not None:
        updates["payment_dates"] = serialize_payment_dates(payment_dates)
    if "payment_completed_dates" in payload or "paymentCompletedDates" in payload:
        updates["payment_completed_dates"] = serialize_payment_dates(
            payload.get("payment_completed_dates") or payload.get("paymentCompletedDates")
        )

    with engine.begin() as conn:
        result = conn.execute(
            invoices_table.update()
            .where(invoices_table.c.id == invoice_id)
            .where(invoices_table.c.user_id == data_owner_id)
            .where(invoices_table.c.company_id == company_id)
            .values(**updates)
        )
        if result.rowcount and supplier:
            store_known_supplier(conn, data_owner_id, company_id, supplier)

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Factura no encontrada."]}), 404

    return jsonify(
        {
            "ok": True,
            "invoice": {
                "id": invoice_id,
                "invoice_date": invoice_date,
                "payment_date": payment_date,
                "payment_dates": payment_dates if payment_dates_payload is not None else None,
                "supplier": supplier,
                "base_amount": base_amount,
                "vat_deductible": vat_deductible,
                "vat_rate": vat_rate,
                "vat_amount": vat_amount,
                "total_amount": total_amount,
                "vat_breakdown": vat_breakdown,
                "withholding_amount": withholding_amount,
                "expense_category": expense_category,
                "vat_deductible": vat_deductible,
                "expense_family": expense_profile.get("expense_family"),
                "expense_subtype": expense_profile.get("expense_subtype"),
                "pnl_bucket": expense_profile.get("pnl_bucket"),
                "tax_model_targets": parse_tax_model_targets(
                    expense_profile.get("tax_model_targets")
                ),
            },
        }
    )


@app.route("/api/invoices/<int:invoice_id>", methods=["DELETE"])
def delete_invoice(invoice_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    with engine.begin() as conn:
        result = conn.execute(
            invoices_table.delete()
            .where(invoices_table.c.id == invoice_id)
            .where(invoices_table.c.user_id == data_owner_id)
            .where(invoices_table.c.company_id == company_id)
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Factura no encontrada."]}), 404

    return jsonify({"ok": True})


@app.route("/api/income-invoices")
def list_income_invoices():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    _, last_day = calendar.monthrange(year, month)
    start = date(year, month, 1).isoformat()
    end = date(year, month, last_day).isoformat()

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                income_invoices_table.c.id,
                income_invoices_table.c.invoice_date,
                income_invoices_table.c.payment_date,
                income_invoices_table.c.payment_dates,
                income_invoices_table.c.payment_completed_dates,
                income_invoices_table.c.client,
                income_invoices_table.c.base_amount,
                income_invoices_table.c.vat_rate,
                income_invoices_table.c.vat_amount,
                income_invoices_table.c.total_amount,
                income_invoices_table.c.vat_breakdown,
                income_invoices_table.c.extraction_source,
                income_invoices_table.c.confidence_score,
                income_invoices_table.c.original_filename,
            )
            .where(income_invoices_table.c.user_id == data_owner_id)
            .where(income_invoices_table.c.company_id == company_id)
            .where(income_invoices_table.c.invoice_date.between(start, end))
            .order_by(income_invoices_table.c.invoice_date.desc(), income_invoices_table.c.id.desc())
        ).mappings().all()

    invoices = [
        {
            "id": row["id"],
            "invoice_date": row["invoice_date"],
            "payment_date": row["payment_date"]
            or compute_payment_date(row["invoice_date"], row["payment_date"]),
            "payment_dates": parse_payment_dates(row.get("payment_dates"))
            or ([row["payment_date"]] if row.get("payment_date") else []),
            "payment_completed_dates": parse_payment_dates(row.get("payment_completed_dates")),
            "client": row["client"],
            "base_amount": float(row["base_amount"]),
            "vat_rate": int(row["vat_rate"]) if row["vat_rate"] is not None else None,
            "vat_amount": float(row["vat_amount"]) if row["vat_amount"] is not None else None,
            "total_amount": float(row["total_amount"]),
            "vat_breakdown": row.get("vat_breakdown"),
            "extraction_source": row.get("extraction_source"),
            "confidence_score": float(row["confidence_score"]) if row["confidence_score"] is not None else None,
            "original_filename": row["original_filename"],
        }
        for row in rows
    ]

    return jsonify({"invoices": invoices})


@app.route("/api/income-invoices", methods=["POST"])
def create_income_invoices():
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    entries = payload.get("entries", [])
    if not entries:
        return jsonify({"ok": False, "errors": ["No se recibieron entradas."]}), 400

    errors = []
    inserted = 0
    with engine.begin() as conn:
        for entry in entries:
            original_name = entry.get("originalFilename") or ""
            invoice_date = entry.get("date") or entry.get("invoice_date") or date.today().isoformat()
            client = (entry.get("client") or "").strip()
            base_amount = parse_amount(str(entry.get("base") or ""))
            vat_rate_raw = vat_rate_to_str(entry.get("vat"))
            vat_amount = parse_amount(str(entry.get("vatAmount") or ""))
            total_amount = parse_amount(str(entry.get("total") or ""))
            is_rectificativa = bool(entry.get("isRectificativa"))
            vat_breakdown = parse_vat_breakdown(
                entry.get("vatBreakdown") or entry.get("vat_breakdown")
            )
            vat_breakdown_json = json.dumps(vat_breakdown) if vat_breakdown else None
            payment_dates = parse_payment_dates(
                entry.get("paymentDates") or entry.get("payment_dates")
            )
            payment_date = compute_payment_date(
                invoice_date,
                entry.get("paymentDate")
                or entry.get("payment_date")
                or (payment_dates[0] if payment_dates else None),
            )
            extraction_source = (
                entry.get("extractionSource") or entry.get("extraction_source")
            )
            confidence_score = entry.get("confidenceScore") or entry.get("confidence_score")

            if not client:
                app.logger.info(
                    "Cliente vacío para %s. Se permite guardado manual.",
                    original_name,
                )
            if client and is_supplier_same_as_company(client, company_id, conn):
                errors.append(
                    f"El cliente no puede ser la empresa activa ({original_name})."
                )
                continue
            if vat_breakdown:
                vat_rate_int = infer_vat_rate_from_breakdown(vat_breakdown)
                summary = summarize_vat_breakdown(vat_breakdown)
                if summary:
                    base_amount, vat_amount, total_amount = summary
            else:
                try:
                    vat_rate_int = int(vat_rate_raw)
                except ValueError:
                    errors.append(f"Tipo de IVA inválido para {original_name}.")
                    continue
                if vat_rate_int not in {0, 4, 10, 21}:
                    errors.append(f"Tipo de IVA inválido para {original_name}.")
                    continue
            if base_amount is None and total_amount is None:
                errors.append(
                    f"Base imponible o total obligatorio para {original_name}."
                )
                continue
            if base_amount is not None and total_amount is not None:
                if (base_amount < 0 or total_amount < 0) and not is_rectificativa:
                    errors.append(f"Factura rectificativa no indicada para {original_name}.")
                    continue

            if vat_rate_int is not None and vat_rate_int >= 0:
                base_amount, vat_amount, total_amount = normalize_vat_amounts(
                    base_amount, vat_rate_int, vat_amount, total_amount
                )

            created_at = datetime.utcnow().isoformat()

            conn.execute(
                income_invoices_table.insert().values(
                    user_id=data_owner_id,
                    company_id=company_id,
                    original_filename=original_name,
                    stored_filename="",
                    invoice_date=invoice_date,
                    client=client,
                    base_amount=base_amount,
                    vat_rate=vat_rate_int,
                    vat_amount=vat_amount,
                    total_amount=total_amount,
                    vat_breakdown=vat_breakdown_json,
                    payment_date=payment_date,
                    payment_dates=json.dumps(payment_dates) if payment_dates else None,
                    ocr_text=None,
                    extraction_source=extraction_source,
                    confidence_score=confidence_score,
                    created_at=created_at,
                )
            )
            inserted += 1

    return jsonify({"ok": True, "inserted": inserted, "errors": errors})


@app.route("/api/income-invoices/<int:invoice_id>", methods=["PUT"])
def update_income_invoice(invoice_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    payment_only = payload.get("payment_only") or payload.get("paymentOnly")
    invoice_date = payload.get("invoice_date") or ""
    payment_dates_payload = payload.get("payment_dates") or payload.get("paymentDates")
    payment_dates = parse_payment_dates(payment_dates_payload)
    payment_date_input = (
        payload.get("payment_date")
        or payload.get("paymentDate")
        or (payment_dates[0] if payment_dates else None)
    )
    payment_date = compute_payment_date(invoice_date, payment_date_input)
    if payment_only:
        mark_paid = bool(payload.get("mark_paid") or payload.get("markPaid"))
        mark_unpaid = bool(payload.get("mark_unpaid") or payload.get("markUnpaid"))
        reference_date = normalize_date(
            payload.get("payment_reference_date") or payload.get("paymentReferenceDate")
        )
        if not payment_date_input and not mark_paid and not mark_unpaid:
            return jsonify({"ok": False, "errors": ["Fecha de pago obligatoria."]}), 400
        with engine.begin() as conn:
            current_row = conn.execute(
                select(
                    income_invoices_table.c.payment_date,
                    income_invoices_table.c.payment_dates,
                    income_invoices_table.c.payment_completed_dates,
                    income_invoices_table.c.invoice_date,
                )
                .where(income_invoices_table.c.id == invoice_id)
                .where(income_invoices_table.c.user_id == data_owner_id)
                .where(income_invoices_table.c.company_id == company_id)
            ).mappings().first()
            if not current_row:
                return jsonify({"ok": False, "errors": ["Factura no encontrada."]}), 404
            existing_dates = parse_payment_dates(current_row.get("payment_dates"))
            if not existing_dates:
                fallback = current_row.get("payment_date") or compute_payment_date(
                    current_row.get("invoice_date"), None
                )
                if fallback:
                    existing_dates = [fallback]
            updated_dates = payment_dates if payment_dates_payload is not None else existing_dates
            if payment_date_input and reference_date and reference_date != normalize_date(payment_date_input):
                updated_dates = replace_payment_date(updated_dates, reference_date, payment_date_input)
            elif payment_date_input and not updated_dates:
                updated_dates = [normalize_date(payment_date_input)]
            completed_dates = parse_payment_dates(current_row.get("payment_completed_dates"))
            if payment_date_input and reference_date and reference_date != normalize_date(payment_date_input):
                completed_dates = replace_payment_date(
                    completed_dates, reference_date, payment_date_input
                )
            if mark_paid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    None,
                    reference_date or normalize_date(payment_date_input),
                )
            if mark_unpaid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    reference_date or normalize_date(payment_date_input),
                    None,
                )
            updates = {
                "payment_date": updated_dates[0] if updated_dates else normalize_date(payment_date_input),
                "payment_dates": serialize_payment_dates(updated_dates),
                "payment_completed_dates": serialize_payment_dates(completed_dates),
            }
            result = conn.execute(
                income_invoices_table.update()
                .where(income_invoices_table.c.id == invoice_id)
                .where(income_invoices_table.c.user_id == data_owner_id)
                .where(income_invoices_table.c.company_id == company_id)
                .values(**updates)
            )
        return jsonify({"ok": True})
    client = (payload.get("client") or "").strip()
    base_amount = parse_amount(str(payload.get("base_amount") or ""))
    vat_rate_raw = vat_rate_to_str(payload.get("vat_rate"))
    vat_amount = parse_amount(str(payload.get("vat_amount") or ""))
    total_amount = parse_amount(str(payload.get("total_amount") or ""))
    vat_breakdown = parse_vat_breakdown(
        payload.get("vat_breakdown") or payload.get("vatBreakdown")
    )
    vat_breakdown_json = json.dumps(vat_breakdown) if vat_breakdown else None

    errors = []
    if not invoice_date:
        errors.append("Fecha obligatoria.")
    if not client:
        app.logger.info("Cliente vacío en actualización de factura emitida %s.", invoice_id)
    if client and company_id:
        with engine.connect() as conn:
            if is_supplier_same_as_company(client, company_id, conn):
                errors.append("El cliente no puede ser la empresa activa.")
    if base_amount is None and total_amount is None:
        errors.append("Base imponible o total obligatorio.")
    if base_amount is not None and total_amount is not None:
        if (base_amount < 0 or total_amount < 0) and not is_rectificativa:
            errors.append("Factura rectificativa no indicada.")
    if vat_breakdown:
        vat_rate = infer_vat_rate_from_breakdown(vat_breakdown)
        summary = summarize_vat_breakdown(vat_breakdown)
        if summary:
            base_amount, vat_amount, total_amount = summary
    else:
        try:
            vat_rate = int(vat_rate_raw)
        except ValueError:
            vat_rate = None
        if vat_rate not in {0, 4, 10, 21}:
            errors.append("Tipo de IVA inválido.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    if vat_rate is not None and vat_rate >= 0:
        base_amount, vat_amount, total_amount = normalize_vat_amounts(
            base_amount, vat_rate, vat_amount, total_amount
        )

    updates = {
        "invoice_date": invoice_date,
        "payment_date": payment_date,
        "client": client,
        "base_amount": base_amount,
        "vat_rate": vat_rate,
        "vat_amount": vat_amount,
        "total_amount": total_amount,
        "vat_breakdown": vat_breakdown_json,
    }
    if payment_dates_payload is not None:
        updates["payment_dates"] = serialize_payment_dates(payment_dates)
    if "payment_completed_dates" in payload or "paymentCompletedDates" in payload:
        updates["payment_completed_dates"] = serialize_payment_dates(
            payload.get("payment_completed_dates") or payload.get("paymentCompletedDates")
        )

    with engine.begin() as conn:
        result = conn.execute(
            income_invoices_table.update()
            .where(income_invoices_table.c.id == invoice_id)
            .where(income_invoices_table.c.user_id == data_owner_id)
            .where(income_invoices_table.c.company_id == company_id)
            .values(**updates)
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Factura no encontrada."]}), 404

    return jsonify({"ok": True})


@app.route("/api/income-invoices/<int:invoice_id>", methods=["DELETE"])
def delete_income_invoice(invoice_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    with engine.begin() as conn:
        result = conn.execute(
            income_invoices_table.delete()
            .where(income_invoices_table.c.id == invoice_id)
            .where(income_invoices_table.c.user_id == data_owner_id)
            .where(income_invoices_table.c.company_id == company_id)
        )
    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Factura no encontrada."]}), 404
    return jsonify({"ok": True})


@app.route("/api/expenses/no-invoice")
def list_no_invoice_expenses():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    _, last_day = calendar.monthrange(year, month)
    start = date(year, month, 1).isoformat()
    end = date(year, month, last_day).isoformat()

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                no_invoice_table.c.id,
                no_invoice_table.c.expense_date,
                no_invoice_table.c.payment_date,
                no_invoice_table.c.payment_dates,
                no_invoice_table.c.payment_completed_dates,
                no_invoice_table.c.concept,
                no_invoice_table.c.amount,
                no_invoice_table.c.interest_amount,
                no_invoice_table.c.vat_deductible,
                no_invoice_table.c.vat_rate,
                no_invoice_table.c.vat_amount,
                no_invoice_table.c.base_amount,
                no_invoice_table.c.withholding_amount,
                no_invoice_table.c.payroll_employee_name,
                no_invoice_table.c.payroll_period,
                no_invoice_table.c.payroll_net_amount,
                no_invoice_table.c.payroll_total_deductions_amount,
                no_invoice_table.c.payroll_employer_cost_amount,
                no_invoice_table.c.expense_type,
                no_invoice_table.c.deductible,
                no_invoice_table.c.expense_family,
                no_invoice_table.c.expense_subtype,
                no_invoice_table.c.pnl_bucket,
                no_invoice_table.c.tax_model_targets,
            )
            .where(no_invoice_table.c.user_id == data_owner_id)
            .where(
                (no_invoice_table.c.company_id == company_id)
                | (no_invoice_table.c.company_id.is_(None))
            )
            .where(no_invoice_table.c.expense_date.between(start, end))
            .order_by(no_invoice_table.c.expense_date.desc(), no_invoice_table.c.id.desc())
        ).mappings().all()

    expenses = [
        {
            "id": row["id"],
            "expense_date": row["expense_date"],
            "payment_date": row["payment_date"] or row["expense_date"],
            "payment_dates": parse_payment_dates(row["payment_dates"])
            or ([row["payment_date"]] if row["payment_date"] else [row["expense_date"]]),
            "payment_completed_dates": parse_payment_dates(row.get("payment_completed_dates")),
            "concept": row["concept"],
            "amount": float(row["amount"]),
            "interest_amount": float(row["interest_amount"] or 0),
            "vat_deductible": bool(row["vat_deductible"]) if row.get("vat_deductible") is not None else False,
            "vat_rate": int(row["vat_rate"]) if row.get("vat_rate") is not None else None,
            "vat_amount": float(row["vat_amount"] or 0),
            "base_amount": float(row["base_amount"] or row["amount"] or 0),
            "withholding_amount": float(row["withholding_amount"] or 0),
            "payroll_employee_name": row.get("payroll_employee_name"),
            "payroll_period": row.get("payroll_period"),
            "payroll_net_amount": float(row["payroll_net_amount"] or 0),
            "payroll_total_deductions_amount": float(row["payroll_total_deductions_amount"] or 0),
            "payroll_employer_cost_amount": float(row["payroll_employer_cost_amount"] or 0),
            "expense_type": row["expense_type"],
            "deductible": bool(row["deductible"]),
            "expense_family": row.get("expense_family"),
            "expense_subtype": row.get("expense_subtype"),
            "pnl_bucket": row.get("pnl_bucket"),
            "tax_model_targets": parse_tax_model_targets(row.get("tax_model_targets")),
        }
        for row in rows
    ]

    return jsonify({"expenses": expenses})


@app.route("/api/expenses/no-invoice", methods=["POST"])
def create_no_invoice_expense():
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    payload = request.get_json(silent=True) or {}

    expense_date = payload.get("expense_date") or ""
    payment_dates = parse_payment_dates(payload.get("payment_dates") or payload.get("paymentDates"))
    concept = (payload.get("concept") or "").strip()
    amount = parse_amount(str(payload.get("amount") or ""))
    expense_type = payload.get("expense_type") or ""
    interest_amount = parse_amount(str(payload.get("interest_amount") or ""))
    vat_deductible = payload.get("vat_deductible")
    vat_rate_raw = payload.get("vat_rate")
    vat_amount_payload = parse_amount(str(payload.get("vat_amount") or ""))
    base_amount_payload = parse_amount(str(payload.get("base_amount") or ""))
    withholding_amount = parse_amount(str(payload.get("withholding_amount") or ""))
    payroll_employee_name = (payload.get("payroll_employee_name") or payload.get("payrollEmployeeName") or "").strip()
    payroll_period = (payload.get("payroll_period") or payload.get("payrollPeriod") or "").strip()
    payroll_net_amount = parse_amount(
        str(payload.get("payroll_net_amount") or payload.get("payrollNetAmount") or "")
    )
    payroll_total_deductions_amount = parse_amount(
        str(
            payload.get("payroll_total_deductions_amount")
            or payload.get("payrollTotalDeductionsAmount")
            or ""
        )
    )
    payroll_employer_cost_amount = parse_amount(
        str(
            payload.get("payroll_employer_cost_amount")
            or payload.get("payrollEmployerCostAmount")
            or ""
        )
    )
    deductible = payload.get("deductible")
    payment_date = compute_payment_date(
        expense_date,
        payload.get("payment_date")
        or payload.get("paymentDate")
        or (payment_dates[0] if payment_dates else expense_date),
    )

    errors = []
    if not expense_date:
        errors.append("Fecha obligatoria.")
    if not concept:
        errors.append("Concepto obligatorio.")
    if amount is None or amount < 0:
        errors.append("Importe inválido.")
    if expense_type not in ALLOWED_NO_INVOICE_EXPENSE_TYPES:
        errors.append("Tipo de gasto inválido.")
    if deductible is None:
        deductible = True
    if withholding_amount is None:
        withholding_amount = 0.0

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    if expense_type == "prestamo":
        if interest_amount is None:
            interest_amount = 0.0
        if interest_amount < 0:
            errors.append("Interés inválido.")
        if amount is not None and interest_amount > amount:
            errors.append("El interés no puede superar el importe.")
        withholding_amount = 0.0
        deductible = False
    else:
        interest_amount = None

    if expense_type in {"nomina", "seguridad_social"}:
        deductible = True
        vat_deductible = False
        vat_rate = None
        vat_amount = None
    if expense_type == "nomina":
        if payroll_total_deductions_amount is None and amount is not None and payroll_net_amount is not None:
            payroll_total_deductions_amount = round(max(amount - payroll_net_amount, 0), 2)
        if payroll_net_amount is None and amount is not None and payroll_total_deductions_amount is not None:
            payroll_net_amount = round(max(amount - payroll_total_deductions_amount, 0), 2)
        if payroll_net_amount is not None and amount is not None and payroll_net_amount > amount:
            errors.append("El líquido no puede superar el bruto de la nómina.")
        if (
            payroll_total_deductions_amount is not None
            and amount is not None
            and payroll_total_deductions_amount > amount
        ):
            errors.append("Las deducciones no pueden superar el bruto de la nómina.")
        if payroll_employer_cost_amount is not None and payroll_employer_cost_amount < 0:
            errors.append("El coste empresa es inválido.")
        if not payroll_employee_name and concept:
            payroll_employee_name = concept.replace("Nomina", "").replace("Nómina", "").strip(" -")

    if withholding_amount < 0:
        errors.append("Retención inválida.")
    elif amount is not None and withholding_amount > amount:
        errors.append("La retención no puede superar el importe.")

    if vat_deductible in (True, "true", "True", 1, "1"):
        vat_deductible = True
    else:
        vat_deductible = False

    vat_rate = None
    vat_amount = None
    base_amount = base_amount_payload
    if vat_deductible and expense_type != "prestamo":
        try:
            vat_rate = int(vat_rate_raw)
        except (TypeError, ValueError):
            vat_rate = None
        if vat_rate not in {0, 4, 10, 21}:
            errors.append("Tipo de IVA inválido.")
        else:
            if amount is None:
                errors.append("Importe inválido.")
            else:
                base_amount = round(amount / (1 + vat_rate / 100), 2)
                vat_amount = round(amount - base_amount, 2)
        deductible = True
    else:
        vat_deductible = False
        vat_rate = None
        vat_amount = None
        if base_amount is None:
            base_amount = amount

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    expense_profile = derive_no_invoice_profile(
        expense_type,
        vat_deductible,
        withholding_amount,
    )

    with engine.begin() as conn:
        conn.execute(
            no_invoice_table.insert().values(
                user_id=data_owner_id,
                company_id=company_id,
                expense_date=expense_date,
                payment_date=payment_date,
                payment_dates=json.dumps(payment_dates) if payment_dates else json.dumps([payment_date]),
                concept=concept,
                amount=amount,
                interest_amount=interest_amount,
                vat_deductible=vat_deductible,
                vat_rate=vat_rate,
                vat_amount=vat_amount,
                base_amount=base_amount,
                withholding_amount=withholding_amount,
                payroll_employee_name=payroll_employee_name or None,
                payroll_period=payroll_period or None,
                payroll_net_amount=payroll_net_amount,
                payroll_total_deductions_amount=payroll_total_deductions_amount,
                payroll_employer_cost_amount=payroll_employer_cost_amount,
                expense_type=expense_type,
                deductible=bool(deductible),
                **expense_profile,
                created_at=datetime.utcnow().isoformat(),
            )
        )

    return jsonify({"ok": True})


@app.route("/api/expenses/no-invoice/<int:expense_id>", methods=["PUT"])
def update_no_invoice_expense(expense_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    payload = request.get_json(silent=True) or {}

    payment_only = payload.get("payment_only") or payload.get("paymentOnly")
    expense_date = payload.get("expense_date") or ""
    if payment_only:
        mark_paid = bool(payload.get("mark_paid") or payload.get("markPaid"))
        mark_unpaid = bool(payload.get("mark_unpaid") or payload.get("markUnpaid"))
        reference_date = normalize_date(
            payload.get("payment_reference_date") or payload.get("paymentReferenceDate")
        )
        if not expense_date and not mark_paid and not mark_unpaid:
            return jsonify({"ok": False, "errors": ["Fecha obligatoria."]}), 400
        payment_dates = parse_payment_dates(payload.get("payment_dates") or payload.get("paymentDates"))
        payment_date = compute_payment_date(
            expense_date,
            payload.get("payment_date")
            or payload.get("paymentDate")
            or (payment_dates[0] if payment_dates else expense_date),
        )
        with engine.begin() as conn:
            current_row = conn.execute(
                select(
                    no_invoice_table.c.expense_date,
                    no_invoice_table.c.payment_date,
                    no_invoice_table.c.payment_dates,
                    no_invoice_table.c.payment_completed_dates,
                )
                .where(no_invoice_table.c.id == expense_id)
                .where(no_invoice_table.c.user_id == data_owner_id)
                .where(no_invoice_table.c.company_id == company_id)
            ).mappings().first()
            if not current_row:
                return jsonify({"ok": False, "errors": ["Gasto no encontrado."]}), 404
            effective_expense_date = expense_date or current_row.get("expense_date")
            existing_dates = parse_payment_dates(current_row.get("payment_dates"))
            if not existing_dates:
                fallback = current_row.get("payment_date") or effective_expense_date
                if fallback:
                    existing_dates = [fallback]
            updated_dates = payment_dates if payment_dates else existing_dates
            if payment_date and reference_date and reference_date != normalize_date(payment_date):
                updated_dates = replace_payment_date(updated_dates, reference_date, payment_date)
            elif payment_date and not updated_dates:
                updated_dates = [normalize_date(payment_date)]
            completed_dates = parse_payment_dates(current_row.get("payment_completed_dates"))
            if payment_date and reference_date and reference_date != normalize_date(payment_date):
                completed_dates = replace_payment_date(
                    completed_dates, reference_date, payment_date
                )
            if mark_paid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    None,
                    reference_date or normalize_date(payment_date),
                )
            if mark_unpaid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    reference_date or normalize_date(payment_date),
                    None,
                )
            result = conn.execute(
                no_invoice_table.update()
                .where(no_invoice_table.c.id == expense_id)
                .where(no_invoice_table.c.user_id == data_owner_id)
                .where(no_invoice_table.c.company_id == company_id)
                .values(
                    expense_date=effective_expense_date,
                    payment_date=updated_dates[0] if updated_dates else normalize_date(payment_date),
                    payment_dates=serialize_payment_dates(updated_dates),
                    payment_completed_dates=serialize_payment_dates(completed_dates),
                )
            )
        return jsonify({"ok": True})
    concept = (payload.get("concept") or "").strip()
    amount = parse_amount(str(payload.get("amount") or ""))
    expense_type = payload.get("expense_type") or ""
    interest_amount = parse_amount(str(payload.get("interest_amount") or ""))
    vat_deductible = payload.get("vat_deductible")
    vat_rate_raw = payload.get("vat_rate")
    vat_amount_payload = parse_amount(str(payload.get("vat_amount") or ""))
    base_amount_payload = parse_amount(str(payload.get("base_amount") or ""))
    withholding_amount = parse_amount(str(payload.get("withholding_amount") or ""))
    payroll_employee_name = (payload.get("payroll_employee_name") or payload.get("payrollEmployeeName") or "").strip()
    payroll_period = (payload.get("payroll_period") or payload.get("payrollPeriod") or "").strip()
    payroll_net_amount = parse_amount(
        str(payload.get("payroll_net_amount") or payload.get("payrollNetAmount") or "")
    )
    payroll_total_deductions_amount = parse_amount(
        str(
            payload.get("payroll_total_deductions_amount")
            or payload.get("payrollTotalDeductionsAmount")
            or ""
        )
    )
    payroll_employer_cost_amount = parse_amount(
        str(
            payload.get("payroll_employer_cost_amount")
            or payload.get("payrollEmployerCostAmount")
            or ""
        )
    )
    deductible = payload.get("deductible")
    payment_dates = parse_payment_dates(payload.get("payment_dates") or payload.get("paymentDates"))
    payment_date = compute_payment_date(
        expense_date,
        payload.get("payment_date")
        or payload.get("paymentDate")
        or (payment_dates[0] if payment_dates else expense_date),
    )

    errors = []
    if not expense_date:
        errors.append("Fecha obligatoria.")
    if not concept:
        errors.append("Concepto obligatorio.")
    if amount is None or amount < 0:
        errors.append("Importe inválido.")
    if expense_type not in ALLOWED_NO_INVOICE_EXPENSE_TYPES:
        errors.append("Tipo de gasto inválido.")
    if deductible is None:
        deductible = True
    if withholding_amount is None:
        withholding_amount = 0.0

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    if expense_type == "prestamo":
        if interest_amount is None:
            interest_amount = 0.0
        if interest_amount < 0:
            errors.append("Interés inválido.")
        if amount is not None and interest_amount > amount:
            errors.append("El interés no puede superar el importe.")
        withholding_amount = 0.0
        deductible = False
    else:
        interest_amount = None

    if expense_type in {"nomina", "seguridad_social"}:
        deductible = True
        vat_deductible = False
        vat_rate = None
        vat_amount = None
    if expense_type == "nomina":
        if payroll_total_deductions_amount is None and amount is not None and payroll_net_amount is not None:
            payroll_total_deductions_amount = round(max(amount - payroll_net_amount, 0), 2)
        if payroll_net_amount is None and amount is not None and payroll_total_deductions_amount is not None:
            payroll_net_amount = round(max(amount - payroll_total_deductions_amount, 0), 2)
        if payroll_net_amount is not None and amount is not None and payroll_net_amount > amount:
            errors.append("El líquido no puede superar el bruto de la nómina.")
        if (
            payroll_total_deductions_amount is not None
            and amount is not None
            and payroll_total_deductions_amount > amount
        ):
            errors.append("Las deducciones no pueden superar el bruto de la nómina.")
        if payroll_employer_cost_amount is not None and payroll_employer_cost_amount < 0:
            errors.append("El coste empresa es inválido.")
        if not payroll_employee_name and concept:
            payroll_employee_name = concept.replace("Nomina", "").replace("Nómina", "").strip(" -")

    if withholding_amount < 0:
        errors.append("Retención inválida.")
    elif amount is not None and withholding_amount > amount:
        errors.append("La retención no puede superar el importe.")

    if vat_deductible in (True, "true", "True", 1, "1"):
        vat_deductible = True
    else:
        vat_deductible = False

    vat_rate = None
    vat_amount = None
    base_amount = base_amount_payload
    if vat_deductible and expense_type != "prestamo":
        try:
            vat_rate = int(vat_rate_raw)
        except (TypeError, ValueError):
            vat_rate = None
        if vat_rate not in {0, 4, 10, 21}:
            errors.append("Tipo de IVA inválido.")
        else:
            if amount is None:
                errors.append("Importe inválido.")
            else:
                base_amount = round(amount / (1 + vat_rate / 100), 2)
                vat_amount = round(amount - base_amount, 2)
        deductible = True
    else:
        vat_deductible = False
        vat_rate = None
        vat_amount = None
        if base_amount is None:
            base_amount = amount

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    expense_profile = derive_no_invoice_profile(
        expense_type,
        vat_deductible,
        withholding_amount,
    )

    with engine.begin() as conn:
        values = {
            "expense_date": expense_date,
            "payment_date": payment_date,
            "payment_dates": serialize_payment_dates(payment_dates if payment_dates else [payment_date]),
            "concept": concept,
            "amount": amount,
            "interest_amount": interest_amount,
            "vat_deductible": vat_deductible,
            "vat_rate": vat_rate,
            "vat_amount": vat_amount,
            "base_amount": base_amount,
            "withholding_amount": withholding_amount,
            "payroll_employee_name": payroll_employee_name or None,
            "payroll_period": payroll_period or None,
            "payroll_net_amount": payroll_net_amount,
            "payroll_total_deductions_amount": payroll_total_deductions_amount,
            "payroll_employer_cost_amount": payroll_employer_cost_amount,
            "expense_type": expense_type,
            "deductible": bool(deductible),
            **expense_profile,
        }
        if "payment_completed_dates" in payload or "paymentCompletedDates" in payload:
            values["payment_completed_dates"] = serialize_payment_dates(
                payload.get("payment_completed_dates") or payload.get("paymentCompletedDates")
            )
        result = conn.execute(
            no_invoice_table.update()
            .where(no_invoice_table.c.id == expense_id)
            .where(no_invoice_table.c.user_id == data_owner_id)
            .where(no_invoice_table.c.company_id == company_id)
            .values(**values)
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Gasto no encontrado."]}), 404

    return jsonify(
        {
            "ok": True,
            "expense": {
                "id": expense_id,
                "expense_date": expense_date,
                "payment_date": payment_date,
                "payment_dates": payment_dates if payment_dates else [payment_date],
                "concept": concept,
                "amount": amount,
                "interest_amount": float(interest_amount or 0),
                "vat_deductible": vat_deductible,
                "vat_rate": vat_rate,
                "vat_amount": float(vat_amount or 0),
                "base_amount": float(base_amount or amount or 0),
                "expense_type": expense_type,
                "deductible": bool(deductible),
                "payroll_employee_name": payroll_employee_name or None,
                "payroll_period": payroll_period or None,
                "payroll_net_amount": float(payroll_net_amount or 0),
                "payroll_total_deductions_amount": float(
                    payroll_total_deductions_amount or 0
                ),
                "payroll_employer_cost_amount": float(
                    payroll_employer_cost_amount or 0
                ),
                **{
                    "expense_family": expense_profile.get("expense_family"),
                    "expense_subtype": expense_profile.get("expense_subtype"),
                    "pnl_bucket": expense_profile.get("pnl_bucket"),
                    "tax_model_targets": parse_tax_model_targets(
                        expense_profile.get("tax_model_targets")
                    ),
                },
            },
        }
    )


@app.route("/api/expenses/no-invoice/<int:expense_id>", methods=["DELETE"])
def delete_no_invoice_expense(expense_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    with engine.begin() as conn:
        result = conn.execute(
            no_invoice_table.delete()
            .where(no_invoice_table.c.id == expense_id)
            .where(no_invoice_table.c.user_id == data_owner_id)
            .where(no_invoice_table.c.company_id == company_id)
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Gasto no encontrado."]}), 404

    return jsonify({"ok": True})


@app.route("/api/loan-installments")
def list_loan_installments():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year
    _, last_day = calendar.monthrange(year, month)
    start = date(year, month, 1).isoformat()
    end = date(year, month, last_day).isoformat()

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                loan_installments_table.c.id,
                loan_installments_table.c.payment_date,
                loan_installments_table.c.payment_completed_dates,
                loan_installments_table.c.bank_name,
                loan_installments_table.c.concept,
                loan_installments_table.c.total_amount,
                loan_installments_table.c.interest_amount,
                loan_installments_table.c.principal_amount,
            )
            .where(loan_installments_table.c.user_id == data_owner_id)
            .where(loan_installments_table.c.company_id == company_id)
            .where(loan_installments_table.c.payment_date.between(start, end))
            .order_by(
                loan_installments_table.c.payment_date.desc(),
                loan_installments_table.c.id.desc(),
            )
        ).mappings().all()

    installments = [
        {
            "id": row["id"],
            "payment_date": row["payment_date"],
            "payment_completed_dates": parse_payment_dates(row.get("payment_completed_dates")),
            "bank_name": row.get("bank_name"),
            "concept": row["concept"],
            "total_amount": float(row["total_amount"] or 0),
            "interest_amount": float(row["interest_amount"] or 0),
            "principal_amount": float(row["principal_amount"] or 0),
        }
        for row in rows
    ]
    return jsonify({"installments": installments})


@app.route("/api/loan-installments", methods=["POST"])
def create_loan_installment():
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    payment_date = payload.get("payment_date")
    concept = (payload.get("concept") or "").strip()
    bank_name = (payload.get("bank_name") or "").strip()
    total_amount = payload.get("total_amount")
    interest_amount = payload.get("interest_amount")

    errors = []
    if not payment_date:
        errors.append("Fecha de pago obligatoria.")
    if not concept:
        errors.append("Concepto obligatorio.")
    try:
        total_amount = float(total_amount)
    except (TypeError, ValueError):
        total_amount = None
    try:
        interest_amount = float(interest_amount)
    except (TypeError, ValueError):
        interest_amount = None

    if total_amount is None or total_amount < 0:
        errors.append("Importe total inválido.")
    if interest_amount is None or interest_amount < 0:
        errors.append("Interés inválido.")
    if total_amount is not None and interest_amount is not None:
        if interest_amount > total_amount:
            errors.append("El interés no puede superar el importe total.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    principal_amount = round(total_amount - interest_amount, 2)
    created_at = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        result = conn.execute(
            loan_installments_table.insert().values(
                user_id=data_owner_id,
                company_id=company_id,
                bank_name=bank_name or None,
                concept=concept,
                payment_date=payment_date,
                payment_completed_dates=None,
                total_amount=total_amount,
                interest_amount=interest_amount,
                principal_amount=principal_amount,
                created_at=created_at,
            )
        )
        new_id = result.inserted_primary_key[0] if result.inserted_primary_key else None

    return jsonify({"ok": True, "id": new_id})


@app.route("/api/loan-installments/<int:installment_id>", methods=["PUT"])
def update_loan_installment(installment_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    payment_only = payload.get("payment_only") or payload.get("paymentOnly")
    payment_date = payload.get("payment_date")
    concept = (payload.get("concept") or "").strip()
    total_amount = payload.get("total_amount")
    interest_amount = payload.get("interest_amount")
    bank_name = (payload.get("bank_name") or payload.get("bankName") or "").strip()

    if payment_only:
        mark_paid = bool(payload.get("mark_paid") or payload.get("markPaid"))
        mark_unpaid = bool(payload.get("mark_unpaid") or payload.get("markUnpaid"))
        reference_date = normalize_date(
            payload.get("payment_reference_date") or payload.get("paymentReferenceDate")
        )
        if not payment_date and not mark_paid and not mark_unpaid:
            return jsonify({"ok": False, "errors": ["Fecha de pago obligatoria."]}), 400
        with engine.begin() as conn:
            current_row = conn.execute(
                select(
                    loan_installments_table.c.payment_date,
                    loan_installments_table.c.payment_completed_dates,
                )
                .where(loan_installments_table.c.id == installment_id)
                .where(loan_installments_table.c.user_id == data_owner_id)
                .where(loan_installments_table.c.company_id == company_id)
            ).mappings().first()
            if not current_row:
                return jsonify({"ok": False, "errors": ["Cuota no encontrada."]}), 404
            completed_dates = parse_payment_dates(current_row.get("payment_completed_dates"))
            if payment_date and reference_date and reference_date != normalize_date(payment_date):
                completed_dates = replace_payment_date(
                    completed_dates, reference_date, payment_date
                )
            if mark_paid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    None,
                    reference_date or normalize_date(payment_date) or current_row.get("payment_date"),
                )
            if mark_unpaid:
                completed_dates = replace_payment_date(
                    completed_dates,
                    reference_date or normalize_date(payment_date) or current_row.get("payment_date"),
                    None,
                )
            result = conn.execute(
                loan_installments_table.update()
                .where(loan_installments_table.c.id == installment_id)
                .where(loan_installments_table.c.user_id == data_owner_id)
                .where(loan_installments_table.c.company_id == company_id)
                .values(
                    payment_date=normalize_date(payment_date) or current_row.get("payment_date"),
                    payment_completed_dates=serialize_payment_dates(completed_dates),
                )
            )
        return jsonify({"ok": True})

    errors = []
    if not payment_date:
        errors.append("Fecha de pago obligatoria.")
    if not concept:
        errors.append("Concepto obligatorio.")
    try:
        total_amount = float(total_amount)
    except (TypeError, ValueError):
        total_amount = None
    try:
        interest_amount = float(interest_amount)
    except (TypeError, ValueError):
        interest_amount = None

    if total_amount is None or total_amount < 0:
        errors.append("Importe total inválido.")
    if interest_amount is None or interest_amount < 0:
        errors.append("Interés inválido.")
    if total_amount is not None and interest_amount is not None:
        if interest_amount > total_amount:
            errors.append("El interés no puede superar el importe total.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    principal_amount = round(total_amount - interest_amount, 2)
    with engine.begin() as conn:
        values = {
            "payment_date": payment_date,
            "bank_name": bank_name or None,
            "concept": concept,
            "total_amount": total_amount,
            "interest_amount": interest_amount,
            "principal_amount": principal_amount,
        }
        if "payment_completed_dates" in payload or "paymentCompletedDates" in payload:
            values["payment_completed_dates"] = serialize_payment_dates(
                payload.get("payment_completed_dates") or payload.get("paymentCompletedDates")
            )
        result = conn.execute(
            loan_installments_table.update()
            .where(loan_installments_table.c.id == installment_id)
            .where(loan_installments_table.c.user_id == data_owner_id)
            .where(loan_installments_table.c.company_id == company_id)
            .values(**values)
        )
        if result.rowcount == 0:
            return jsonify({"ok": False, "errors": ["Cuota no encontrada."]}), 404

    return jsonify({"ok": True})


@app.route("/api/loan-installments/<int:installment_id>", methods=["DELETE"])
def delete_loan_installment(installment_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    with engine.begin() as conn:
        result = conn.execute(
            loan_installments_table.delete()
            .where(loan_installments_table.c.id == installment_id)
            .where(loan_installments_table.c.user_id == data_owner_id)
            .where(loan_installments_table.c.company_id == company_id)
        )
        if result.rowcount == 0:
            return jsonify({"ok": False, "errors": ["Cuota no encontrada."]}), 404

    return jsonify({"ok": True})


@app.route("/api/loan-installments/import", methods=["POST"])
def import_loan_installments():
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    uploaded_file = request.files.get("file")
    if not uploaded_file:
        return jsonify({"ok": False, "errors": ["Archivo no recibido."]}), 400

    filename = secure_filename(uploaded_file.filename or "")
    if not filename:
        return jsonify({"ok": False, "errors": ["Nombre de archivo inválido."]}), 400

    concept = (request.form.get("concept") or "Préstamo bancario").strip()
    file_bytes = uploaded_file.read()
    extension = os.path.splitext(filename)[1].lower()
    installments = []

    try:
        if extension in {".xlsx", ".xls"}:
            installments = parse_loan_installments_from_excel(file_bytes)
        elif extension == ".pdf":
            text = _extract_pdf_text_from_bytes(file_bytes)
            if not text or len(text.strip()) < 50:
                text = _extract_pdf_text_ocr_from_bytes(file_bytes)
            installments = parse_loan_installments_from_text(text)
            if not installments and text and len(text.strip()) >= 50:
                installments = extract_loan_schedule(text)
        elif extension in {".jpg", ".jpeg", ".png"}:
            text = _extract_image_text_ocr_from_bytes(file_bytes)
            installments = parse_loan_installments_from_text(text)
            if not installments and text and len(text.strip()) >= 50:
                installments = extract_loan_schedule(text)
        else:
            return jsonify({"ok": False, "errors": ["Formato no soportado."]}), 400
    except Exception:
        return jsonify({"ok": False, "errors": ["No se pudo leer el archivo."]}), 400

    if not installments:
        return jsonify({"ok": False, "errors": ["No se detectaron cuotas válidas."]}), 400

    if request.args.get("preview") or request.form.get("preview"):
        preview_items = [
            {
                "payment_date": item["payment_date"],
                "bank_name": item.get("bank_name"),
                "concept": concept,
                "total_amount": item["total_amount"],
                "interest_amount": item["interest_amount"],
                "principal_amount": item["principal_amount"],
            }
            for item in installments
        ]
        return jsonify({"ok": True, "installments": preview_items})

    created_at = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        for item in installments:
            conn.execute(
                loan_installments_table.insert().values(
                    user_id=data_owner_id,
                    company_id=company_id,
                    bank_name=item.get("bank_name"),
                    concept=concept,
                    payment_date=item["payment_date"],
                    total_amount=item["total_amount"],
                    interest_amount=item["interest_amount"],
                    principal_amount=item["principal_amount"],
                    created_at=created_at,
                )
            )

    return jsonify({"ok": True, "count": len(installments)})


@app.route("/api/loan-installments/batch", methods=["POST"])
def create_loan_installments_batch():
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    payload = request.get_json(silent=True) or {}
    installments = payload.get("installments")
    if not isinstance(installments, list) or not installments:
        return jsonify({"ok": False, "errors": ["No hay cuotas para guardar."]}), 400

    created_at = datetime.utcnow().isoformat()
    errors = []
    clean_items = []
    for item in installments:
        if not isinstance(item, dict):
            continue
        payment_date = (item.get("payment_date") or "").strip()
        bank_name = (item.get("bank_name") or "").strip()
        if not payment_date:
            errors.append("Fecha de pago inválida.")
            continue
        try:
            total_amount = float(item.get("total_amount") or 0)
        except (TypeError, ValueError):
            errors.append("Importe total inválido.")
            continue
        try:
            interest_amount = float(item.get("interest_amount") or 0)
        except (TypeError, ValueError):
            interest_amount = 0.0
        try:
            principal_amount = float(item.get("principal_amount") or 0)
        except (TypeError, ValueError):
            principal_amount = 0.0
        concept = (item.get("concept") or "Préstamo bancario").strip()
        if total_amount < 0 or interest_amount < 0 or principal_amount < 0:
            errors.append("Importes inválidos.")
            continue
        if principal_amount == 0 and total_amount and interest_amount:
            principal_amount = max(total_amount - interest_amount, 0)
        clean_items.append(
            {
                "payment_date": payment_date,
                "concept": concept,
                "bank_name": bank_name or None,
                "total_amount": round(total_amount, 2),
                "interest_amount": round(interest_amount, 2),
                "principal_amount": round(principal_amount, 2),
            }
        )

    if errors and not clean_items:
        return jsonify({"ok": False, "errors": errors}), 400

    with engine.begin() as conn:
        for item in clean_items:
            conn.execute(
                loan_installments_table.insert().values(
                    user_id=data_owner_id,
                    company_id=company_id,
                    bank_name=item.get("bank_name"),
                    concept=item["concept"],
                    payment_date=item["payment_date"],
                    total_amount=item["total_amount"],
                    interest_amount=item["interest_amount"],
                    principal_amount=item["principal_amount"],
                    created_at=created_at,
                )
            )

    return jsonify({"ok": True, "count": len(clean_items)})


@app.route("/api/billing/entries")
def billing_entries():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    with engine.connect() as conn:
        rows = conn.execute(
            select(
                facturacion_table.c.id,
                facturacion_table.c.mes,
                facturacion_table.c.anio,
                facturacion_table.c.invoice_date,
                facturacion_table.c.concept,
                facturacion_table.c.base_facturada,
                facturacion_table.c.tipo_iva,
                facturacion_table.c.iva_repercutido,
                facturacion_table.c.total_amount,
            )
            .where(
                facturacion_table.c.mes == month,
                facturacion_table.c.anio == year,
                facturacion_table.c.user_id == data_owner_id,
                facturacion_table.c.company_id == company_id,
            )
            .order_by(facturacion_table.c.id.desc())
        ).mappings().all()

    entries = [
        {
            "id": row["id"],
            "month": row["mes"],
            "year": row["anio"],
            "invoice_date": row["invoice_date"],
            "concept": row["concept"],
            "base": float(row["base_facturada"]),
            "vat": int(row["tipo_iva"]),
            "vatAmount": float(row["iva_repercutido"]),
            "total": float(row["total_amount"] or 0),
        }
        for row in rows
    ]

    return jsonify({"entries": entries})


@app.route("/api/billing/<int:billing_id>", methods=["PUT"])
def update_billing(billing_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    payload = request.get_json(silent=True) or request.form

    base_amount = parse_amount(str(payload.get("base") or ""))
    vat_rate_raw = vat_rate_to_str(payload.get("vat"))

    errors = []
    if base_amount is None or base_amount < 0:
        errors.append("Base facturada inválida.")
    try:
        vat_rate = int(vat_rate_raw)
    except ValueError:
        vat_rate = None
    if vat_rate not in {0, 4, 10, 21}:
        errors.append("Tipo de IVA inválido.")

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    iva_repercutido = round(base_amount * (vat_rate / 100), 2)
    total_amount = round(base_amount + iva_repercutido, 2)

    with engine.begin() as conn:
        result = conn.execute(
            facturacion_table.update()
            .where(facturacion_table.c.id == billing_id)
            .where(facturacion_table.c.user_id == data_owner_id)
            .where(facturacion_table.c.company_id == company_id)
            .values(
                base_facturada=base_amount,
                tipo_iva=vat_rate,
                iva_repercutido=iva_repercutido,
                total_amount=total_amount,
            )
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Registro no encontrado."]}), 404

    return jsonify(
        {
            "ok": True,
            "entry": {
                "id": billing_id,
                "base": base_amount,
                "vat": vat_rate,
                "vatAmount": iva_repercutido,
                "total": total_amount,
            },
        }
    )


@app.route("/api/billing/<int:billing_id>", methods=["DELETE"])
def delete_billing(billing_id):
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    with engine.begin() as conn:
        result = conn.execute(
            facturacion_table.delete()
            .where(facturacion_table.c.id == billing_id)
            .where(facturacion_table.c.user_id == data_owner_id)
            .where(facturacion_table.c.company_id == company_id)
        )

    if result.rowcount == 0:
        return jsonify({"ok": False, "errors": ["Registro no encontrado."]}), 404

    return jsonify({"ok": True})


@app.route("/api/accounting-integrations/summary")
def accounting_integrations_summary():
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    try:
        start_date, end_date, period_label = build_accounting_export_range()
    except ValueError as exc:
        return jsonify({"ok": False, "errors": [str(exc)]}), 400

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.id,
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.tax_id,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(company_query).mappings().first()
        if not company:
            return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
        source_data = load_accounting_export_source_data(
            conn,
            data_owner_id,
            company_id,
            start_date,
            end_date,
        )

    purchase_rows = build_purchase_export_rows(source_data)
    sales_rows = build_sales_export_rows(source_data)
    journal_rows = build_journal_export_rows(source_data)
    document_rows = build_document_manifest_rows(source_data)

    return jsonify(
        {
            "ok": True,
            "company": {
                "id": company["id"],
                "displayName": company.get("display_name"),
                "legalName": company.get("legal_name"),
                "taxId": company.get("tax_id"),
            },
            "periodLabel": period_label,
            "startDate": start_date.isoformat(),
            "endDate": end_date.isoformat(),
            "stats": {
                "purchaseDocuments": len(purchase_rows),
                "salesDocuments": len(sales_rows),
                "journalLines": len(journal_rows),
                "journalEntries": len({row["asiento_id"] for row in journal_rows}),
                "documentFiles": len(document_rows),
            },
            "targets": [
                {
                    "id": "generic",
                    "label": "Paquete contable Ledged",
                    "status": "ready",
                    "description": "Exportación normalizada PGC con compras, ventas, asientos y trazabilidad documental.",
                },
                {
                    "id": "contasol",
                    "label": "ContaSOL / importación asistida",
                    "status": "ready",
                    "description": "Usa CSV o XLSX como capa puente para importar y revisar asientos desde la gestoría.",
                },
                {
                    "id": "odoo_a3_cegid",
                    "label": "Odoo / A3 / Cegid",
                    "status": "bridge",
                    "description": "Primera fase por exportación puente. La conexión API directa queda preparada como siguiente capa.",
                },
            ],
        }
    )


@app.route("/api/accounting-integrations/export/<export_kind>")
def accounting_integrations_export(export_kind):
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    normalized_kind = (export_kind or "").strip().lower()
    export_format = (request.args.get("format") or "csv").strip().lower()
    if normalized_kind not in {"purchases", "sales", "journal"}:
        return jsonify({"ok": False, "errors": ["Tipo de exportación no soportado."]}), 400
    if export_format not in {"csv", "xlsx"}:
        return jsonify({"ok": False, "errors": ["Formato no soportado."]}), 400

    try:
        start_date, end_date, _ = build_accounting_export_range()
    except ValueError as exc:
        return jsonify({"ok": False, "errors": [str(exc)]}), 400

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.id,
            companies_table.c.display_name,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(company_query).mappings().first()
        if not company:
            return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
        source_data = load_accounting_export_source_data(
            conn,
            data_owner_id,
            company_id,
            start_date,
            end_date,
        )

    if normalized_kind == "purchases":
        rows = build_purchase_export_rows(source_data)
        columns = [
            "fecha",
            "documento_tipo",
            "origen_tipo",
            "origen_id",
            "contraparte",
            "concepto",
            "base",
            "iva",
            "retencion",
            "total",
            "iva_deducible",
            "cuenta_sugerida",
            "familia",
            "subtipo",
            "bucket_pyg",
            "modelos_fiscales",
        ]
        sheet_name = "Compras"
        prefix = "compras"
    elif normalized_kind == "sales":
        rows = build_sales_export_rows(source_data)
        columns = [
            "fecha",
            "documento_tipo",
            "origen_tipo",
            "origen_id",
            "cliente",
            "concepto",
            "base",
            "iva",
            "total",
            "tipo_iva",
            "vencimiento",
            "estado_pago",
        ]
        sheet_name = "Ventas"
        prefix = "ventas"
    else:
        rows = build_journal_export_rows(source_data)
        columns = [
            "asiento_id",
            "linea",
            "fecha",
            "diario",
            "concepto",
            "cuenta",
            "descripcion_cuenta",
            "debe",
            "haber",
            "tercero",
            "documento_origen",
            "origen_tipo",
            "origen_id",
        ]
        sheet_name = "Asientos"
        prefix = "asientos"

    filename = build_export_row_filename(
        prefix,
        company.get("display_name") or "empresa",
        "xlsx" if export_format == "xlsx" else "csv",
        start_date,
        end_date,
    )
    return export_response_from_rows(
        rows,
        columns,
        export_format=export_format,
        sheet_name=sheet_name,
        download_name=filename,
    )


@app.route("/api/accounting-integrations/export-package")
def accounting_integrations_export_package():
    data_owner_id = get_data_owner_id()
    user_role = (g.current_user or {}).get("role")
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400
    try:
        start_date, end_date, period_label = build_accounting_export_range()
    except ValueError as exc:
        return jsonify({"ok": False, "errors": [str(exc)]}), 400

    with engine.connect() as conn:
        company_query = select(
            companies_table.c.id,
            companies_table.c.display_name,
            companies_table.c.legal_name,
            companies_table.c.tax_id,
        ).where(companies_table.c.id == company_id)
        if user_role != "owner":
            company_query = company_query.where(companies_table.c.agency_id == data_owner_id)
        company = conn.execute(company_query).mappings().first()
        if not company:
            return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
        source_data = load_accounting_export_source_data(
            conn,
            data_owner_id,
            company_id,
            start_date,
            end_date,
        )

    package_stream = build_accounting_export_package(
        source_data,
        {
            "company_id": company["id"],
            "company_name": company.get("display_name"),
            "company_legal_name": company.get("legal_name"),
            "company_tax_id": company.get("tax_id"),
            "period_label": period_label,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "generated_at": datetime.utcnow().isoformat(),
        },
    )
    filename = build_export_row_filename(
        "paquete_contable",
        company.get("display_name") or "empresa",
        "zip",
        start_date,
        end_date,
    )
    return send_file(
        package_stream,
        mimetype="application/zip",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/financial-metrics")
def financial_metrics():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    period = (request.args.get("period") or "monthly").strip().lower()
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year
    if period not in {"monthly", "quarterly"}:
        period = "monthly"

    selected_months = _quarter_months_for_month(month) if period == "quarterly" else [month]
    selected_periods = _periods_for_months(year, selected_months)
    ytd_periods = _periods_for_months(year, range(1, month + 1))
    full_year_periods = _periods_for_months(year, range(1, 13))

    with engine.connect() as conn:
        company = _fetch_company_fiscal_profile(conn, company_id)
        if not company:
            return jsonify({"ok": False, "errors": ["Empresa no encontrada."]}), 404
        selected_metrics = _build_financial_metrics(data_owner_id, company_id, selected_periods, conn)
        ytd_metrics = _build_financial_metrics(data_owner_id, company_id, ytd_periods, conn)
        full_year_metrics = _build_financial_metrics(data_owner_id, company_id, full_year_periods, conn)

    return jsonify(
        {
            "ok": True,
            "companyType": company.get("company_type"),
            "period": period,
            "selectedMonths": selected_months,
            "selected": selected_metrics,
            "yearToDate": ytd_metrics,
            "fullYear": full_year_metrics,
        }
    )


@app.route("/api/summary")
def summary():
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    data_owner_id = get_data_owner_id()
    company_id = get_company_id(required=True)
    if company_id is None:
        return jsonify({"ok": False, "errors": ["Empresa no seleccionada."]}), 400

    today = date.today()
    month = month or today.month
    year = year or today.year

    _, last_day = calendar.monthrange(year, month)
    month_prefix = f"{year}-{month:02d}"

    with engine.connect() as conn:
        metrics = _build_financial_metrics(data_owner_id, company_id, [(year, month)], conn)
        rows = conn.execute(
            select(
                invoices_table.c.expense_category,
                invoices_table.c.vat_deductible,
                invoices_table.c.vat_rate,
                invoices_table.c.base_amount,
                invoices_table.c.vat_amount,
                invoices_table.c.vat_breakdown,
            )
            .where(invoices_table.c.user_id == data_owner_id)
            .where(invoices_table.c.company_id == company_id)
            .where(invoices_table.c.invoice_date.like(f"{month_prefix}%"))
        ).mappings().all()
        no_invoice_rows = conn.execute(
            select(
                no_invoice_table.c.vat_deductible,
                no_invoice_table.c.vat_rate,
                no_invoice_table.c.vat_amount,
            )
            .where(no_invoice_table.c.user_id == data_owner_id)
            .where(no_invoice_table.c.company_id == company_id)
            .where(no_invoice_table.c.expense_date.like(f"{month_prefix}%"))
        ).mappings().all()

    daily_totals = {day: 0.0 for day in range(1, last_day + 1)}
    vat_totals = {0: 0.0, 4: 0.0, 10: 0.0, 21: 0.0}

    for row in rows:
        if row.get("expense_category") != "non_deductible" and row.get("vat_deductible") is not False:
            breakdown = parse_vat_breakdown(row.get("vat_breakdown"))
            if breakdown:
                for line in breakdown:
                    rate = int(line.get("rate") or 0)
                    vat_value = float(line.get("vat_amount") or 0)
                    if rate in vat_totals:
                        vat_totals[rate] += vat_value
            else:
                base_amount = float(row.get("base_amount") or 0)
                vat_rate = int(row["vat_rate"])
                vat_totals[vat_rate] += float(row.get("vat_amount") or base_amount * (vat_rate / 100))

    for row in no_invoice_rows:
        if not row.get("vat_deductible"):
            continue
        rate = int(row.get("vat_rate") or 0)
        vat_value = float(row.get("vat_amount") or 0)
        if rate in vat_totals:
            vat_totals[rate] += vat_value

    for day, amount in metrics["daily_expense_totals"].items():
        day_number = int(day)
        if 1 <= day_number <= last_day:
            daily_totals[day_number] = round(float(amount or 0), 2)

    cumulative = []
    running = 0.0
    for day in range(1, last_day + 1):
        running += daily_totals[day]
        cumulative.append(round(running, 2))

    suppliers = list(metrics["supplier_totals"].keys())
    supplier_values = [round(metrics["supplier_totals"][name], 2) for name in suppliers]

    vat_total_deductible = round(sum(vat_totals.values()), 2)

    return jsonify(
        {
            "days": list(range(1, last_day + 1)),
            "cumulative": cumulative,
            "suppliers": suppliers,
            "supplierTotals": supplier_values,
            "totalSpent": metrics["expense_gross_total"],
            "vatTotals": {
                "0": round(vat_totals[0], 2),
                "4": round(vat_totals[4], 2),
                "10": round(vat_totals[10], 2),
                "21": round(vat_totals[21], 2),
            },
            "vatTotalDeductible": vat_total_deductible,
        }
    )


init_db()

if __name__ == "__main__":
    app.run()
